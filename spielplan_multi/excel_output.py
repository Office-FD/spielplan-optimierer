"""Excel-Export pro Liga (Format identisch zu v6) + Co-Home-Zusammenfassung."""

from __future__ import annotations

import os
import re
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import numpy as np
from openpyxl import Workbook
from ortools.sat.python import cp_model as _cp_model
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

from .league_types import LeagueConfig, LeagueResult
from .config import TEAM_COLORS, WEIGHT_LABELS, KM_PAUSCHALE, get_team_color
from .ui import step, ok, warn


# ── Style-Helfer ─────────────────────────────────────────────────────────────

def _fill(color): return PatternFill(start_color=color, end_color=color, fill_type='solid')

HDR_FILL   = _fill('4472C4')
TITLE_FILL = _fill('1F3864')
SEC_FILL   = _fill('D9E1F2')
GREEN_FILL = _fill('C6EFCE')
RED_FILL   = _fill('FFC7CE')
GRAY_FILL  = _fill('E0E0E0')
AWAY_FILL  = _fill('FFE0E0')
DST_FILL   = _fill('FFF2CC')


# ── Eine Liga als Workbook ────────────────────────────────────────────────────

def build_league_excel(result: LeagueResult) -> Workbook:
    cfg = result.cfg
    teams    = cfg.teams
    n        = cfg.n_teams
    days     = cfg.days
    dst_days = cfg.dst_days
    dist     = cfg.dist
    t_idx    = {t: i for i, t in enumerate(teams)}

    date_str = datetime.now().strftime('%Y-%m-%d')
    total_km = int(sum(result.travels))

    wb = Workbook()

    # ── Sheet 1: Konfiguration ───────────────────────────────────────────────
    ws = wb.active
    ws.title = 'Konfiguration'
    ws.column_dimensions['A'].width = 34
    ws.column_dimensions['B'].width = 56

    def kv(row, k, v):
        ws.cell(row, 1, k); ws.cell(row, 2, v)

    def sec(row, text):
        ws.merge_cells(f'A{row}:B{row}')
        c = ws.cell(row, 1, text)
        c.fill = SEC_FILL; c.font = Font(bold=True)

    ws.merge_cells('A1:B1')
    c = ws.cell(1, 1, f'SPIELPLAN – {cfg.name}')
    c.fill = TITLE_FILL; c.font = Font(bold=True, color='FFFFFF', size=13)
    c.alignment = Alignment(horizontal='center')

    r = 3
    sec(r, 'ALLGEMEINE INFORMATIONEN'); r += 1
    kv(r, 'Erstellungsdatum:', date_str); r += 1
    kv(r, 'Liga:', cfg.name); r += 1
    kv(r, 'Anzahl Teams:', n); r += 1
    kv(r, 'Spieltage:', cfg.n_matchdays); r += 1
    kv(r, 'Runden:', cfg.n_rounds); r += 1
    if cfg.games_per_team_per_day > 1:
        K = cfg.n_teams_per_group
        G = cfg.n_groups_per_day
        kv(r, 'Format:', 'Turniertag'); r += 1
        kv(r, 'Spiele/Team/Tag:', cfg.games_per_team_per_day); r += 1
        n_act = cfg.n_active_per_day
        if n_act > 0 and n_act < n:
            kv(r, 'Teams/Spieltag:', f'{n_act} von {n} ({n - n_act} Spielfrei)'); r += 1
        elif K > 0:
            kv(r, 'Stufe:', f'2 ({G} Gruppen à {K} Teams)'); r += 1
        else:
            kv(r, 'Stufe:', '1 (alle Teams an einem Ort)'); r += 1
        # Spielreihenfolge-Einstellungen
        tt = cfg.tt_settings
        if tt:
            kv(r, 'Spielreihenfolge:', 'Ja'); r += 1
            _hs = tt.get('host_slots', [])
            if not _hs and tt.get('host_position'):
                _hs = ['2', 'N-1']
            kv(r, '  Ausrichterslots:', str(_hs) if _hs else 'Keine'); r += 1
            kv(r, '  Mindestpause (Spiele):', tt.get('min_gap', 1)); r += 1
            kv(r, '  Maximalpause (Spiele):', tt.get('max_gap', 3)); r += 1
            mode = tt.get('host_mode', 'per_team')
            if mode == 'per_day':
                hdp = tt.get('host_per_day', {})
                kv(r, '  Ausrichter pro Spieltag:',
                   ', '.join(f'ST{d}:{h}' for d, h in sorted(hdp.items()))); r += 1
            else:
                hc = tt.get('host_counts', {})
                kv(r, '  Ausrichter-Anzahl:',
                   ', '.join(f'{t}:{n}' for t, n in hc.items() if n > 0)); r += 1
    kv(r, 'Solver-Status:', _cp_model.CpSolver().StatusName(result.status)); r += 1
    kv(r, 'Laufzeit:', f'{result.mins:02d}:{result.secs:02d} (mm:ss)'); r += 1
    kv(r, 'Objective Value:', round(result.objective, 2)); r += 2

    sec(r, 'TEAMS & ORTE'); r += 1
    for i, (t, loc) in enumerate(zip(teams, cfg.locations), 1):
        kv(r, f'{i}. {t}', loc); r += 1
    r += 1

    sec(r, 'DOPPELSPIELTAGE'); r += 1
    kv(r, 'Anzahl:', len(cfg.dst_blocks)); r += 1
    if cfg.dst_blocks:
        kv(r, 'Paare:', ', '.join(f'ST{a}+{b}' for a, b in cfg.dst_blocks)); r += 1
    r += 1

    sec(r, 'DST ROUTING'); r += 1
    kv(r, 'Aktiv:', 'Ja' if cfg.apply_routing else 'Nein'); r += 1
    if cfg.apply_routing:
        kv(r, 'Max:', f'{cfg.f_num}% der Direktstrecke'); r += 1
    r += 1

    sec(r, 'GEWICHTUNGEN (0-10)'); r += 1
    for key, label in WEIGHT_LABELS:
        kv(r, label + ':', cfg.raw_weights.get(key, '-')); r += 1
    kv(r, 'Ligahierarchie:', cfg.hier_weight); r += 1
    r += 1

    sec(r, 'PFLICHTSPIELE'); r += 1
    if cfg.pinned:
        for pm in cfg.pinned:
            kv(r, f"ST{pm['day']}: {pm['teamA']} vs. {pm['teamB']}",
               f"Heim: {pm['home']}" if pm['home'] else 'Heim: beliebig'); r += 1
    else:
        kv(r, 'Keine', ''); r += 1
    r += 1

    sec(r, 'HEIMSPIEL-SPERRTAGE'); r += 1
    if cfg.blocked:
        for t, bdays in cfg.blocked.items():
            kv(r, t + ':', ', '.join(f'ST{d}' for d in bdays)); r += 1
    else:
        kv(r, 'Keine', ''); r += 1
    r += 1

    sec(r, 'HEIMSPIEL-PFLICHTTAGE'); r += 1
    forced_home = getattr(cfg, 'forced_home', {})
    if forced_home:
        for t, fdays in forced_home.items():
            kv(r, t + ':', ', '.join(f'ST{d}' for d in fdays)); r += 1
    else:
        kv(r, 'Keine', ''); r += 1
    r += 1

    sec(r, 'STATISTIKEN'); r += 1
    kv(r, 'Gesamt-km:', f'{total_km} km'); r += 1
    _tr = result.travels or []
    _sr = result.sw_rates or []
    kv(r, 'O km/Team:', f'{np.mean(_tr):.1f} km' if _tr else '–'); r += 1
    kv(r, 'Std. km:', f'{np.std(_tr):.1f} km' if _tr else '–'); r += 1
    kv(r, 'Min. km:', f'{min(_tr)} km' if _tr else '–'); r += 1
    kv(r, 'Max. km:', f'{max(_tr)} km' if _tr else '–'); r += 1
    kv(r, 'O Wechselquote:', f'{np.mean(_sr):.1f}%' if _sr else '–'); r += 1
    kv(r, 'km-Pauschale:', f'{KM_PAUSCHALE:.2f} EUR/km'); r += 1

    # ── Sheet 2: Spielplan ───────────────────────────────────────────────────
    ws_sp = wb.create_sheet('Spielplan')
    K = cfg.n_teams_per_group
    has_groups = K > 0 and bool(result.groups)
    is_tt      = cfg.games_per_team_per_day > 1
    has_times  = bool(result.game_times)
    typ_col    = 'Ausrichter' if is_tt else 'Typ'
    if has_groups:
        hdr_cols   = ['Spieltag', 'Phase', typ_col, 'Gruppe', 'Ort', 'Nr']
        col_widths = [10, 12, 22 if is_tt else 8, 8, 22, 5]
    else:
        hdr_cols   = ['Spieltag', 'Phase', typ_col, 'Nr']
        col_widths = [10, 12, 22 if is_tt else 8, 5]
    if has_times:
        hdr_cols.append('Uhrzeit'); col_widths.append(9)
    hdr_cols  += ['Heimteam', 'Gastteam']
    col_widths += [24, 24]

    for col, txt in enumerate(hdr_cols, 1):
        c = ws_sp.cell(1, col, txt)
        c.fill = HDR_FILL; c.font = Font(bold=True, color='FFFFFF')
        c.alignment = Alignment(horizontal='center', vertical='center')

    round_len  = cfg.n_matchdays // max(1, cfg.n_rounds)
    _phase_lbl = {1: 'Hinrunde', 2: 'Rueckrunde', 3: 'Dritte Runde'}
    # dynamische Spalten-Offsets
    _uhr_col  = hdr_cols.index('Uhrzeit') + 1 if has_times else None
    _heim_col = hdr_cols.index('Heimteam') + 1
    _gast_col = hdr_cols.index('Gastteam') + 1
    r = 2
    for d in days:
        typ   = result.hosts.get(d, '') if is_tt else ('DST' if d in dst_days else 'EST')
        rnd   = min(cfg.n_rounds, (d - 1) // round_len + 1)
        phase = _phase_lbl.get(rnd, f'Runde {rnd}')
        day_times = result.game_times.get(d, [])
        team_to_group: Dict[str, int] = {}
        grp_host_loc: Dict[int, str] = {}   # g_idx (1-based) -> Spielort
        if has_groups:
            for g_idx, grp in enumerate(result.groups.get(d, []), 1):
                for tm in grp:
                    team_to_group[tm] = g_idx
                # Spielort = Location des Heimteams in dieser Gruppe
                _host = next(
                    (tm for tm in grp
                     if result.home_vals.get((t_idx.get(tm, -1), d), 0) >= 1),
                    grp[0] if grp else '',
                )
                _hi = t_idx.get(_host, -1)
                grp_host_loc[g_idx] = (cfg.locations[_hi]
                                       if 0 <= _hi < len(cfg.locations) else _host)
        _ort_col = hdr_cols.index('Ort') + 1 if 'Ort' in hdr_cols else None
        for nr, (ht, at) in enumerate(result.schedule.get(d, []), 1):
            hi = t_idx.get(ht, -1); ai = t_idx.get(at, -1)
            ws_sp.cell(r, 1, d); ws_sp.cell(r, 2, phase); ws_sp.cell(r, 3, typ)
            if has_groups:
                _g = team_to_group.get(ht, '')
                ws_sp.cell(r, 4, _g)
                if _ort_col:
                    ws_sp.cell(r, _ort_col, grp_host_loc.get(_g, ''))
                ws_sp.cell(r, hdr_cols.index('Nr') + 1, nr)
            else:
                ws_sp.cell(r, 4, nr)
            if _uhr_col:
                t_val = day_times[nr - 1] if nr - 1 < len(day_times) else ''
                ws_sp.cell(r, _uhr_col, t_val).alignment = Alignment(horizontal='center')
            ch = ws_sp.cell(r, _heim_col, ht)
            ca = ws_sp.cell(r, _gast_col, at)
            ch.fill = _fill(get_team_color(hi))
            ca.fill = _fill(get_team_color(ai))
            r += 1
    for col, w in zip(range(1, len(hdr_cols) + 1), col_widths):
        ws_sp.column_dimensions[get_column_letter(col)].width = w

    # ── Sheet 3: Gruppen-Übersicht (nur Stufe 2) ────────────────────────────
    if has_groups:
        G = cfg.n_groups_per_day
        t_idx = {t: i for i, t in enumerate(teams)}
        ws_gr = wb.create_sheet('Gruppen-Uebersicht')
        # Header: Spieltag + je eine Spalte pro Gruppe
        ws_gr.cell(1, 1, 'Spieltag').fill = HDR_FILL
        ws_gr.cell(1, 1).font = Font(bold=True, color='FFFFFF')
        ws_gr.column_dimensions['A'].width = 10
        for g in range(1, G + 1):
            c = ws_gr.cell(1, g + 1, f'Gruppe {g}')
            c.fill = HDR_FILL; c.font = Font(bold=True, color='FFFFFF')
            c.alignment = Alignment(horizontal='center')
            ws_gr.column_dimensions[get_column_letter(g + 1)].width = max(20, K * 10)
        grp_fills = [_fill('D9E1F2'), _fill('E2EFDA'), _fill('FFF2CC'),
                     _fill('FCE4D6'), _fill('DDEBF7'), _fill('F2CEEF')]
        r = 2
        for d in days:
            ws_gr.cell(r, 1, d)
            for g_idx, grp in enumerate(result.groups.get(d, []), 1):
                # Spielort: das Heimteam der Gruppe (home_vals == 1)
                host = next(
                    (t for t in grp
                     if result.home_vals.get((t_idx.get(t, -1), d), 0) >= 1
                     and t_idx.get(t, -1) >= 0),
                    grp[0],
                )
                _hi = t_idx.get(host, -1)
                host_loc = cfg.locations[_hi] if 0 <= _hi < len(cfg.locations) else host
                cell_val = ' · '.join(grp) + f'\n@ {host_loc}'
                c = ws_gr.cell(r, g_idx + 1, cell_val)
                c.fill = grp_fills[(g_idx - 1) % len(grp_fills)]
                c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws_gr.row_dimensions[r].height = 32
            r += 1

    # ── Sheet: Heatmap Heimrecht ─────────────────────────────────────────────
    ws_hm = wb.create_sheet('Heatmap Heimrecht')
    ws_hm.column_dimensions['A'].width = 24
    ws_hm.cell(1, 1, 'Team').fill = HDR_FILL
    ws_hm.cell(1, 1).font = Font(bold=True, color='FFFFFF')

    day_to_col = {d: i + 2 for i, d in enumerate(days)}
    for d in days:
        col = day_to_col[d]
        c = ws_hm.cell(1, col, d)
        c.fill = HDR_FILL; c.font = Font(bold=True, color='FFFFFF')
        c.alignment = Alignment(horizontal='center')
        ws_hm.column_dimensions[get_column_letter(col)].width = 4

    for ti, t in enumerate(teams):
        ws_hm.cell(ti + 2, 1, t).font = Font(bold=True)
        for d in days:
            col = day_to_col[d]
            is_home = result.home_vals.get((ti, d), 0) >= 1
            c = ws_hm.cell(ti + 2, col, 'H' if is_home else 'A')
            c.fill = GREEN_FILL if is_home else RED_FILL
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal='center', vertical='center')

    # ── Sheet: Kilometerstatistik ────────────────────────────────────────────
    ws_km = wb.create_sheet('Kilometerstatistik')
    for col, txt in enumerate(['Team', 'Kilometer', 'Switches', 'Wechselquote %'], 1):
        c = ws_km.cell(1, col, txt)
        c.fill = HDR_FILL; c.font = Font(bold=True, color='FFFFFF')
        c.alignment = Alignment(horizontal='center')
    r = 2
    for ti, t in enumerate(teams):
        ws_km.cell(r, 1, t)
        ws_km.cell(r, 2, result.travels[ti])
        ws_km.cell(r, 3, result.sw_counts[ti])
        ws_km.cell(r, 4, round(result.sw_rates[ti], 1))
        r += 1
    r += 1
    ws_km.cell(r, 1, 'GESAMT').font = Font(bold=True)
    ws_km.cell(r, 2, total_km).font = Font(bold=True)
    r += 1
    ws_km.cell(r, 1, 'Durchschnitt').font = Font(bold=True)
    ws_km.cell(r, 2, round(np.mean(result.travels), 1)).font = Font(bold=True)
    ws_km.cell(r, 3, round(np.mean(result.sw_counts), 1)).font = Font(bold=True)
    ws_km.cell(r, 4, round(np.mean(result.sw_rates), 1)).font = Font(bold=True)
    for col, w in zip(range(1, 5), [24, 12, 10, 16]):
        ws_km.column_dimensions[get_column_letter(col)].width = w

    # ── Sheet: Distanzmatrix ─────────────────────────────────────────────────
    ws_dm = wb.create_sheet('Distanzmatrix')
    ws_dm.column_dimensions['A'].width = 22
    ws_dm.cell(1, 1, 'von \\ nach').fill = HDR_FILL
    ws_dm.cell(1, 1).font = Font(bold=True, color='FFFFFF')
    ws_dm.cell(1, 1).alignment = Alignment(horizontal='center')
    for j, t in enumerate(teams):
        c = ws_dm.cell(1, j + 2, t)
        c.fill = HDR_FILL; c.font = Font(bold=True, color='FFFFFF')
        c.alignment = Alignment(horizontal='center')
        ws_dm.column_dimensions[get_column_letter(j + 2)].width = 14
    for i, t in enumerate(teams):
        ws_dm.cell(i + 2, 1, t).font = Font(bold=True)
        for j in range(n):
            c = ws_dm.cell(i + 2, j + 2, int(dist[i, j]))
            c.alignment = Alignment(horizontal='center')
            if i == j: c.fill = GRAY_FILL

    # ── Sheet: Fahrtkostenausgleich ──────────────────────────────────────────
    ws_fk = wb.create_sheet('Fahrtkostenausgleich')
    for col, txt in enumerate(['Team', 'Kilometer', 'O Kilometer', 'Abweichung (km)', 'Ausgleich (EUR)'], 1):
        c = ws_fk.cell(1, col, txt)
        c.fill = HDR_FILL; c.font = Font(bold=True, color='FFFFFF')
        c.alignment = Alignment(horizontal='center')

    avg_km     = float(np.mean(result.travels))
    r          = 2
    total_comp = 0.0
    for ti, t in enumerate(teams):
        abw = result.travels[ti] - avg_km
        aus = abw * KM_PAUSCHALE
        total_comp += aus
        ws_fk.cell(r, 1, t)
        ws_fk.cell(r, 2, result.travels[ti]).alignment = Alignment(horizontal='center')
        ws_fk.cell(r, 3, round(avg_km, 1)).alignment = Alignment(horizontal='center')
        ws_fk.cell(r, 4, round(abw, 1)).alignment = Alignment(horizontal='center')
        ws_fk.cell(r, 5, f'{aus:.2f} EUR').alignment = Alignment(horizontal='right')
        if abw > 0:
            fill = GREEN_FILL; color = '006100'
        elif abw < 0:
            fill = RED_FILL;   color = '9C0006'
        else:
            fill = None;       color = '000000'
        if fill:
            ws_fk.cell(r, 4).fill = fill; ws_fk.cell(r, 4).font = Font(color=color)
            ws_fk.cell(r, 5).fill = fill; ws_fk.cell(r, 5).font = Font(color=color, bold=True)
        r += 1
    r += 1
    ws_fk.cell(r, 4, 'Summe:').font = Font(bold=True)
    ws_fk.cell(r, 4).alignment = Alignment(horizontal='right')
    ws_fk.cell(r, 5, f'{total_comp:.2f} EUR').font = Font(bold=True)
    r += 2
    ws_fk.cell(r, 1, f'km-Pauschale: {KM_PAUSCHALE:.2f} EUR/km').font = Font(italic=True, color='666666')
    for col, w in zip(range(1, 6), [24, 12, 12, 16, 18]):
        ws_fk.column_dimensions[get_column_letter(col)].width = w

    # ── Sheet: Fairness-Analyse ──────────────────────────────────────────────
    ws_fa = wb.create_sheet('Fairness-Analyse')
    ws_fa.column_dimensions['A'].width = 26

    OK_FILL  = _fill('FFEB9C')   # gelb = ausreichend
    # Bewertungs-Helfer
    def _bew_km(pct):
        p = abs(pct)
        if p <= 10: return 'gut',       GREEN_FILL, '006100'
        if p <= 25: return 'ok',        OK_FILL,    '9C6500'
        return             'ungünstig', RED_FILL,   '9C0006'

    def _bew_sw(rate):
        if rate >= 70: return 'gut',       GREEN_FILL, '006100'
        if rate >= 50: return 'ok',        OK_FILL,    '9C6500'
        return              'ungünstig', RED_FILL,   '9C0006'

    def _bew_cons(max_c):
        if max_c <= 2: return 'gut',       GREEN_FILL, '006100'
        if max_c <= 3: return 'ok',        OK_FILL,    '9C6500'
        return              'ungünstig', RED_FILL,   '9C0006'

    def _bew_home_bal(home_cnt, away_cnt):
        total = home_cnt + away_cnt
        if total == 0: return 'gut', GREEN_FILL, '006100'
        pct = home_cnt / total * 100
        if 40 <= pct <= 60: return 'gut',       GREEN_FILL, '006100'
        if 30 <= pct <= 70: return 'ok',        OK_FILL,    '9C6500'
        return                     'ungünstig', RED_FILL,   '9C0006'

    def _write_cell(ws, r, c, val, bold=False, center=True, fill=None, color='000000'):
        cell = ws.cell(r, c, val)
        cell.font = Font(bold=bold, color=color)
        cell.alignment = Alignment(horizontal='center' if center else 'left',
                                   vertical='center')
        if fill: cell.fill = fill
        return cell

    def _section_hdr(ws, r, txt, ncols):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
        c = ws.cell(r, 1, txt)
        c.fill = SEC_FILL; c.font = Font(bold=True); c.alignment = Alignment(horizontal='left')

    # Titel
    ws_fa.merge_cells('A1:H1')
    tc = ws_fa.cell(1, 1, f'FAIRNESS-ANALYSE – {cfg.name}')
    tc.fill = TITLE_FILL
    tc.font = Font(bold=True, color='FFFFFF', size=13)
    tc.alignment = Alignment(horizontal='center', vertical='center')
    ws_fa.row_dimensions[1].height = 22

    # ── A: Reise-Fairness ───────────────────────────────────────────────────
    avg_km_fa = float(np.mean(result.travels)) if result.travels else 0.0
    r = 3
    _section_hdr(ws_fa, r, 'A  REISE-FAIRNESS', 6); r += 1
    for col, txt in enumerate(['Team', 'Kilometer', 'Abw. (km)', 'Abw. (%)', 'Heim-km', 'Bewertung'], 1):
        c = ws_fa.cell(r, col, txt)
        c.fill = HDR_FILL; c.font = Font(bold=True, color='FFFFFF')
        c.alignment = Alignment(horizontal='center')
    r += 1

    home_km_per_team = [0] * n  # km the team earns as host (opponents travel)
    for d, games in result.schedule.items():
        for ht, at in games:
            hi = t_idx.get(ht, -1); ai = t_idx.get(at, -1)
            if 0 <= hi < n and 0 <= ai < n:
                home_km_per_team[hi] += int(dist[ai, hi])

    for ti, t in enumerate(teams):
        km   = result.travels[ti]
        abw  = km - avg_km_fa
        pct  = abw / avg_km_fa * 100 if avg_km_fa > 0 else 0.0
        lbl, fill, color = _bew_km(pct)
        _write_cell(ws_fa, r, 1, t, bold=True, center=False)
        _write_cell(ws_fa, r, 2, km)
        c3 = ws_fa.cell(r, 3, round(abw, 0))
        c3.alignment = Alignment(horizontal='center')
        if abw < 0:   c3.fill = GREEN_FILL
        elif abw > 0: c3.fill = RED_FILL
        _write_cell(ws_fa, r, 4, f'{pct:+.1f}%')
        _write_cell(ws_fa, r, 5, home_km_per_team[ti])
        _write_cell(ws_fa, r, 6, lbl, bold=True, fill=fill, color=color)
        r += 1
    # Ø-Zeile
    ws_fa.cell(r, 1, 'Durchschnitt').font = Font(italic=True)
    ws_fa.cell(r, 2, round(avg_km_fa, 0)).font = Font(italic=True)
    ws_fa.cell(r, 3, 0).font = Font(italic=True)
    ws_fa.cell(r, 4, '0.0%').font = Font(italic=True)
    r += 2

    # ── B: Heimrecht pro Phase ───────────────────────────────────────────────
    round_len_fa = cfg.n_matchdays // max(1, cfg.n_rounds)
    _ph_lbl = {1: 'Hinrunde', 2: 'Rückrunde', 3: 'Dritte Runde'}
    # phase_home[ti][rnd] – 1-basiert
    phase_home = [[0] * (cfg.n_rounds + 1) for _ in range(n)]
    phase_away = [[0] * (cfg.n_rounds + 1) for _ in range(n)]
    for d, games in result.schedule.items():
        rnd = min(cfg.n_rounds, (d - 1) // round_len_fa + 1)
        for ht, at in games:
            hi = t_idx.get(ht, -1); ai = t_idx.get(at, -1)
            if hi >= 0: phase_home[hi][rnd] += 1
            if ai >= 0: phase_away[ai][rnd]  += 1

    phase_cols = []
    for rnd in range(1, cfg.n_rounds + 1):
        phase_cols += [f'{_ph_lbl.get(rnd, f"R{rnd}")}\nHeim',
                       f'{_ph_lbl.get(rnd, f"R{rnd}")}\nAusw.']
    n_bcols = 4 + 2 * cfg.n_rounds
    _section_hdr(ws_fa, r, 'B  HEIMRECHT PRO PHASE', n_bcols); r += 1
    for col, txt in enumerate(['Team', 'Heim ges.', 'Ausw. ges.'] + phase_cols + ['Bewertung'], 1):
        c = ws_fa.cell(r, col, txt)
        c.fill = HDR_FILL; c.font = Font(bold=True, color='FFFFFF')
        c.alignment = Alignment(horizontal='center', wrap_text=True)
    ws_fa.row_dimensions[r].height = 28
    r += 1

    for ti, t in enumerate(teams):
        total_h = sum(phase_home[ti][1:])
        total_a = sum(phase_away[ti][1:])
        lbl, fill, color = _bew_home_bal(total_h, total_a)
        _write_cell(ws_fa, r, 1, t, bold=True, center=False)
        _write_cell(ws_fa, r, 2, total_h)
        _write_cell(ws_fa, r, 3, total_a)
        for rnd in range(1, cfg.n_rounds + 1):
            base_col = 4 + (rnd - 1) * 2
            _write_cell(ws_fa, r, base_col,     phase_home[ti][rnd], fill=GREEN_FILL)
            _write_cell(ws_fa, r, base_col + 1, phase_away[ti][rnd], fill=AWAY_FILL)
        _write_cell(ws_fa, r, 4 + 2 * cfg.n_rounds, lbl, bold=True, fill=fill, color=color)
        r += 1
    r += 1

    # ── C: Abwechslung & Konsekutive Sequenzen ──────────────────────────────
    max_h_streak = [0] * n
    max_a_streak = [0] * n
    cur_h_streak = [0] * n
    cur_a_streak = [0] * n
    for d in sorted(days):
        for ti in range(n):
            hv = result.home_vals.get((ti, d))
            if hv is None:
                cur_h_streak[ti] = 0; cur_a_streak[ti] = 0
                continue
            if hv >= 1:
                cur_h_streak[ti] += 1; cur_a_streak[ti] = 0
                max_h_streak[ti] = max(max_h_streak[ti], cur_h_streak[ti])
            else:
                cur_a_streak[ti] += 1; cur_h_streak[ti] = 0
                max_a_streak[ti] = max(max_a_streak[ti], cur_a_streak[ti])

    _section_hdr(ws_fa, r, 'C  ABWECHSLUNG & KONSEKUTIVE SEQUENZEN', 7); r += 1
    for col, txt in enumerate(
            ['Team', 'Wechsel', 'Quote (%)',
             'Max. kons. Heim', 'Max. kons. Ausw.', 'Bewertung'], 1):
        c = ws_fa.cell(r, col, txt)
        c.fill = HDR_FILL; c.font = Font(bold=True, color='FFFFFF')
        c.alignment = Alignment(horizontal='center', wrap_text=True)
    ws_fa.row_dimensions[r].height = 28
    r += 1

    for ti, t in enumerate(teams):
        sw   = result.sw_counts[ti]
        rate = result.sw_rates[ti]
        mh   = max_h_streak[ti]
        ma   = max_a_streak[ti]
        lbl_sw,   fill_sw,   color_sw   = _bew_sw(rate)
        lbl_mh,   fill_mh,   _          = _bew_cons(mh)
        lbl_ma,   fill_ma,   _          = _bew_cons(ma)
        overall_lbl = lbl_sw if lbl_sw != 'gut' else (lbl_mh if lbl_mh != 'gut' else lbl_ma)
        _, overall_fill, overall_color = _bew_sw(rate) if lbl_sw != 'gut' else \
                                         (_bew_cons(mh) if lbl_mh != 'gut' else _bew_cons(ma))
        _write_cell(ws_fa, r, 1, t, bold=True, center=False)
        _write_cell(ws_fa, r, 2, sw)
        _write_cell(ws_fa, r, 3, f'{rate:.1f}%', fill=fill_sw)
        _write_cell(ws_fa, r, 4, mh, fill=fill_mh)
        _write_cell(ws_fa, r, 5, ma, fill=fill_ma)
        _write_cell(ws_fa, r, 6, overall_lbl, bold=True, fill=overall_fill, color=overall_color)
        r += 1
    r += 1

    # ── Zusammenfassung ─────────────────────────────────────────────────────
    _section_hdr(ws_fa, r, 'ZUSAMMENFASSUNG', 4); r += 1
    km_scores  = [abs(result.travels[i] - avg_km_fa) / avg_km_fa * 100 if avg_km_fa > 0 else 0
                  for i in range(n)]
    km_fair_lbl, km_fair_fill, _ = _bew_km(max(km_scores) if km_scores else 0)
    sw_scores  = result.sw_rates
    sw_fair_lbl, sw_fair_fill, _ = _bew_sw(min(sw_scores) if sw_scores else 0)
    cons_max   = max(max_h_streak + max_a_streak) if n > 0 else 0
    cons_lbl, cons_fill, _  = _bew_cons(cons_max)
    for col, (label, val, fill) in enumerate([
        ('Reise-Fairness',  km_fair_lbl, km_fair_fill),
        ('Heimrecht-Abwechslung', sw_fair_lbl, sw_fair_fill),
        ('Konsekutive Spiele', cons_lbl, cons_fill),
    ], 1):
        ws_fa.cell(r, col, label).font = Font(italic=True, color='666666')
        ws_fa.cell(r, col).alignment = Alignment(horizontal='center')
        c = ws_fa.cell(r + 1, col, val)
        c.fill = fill; c.font = Font(bold=True)
        c.alignment = Alignment(horizontal='center')

    # Spaltenbreiten
    _fa_widths = [26, 12, 12, 14] + [10, 10] * cfg.n_rounds + [14]
    for col, w in zip(range(1, len(_fa_widths) + 1), _fa_widths):
        ws_fa.column_dimensions[get_column_letter(col)].width = w

    # ── Sheet: Team-Ansichten ────────────────────────────────────────────────
    ws_ta = wb.create_sheet('Team-Ansichten')
    is_tt_ta = cfg.games_per_team_per_day > 1

    # (day, team) → [(is_home, opponent)]
    tdg: Dict = {}
    for d in days:
        for ht, at in result.schedule.get(d, []):
            tdg.setdefault((d, ht), []).append((True,  at))
            tdg.setdefault((d, at), []).append((False, ht))

    # Header
    hdr_ta = ['ST', 'KW', 'Phase'] + list(teams)
    for col, txt in enumerate(hdr_ta, 1):
        c = ws_ta.cell(1, col, txt)
        c.fill = HDR_FILL
        c.font = Font(bold=True, color='FFFFFF')
        c.alignment = Alignment(horizontal='center', wrap_text=True)

    round_len_ta = cfg.n_matchdays // max(1, cfg.n_rounds)
    _phase_s = {1: 'Hin', 2: 'Rück', 3: '3.Rd'}

    for row_i, d in enumerate(days, 2):
        rnd   = min(cfg.n_rounds, (d - 1) // round_len_ta + 1)
        phase = _phase_s.get(rnd, f'R{rnd}')
        kw    = cfg.calendar.get(d, {}).get('kw', '')

        st_cell = ws_ta.cell(row_i, 1, d)
        st_cell.alignment = Alignment(horizontal='center')
        if d in dst_days:
            st_cell.fill = DST_FILL

        ws_ta.cell(row_i, 2, f'KW {kw}' if kw else '').alignment = Alignment(horizontal='center')
        ws_ta.cell(row_i, 3, phase).alignment = Alignment(horizontal='center')

        host_d = result.hosts.get(d, '') if is_tt_ta else ''

        for ti, t in enumerate(teams):
            col   = ti + 4
            games = tdg.get((d, t), [])

            if not games:
                c = ws_ta.cell(row_i, col, 'Spielfrei')
                c.fill = GRAY_FILL
            elif is_tt_ta:
                opps = ', '.join(opp for _, opp in games[:3])
                if t == host_d:
                    c = ws_ta.cell(row_i, col, f'Ausrichter\n{opps}')
                    c.fill = GREEN_FILL
                else:
                    c = ws_ta.cell(row_i, col, f'Gast bei: {host_d or "?"}\n{opps}')
                    c.fill = DST_FILL
            else:
                is_home, opp = games[0]
                oi = t_idx.get(opp, -1)
                if is_home:
                    c = ws_ta.cell(row_i, col, f'Heim\n{opp}')
                    c.fill = GREEN_FILL
                else:
                    km = int(dist[ti, oi]) if 0 <= oi < n else 0
                    km_s = f' · {km} km' if km else ''
                    c = ws_ta.cell(row_i, col, f'Auswärts{km_s}\n{opp}')
                    c.fill = AWAY_FILL

            c.font = Font(size=9)
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        ws_ta.row_dimensions[row_i].height = 32

    ws_ta.column_dimensions['A'].width = 5
    ws_ta.column_dimensions['B'].width = 7
    ws_ta.column_dimensions['C'].width = 6
    _ta_w = max(20, max(len(t) for t in teams) + 10)
    for i in range(n):
        ws_ta.column_dimensions[get_column_letter(4 + i)].width = _ta_w
    ws_ta.freeze_panes = 'D2'

    return wb


def save_league_excel(wb: Workbook, result: LeagueResult, output_dir: Path) -> str:
    cfg      = result.cfg
    date_str = datetime.now().strftime('%Y-%m-%d')
    total_km = int(sum(result.travels))

    if cfg.dst_blocks:
        cnt   = len(cfg.dst_blocks)
        pairs = ';'.join(f'{a},{b}' for a, b in cfg.dst_blocks)
        pct   = cfg.f_num - 100 if cfg.apply_routing else 0
        dst_p = f'{cnt}-DST{pct}%'
    else:
        dst_p = '0-DST'

    rw = cfg.raw_weights
    w  = f"W{int(rw.get('switch',5))}-FW{int(rw.get('sw_fair',5))}-FK{int(rw.get('trav_fair',5))}-GK{int(rw.get('travel',5))}"
    raw_name = f'{date_str}_{cfg.league_id}_{cfg.n_teams}Teams_{dst_p}_{w}_{total_km}km.xlsx'
    safe     = re.sub(r'[<>:"/\\|?*]', '_', raw_name)
    if len(safe) > 220:
        safe = safe[:216] + '.xlsx'

    output_dir.mkdir(parents=True, exist_ok=True)
    base, ext = os.path.splitext(safe)
    suffix = 0
    while True:
        target = output_dir / (safe if suffix == 0 else f'{base}_{suffix}{ext}')
        try:
            wb.save(target)
            return str(target)
        except PermissionError:
            suffix += 1
            if suffix > 20:
                raise


# ── Co-Home-Zusammenfassung (Liga-uebergreifend) ─────────────────────────────

def build_cohome_summary(results: Dict[str, Optional[LeagueResult]],
                          clubs: Dict[str, Dict[str, str]],
                          kw_compat: Dict[int, Dict[str, List[int]]]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = 'Co-Home Uebersicht'

    ws.merge_cells('A1:F1')
    c = ws.cell(1, 1, 'CO-HOME REALISIERUNG – MEHRSPARTEN-VEREINE')
    c.fill = TITLE_FILL; c.font = Font(bold=True, color='FFFFFF', size=12)
    c.alignment = Alignment(horizontal='center')

    headers = ['Verein', 'KW', 'Woche von', 'Woche bis', 'Ligen', 'Alle Heim?']
    for col, txt in enumerate(headers, 1):
        c = ws.cell(3, col, txt)
        c.fill = HDR_FILL; c.font = Font(bold=True, color='FFFFFF')
        c.alignment = Alignment(horizontal='center')

    r = 4
    for club_name, liga_team_map in clubs.items():
        for kw, kw_data in sorted(kw_compat.items()):
            entries = []
            for lid, tname in liga_team_map.items():
                if lid not in kw_data or lid not in results or results[lid] is None:
                    continue
                sts = kw_data[lid]
                if not sts:
                    continue
                st = sts[0]
                res = results[lid]
                ti  = {t: i for i, t in enumerate(res.cfg.teams)}.get(tname, -1)
                if ti < 0:
                    continue
                is_home = res.home_vals.get((ti, st), 0) == 1
                entries.append((lid, tname, is_home))

            if len(entries) < 2:
                continue

            all_home = all(e[2] for e in entries)
            ligen    = ', '.join(f'{lid}({tname})' for lid, tname, _ in entries)

            # Datum aus erstem Eintrag holen
            first_lid = entries[0][0]
            first_st  = (kw_compat[kw].get(first_lid) or [None])[0]
            cal_entry = results[first_lid].cfg.calendar.get(first_st, {}) if results.get(first_lid) else {}

            ws.cell(r, 1, club_name)
            ws.cell(r, 2, f'KW {kw}')
            ws.cell(r, 3, cal_entry.get('week_start', ''))
            ws.cell(r, 4, cal_entry.get('week_end', ''))
            ws.cell(r, 5, ligen)
            c = ws.cell(r, 6, 'JA' if all_home else 'nein')
            c.fill = GREEN_FILL if all_home else RED_FILL
            c.font = Font(bold=True, color='006100' if all_home else '9C0006')
            c.alignment = Alignment(horizontal='center')
            r += 1

    for col, w in zip(range(1, 7), [24, 8, 12, 12, 60, 10]):
        ws.column_dimensions[get_column_letter(col)].width = w

    return wb


def build_hall_schedule(results: Dict[str, Optional['LeagueResult']]) -> Workbook:
    """Erzeugt einen ligaübergreifenden Hallenbelegungsplan.

    Zeigt alle Heimspiele aller Ligen nach Halle/Spielort sortiert,
    damit Hallenbetreiber und Vereine ihre Belegung auf einen Blick sehen.
    """
    # Alle Heimspiele sammeln
    rows: list = []
    for lid, res in results.items():
        if res is None or not res.cfg:
            continue
        cfg   = res.cfg
        t_idx = {t: i for i, t in enumerate(cfg.teams)}
        round_len = cfg.n_matchdays // max(1, cfg.n_rounds)
        _ph   = {1: 'Hinrunde', 2: 'Rückrunde', 3: 'Dritte Runde'}
        is_tt = cfg.games_per_team_per_day > 1

        for d in cfg.days:
            rnd      = min(cfg.n_rounds, (d - 1) // round_len + 1)
            phase    = _ph.get(rnd, f'Runde {rnd}')
            cal_d    = cfg.calendar.get(d, {})
            kw       = cal_d.get('kw', '')
            w_start  = cal_d.get('week_start', '')
            w_end    = cal_d.get('week_end', '')
            day_times = res.game_times.get(d, [])
            typ_lbl  = ('DST' if d in cfg.dst_days else 'EST') if not is_tt else res.hosts.get(d, '')

            for g_idx, (ht, at) in enumerate(res.schedule.get(d, [])):
                hi  = t_idx.get(ht, -1)
                loc = cfg.locations[hi] if 0 <= hi < len(cfg.locations) else ht
                uhr = day_times[g_idx] if g_idx < len(day_times) else ''
                rows.append({
                    'liga_id':  lid,
                    'liga':     cfg.name,
                    'spieltag': d,
                    'kw':       kw,
                    'w_start':  w_start,
                    'w_end':    w_end,
                    'phase':    phase,
                    'typ':      typ_lbl,
                    'halle':    loc,
                    'heimteam': ht,
                    'gastteam': at,
                    'hi':       hi,
                    'uhrzeit':  uhr,
                })

    # Sortierung: KW (numerisch, leer zuletzt) → Halle → Liga → Spieltag
    def _sort_key(r):
        kw_num = r['kw'] if isinstance(r['kw'], int) else (int(r['kw']) if str(r['kw']).isdigit() else 999)
        return (kw_num, str(r['halle']), str(r['liga']), r['spieltag'])
    rows.sort(key=_sort_key)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Hallenbelegungsplan'

    # Titel
    ws.merge_cells('A1:J1')
    tc = ws.cell(1, 1, 'HALLENBELEGUNGSPLAN')
    tc.fill = TITLE_FILL; tc.font = Font(bold=True, color='FFFFFF', size=13)
    tc.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 20

    # Hinweis
    ws.merge_cells('A2:J2')
    hc = ws.cell(2, 1, f'Erstellt am {datetime.now().strftime("%d.%m.%Y")}  •  '
                        f'{len(rows)} Heimspiele  •  {len(results)} Liga(en)')
    hc.font = Font(italic=True, color='666666', size=9)
    hc.alignment = Alignment(horizontal='left')

    has_uhr   = any(r['uhrzeit'] for r in rows)
    has_dates = any(r['w_start'] for r in rows)

    hdr = ['Liga', 'Spieltag', 'Phase', 'Typ']
    if has_dates:
        hdr += ['KW', 'Woche von', 'Woche bis']
    else:
        hdr.append('KW')
    if has_uhr:
        hdr.append('Uhrzeit')
    hdr += ['Halle / Ort', 'Heimteam', 'Gastteam']

    for col, txt in enumerate(hdr, 1):
        c = ws.cell(3, col, txt)
        c.fill = HDR_FILL; c.font = Font(bold=True, color='FFFFFF')
        c.alignment = Alignment(horizontal='center', vertical='center')

    _halle_col  = hdr.index('Halle / Ort') + 1
    _heim_col   = hdr.index('Heimteam') + 1
    _gast_col   = hdr.index('Gastteam') + 1
    _uhr_col    = hdr.index('Uhrzeit') + 1 if 'Uhrzeit' in hdr else None

    prev_venue = None
    r = 4
    for row in rows:
        venue_changed = (row['halle'] != prev_venue)
        if venue_changed and r > 4:
            # Trennzeile zwischen verschiedenen Hallen
            for col in range(1, len(hdr) + 1):
                ws.cell(r, col).fill = _fill('E8EDF5')
            r += 1
        prev_venue = row['halle']

        col = 1
        ws.cell(r, col, row['liga']); col += 1
        ws.cell(r, col, row['spieltag']).alignment = Alignment(horizontal='center'); col += 1
        ws.cell(r, col, row['phase']); col += 1
        ws.cell(r, col, row['typ']).alignment = Alignment(horizontal='center'); col += 1
        if has_dates:
            ws.cell(r, col, f'KW {row["kw"]}' if row['kw'] else '').alignment = Alignment(horizontal='center'); col += 1
            ws.cell(r, col, row['w_start']); col += 1
            ws.cell(r, col, row['w_end']); col += 1
        else:
            ws.cell(r, col, f'KW {row["kw"]}' if row['kw'] else '').alignment = Alignment(horizontal='center'); col += 1
        if _uhr_col:
            ws.cell(r, _uhr_col, row['uhrzeit']).alignment = Alignment(horizontal='center')
            col += 1
        ws.cell(r, _halle_col, row['halle']).font = Font(bold=True)
        hi = row['hi']
        c_h = ws.cell(r, _heim_col, row['heimteam'])
        c_g = ws.cell(r, _gast_col, row['gastteam'])
        c_h.fill = _fill(get_team_color(hi))
        c_g.fill = GRAY_FILL
        r += 1

    # Spaltenbreiten
    widths = {'Liga': 22, 'Spieltag': 9, 'Phase': 12, 'Typ': 6,
              'KW': 7, 'Woche von': 12, 'Woche bis': 12,
              'Uhrzeit': 9, 'Halle / Ort': 28, 'Heimteam': 24, 'Gastteam': 24}
    for col, h in enumerate(hdr, 1):
        ws.column_dimensions[get_column_letter(col)].width = widths.get(h, 12)

    ws.freeze_panes = 'A4'
    ws.auto_filter.ref = f'A3:{get_column_letter(len(hdr))}3'

    return wb


def save_cohome_summary(wb: Workbook, output_dir: Path) -> str:
    date_str  = datetime.now().strftime('%Y-%m-%d')
    filename  = output_dir / f'{date_str}_CoHome_Zusammenfassung.xlsx'
    filename.parent.mkdir(parents=True, exist_ok=True)
    wb.save(filename)
    return str(filename)


def save_hall_schedule(wb: Workbook, output_dir: Path) -> str:
    date_str = datetime.now().strftime('%Y-%m-%d')
    filename = output_dir / f'{date_str}_Hallenbelegungsplan.xlsx'
    filename.parent.mkdir(parents=True, exist_ok=True)
    wb.save(filename)
    return str(filename)
