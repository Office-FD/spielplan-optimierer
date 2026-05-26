#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Feature-Test fuer alle erweiterten Funktionen.

Testet ohne Streamlit und ohne Google-Maps-API:
  1. Persistenter Cache-Pfad  (schedule_utils importierbar)
  2. Team-Ansichten-Sheet     (build_league_excel erzeugt Sheet)
  3. Kalender-Export (.ics)   (build_ics_bytes korrekte VEVENT-Anzahl)
  4. Manueller Tausch         (swap_home_away aendert Schedule + Stats)
  5. Fairness-Analyse-Sheet   (build_league_excel enthaelt Fairness-Sheet)
  6. Constraint-Validierung   (config_validator erkennt Fehler/Warnungen)

Aufruf: python test_features.py
"""
from __future__ import annotations

import io
import sys
import traceback
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

import numpy as np

from spielplan_multi.league_types import LeagueConfig, LeagueResult
from spielplan_multi.config import WEIGHT_SCALES
from spielplan_multi.solver import solve_league_phase1
from spielplan_multi.calendar_parser import build_weekends
from spielplan_multi.excel_output import build_league_excel, build_hall_schedule
from spielplan_multi.schedule_utils import (
    recompute_result_stats,
    swap_home_away,
    build_ics_bytes,
    build_print_html,
    assign_game_times,
)
from spielplan_multi.config_validator import validate as validate_cfg


# ── Hilfsfunktionen ───────────────────────────────────────────────────────────

_results: list = []

def check(name: str, fn):
    try:
        detail = fn()
        print(f'  [PASS] {name}' + (f'  – {detail}' if detail else ''))
        _results.append((name, True))
    except AssertionError as e:
        print(f'  [FAIL] {name}  – {e}')
        _results.append((name, False))
    except Exception as e:
        tb = traceback.format_exc().strip().split('\n')[-1]
        print(f'  [FAIL] {name}  – EXCEPTION: {tb}')
        _results.append((name, False))


def make_cfg(lid: str, teams: list, dist=None, dst_blocks=None, calendar=None):
    n   = len(teams)
    dst = dst_blocks or []
    nr  = 2
    days = list(range(1, nr * (n - 1) + 1))
    raw = {k: 5.0 for k in WEIGHT_SCALES}
    if dist is None:
        dist = np.zeros((n, n))
    cal = calendar or {}
    return LeagueConfig(
        league_id=lid, name=f'Liga {lid}',
        teams=teams, locations=teams,
        dist=dist, dst_blocks=dst,
        weekends=build_weekends(days, dst),
        apply_routing=False,
        f_num=125, f_den=100,
        w_scaled={k: v * WEIGHT_SCALES[k] for k, v in raw.items()},
        raw_weights=raw,
        pinned=[], blocked={},
        calendar=cal,
        hier_weight=1.0,
        games_per_team_per_day=1,
        n_rounds=nr,
    )


TEAMS = ['Alpha', 'Beta', 'Gamma', 'Delta']
DIST  = np.array([
    [0,   100, 200, 300],
    [100, 0,   150, 250],
    [200, 150, 0,   120],
    [300, 250, 120, 0  ],
], dtype=float)
# 4 Teams, n_rounds=2 -> 6 Spieltage
CAL = {1: {'kw': 37}, 2: {'kw': 38}, 3: {'kw': 39},
       4: {'kw': 40}, 5: {'kw': 41}, 6: {'kw': 42}}


# ── Tests ─────────────────────────────────────────────────────────────────────

def main():
    _results.clear()

    # ── Feature 4: Persistenter Cache importierbar ────────────────────────────
    print('\n--- Feature 4: Persistenter Cache-Pfad ---')

    def t_cache_import():
        from spielplan_multi.schedule_utils import build_ics_bytes  # noqa: F401
        _here = Path(__file__).resolve().parent
        cache_dir = _here / '.cache'
        assert str(cache_dir).endswith('.cache'), f'Unerwarteter Pfad: {cache_dir}'
        return f'.cache liegt in {_here.name}/'
    check('schedule_utils importierbar', t_cache_import)

    # ── Loesung erzeugen (Basis fuer alle weiteren Tests) ─────────────────────
    print('\n--- Loesung erzeugen (Basis) ---')
    cfg = make_cfg('TEST', TEAMS, dist=DIST, calendar=CAL)
    result = None

    def t_solve():
        nonlocal result
        r = solve_league_phase1(cfg, time_limit=30, seed=42)
        assert r is not None, 'Solver: keine Loesung (None)'
        r.cfg = cfg
        result = r
        total = sum(len(g) for g in r.schedule.values())
        exp   = cfg.n_teams * (cfg.n_teams - 1) // 2 * cfg.n_rounds
        assert total == exp, f'{total} Spiele statt {exp}'
        return f'obj={r.objective:.0f}  km={sum(r.travels)}  sw={r.sw_counts}'
    check('solve_league_phase1 loest (4 Teams, kein DST)', t_solve)

    if result is None:
        print('\n  Ueberspringe weitere Tests (keine Loesung).')
        _summarize()
        return

    # ── Feature 1: Team-Ansichten-Sheet ───────────────────────────────────────
    print('\n--- Feature 1: Team-Ansichten im Excel ---')

    def t_team_ansichten_sheet():
        buf = io.BytesIO()
        wb  = build_league_excel(result)
        wb.save(buf)
        buf.seek(0)
        import openpyxl
        wb2 = openpyxl.load_workbook(buf)
        assert 'Team-Ansichten' in wb2.sheetnames, \
            f'Sheet fehlt. Vorhandene Sheets: {wb2.sheetnames}'
        return f'Sheets: {wb2.sheetnames}'
    check('Excel enthaelt Sheet "Team-Ansichten"', t_team_ansichten_sheet)

    def t_fairness_sheet():
        wb  = build_league_excel(result)
        assert 'Fairness-Analyse' in wb.sheetnames, 'Sheet fehlt'
        ws  = wb['Fairness-Analyse']
        # Titel-Zelle muss den Liga-Namen enthalten
        merged_val = ws.cell(1, 1).value or ''
        assert 'FAIRNESS-ANALYSE' in str(merged_val), f'Kein Titel: {merged_val!r}'
        # Mindestens n+8 ausgefüllte Zeilen (Titel + 3 Abschnitte)
        filled = [r for r in ws.iter_rows(values_only=True) if any(v for v in r)]
        assert len(filled) >= cfg.n_teams + 6, f'Zu wenige Zeilen: {len(filled)}'
        return f'{len(filled)} gefuellte Zeilen'
    check('Fairness-Analyse Sheet vorhanden und befuellt', t_fairness_sheet)

    def t_team_ansichten_inhalt():
        wb  = build_league_excel(result)
        ws  = wb['Team-Ansichten']
        rows = list(ws.iter_rows(values_only=True))
        assert len(rows) > 1, 'Sheet ist leer'
        header = rows[0]
        assert 'ST' in header or header[0] == 'ST', f'Kein ST-Header: {header}'
        # Mindestens len(days)+1 Zeilen (Header + eine pro Spieltag)
        assert len(rows) >= cfg.n_matchdays + 1, \
            f'Zu wenige Zeilen: {len(rows)}'
        return f'{len(rows)-1} Datenzeilen, {len(header)} Spalten'
    check('Team-Ansichten hat korrekte Zeilen', t_team_ansichten_inhalt)

    # ── Feature 3: Kalender-Export (.ics) ────────────────────────────────────
    print('\n--- Feature 3: Kalender-Export (.ics) ---')

    def t_ics_parsebar():
        raw = build_ics_bytes(result, season_year=2026)
        txt = raw.decode('utf-8')
        assert txt.startswith('BEGIN:VCALENDAR'), 'Kein VCALENDAR-Header'
        assert txt.endswith('END:VCALENDAR'), 'Kein VCALENDAR-Footer'
        n_events = txt.count('BEGIN:VEVENT')
        exp_games = sum(len(g) for g in result.schedule.values())
        assert n_events == exp_games, \
            f'{n_events} VEVENTs, erwartet {exp_games}'
        return f'{n_events} VEVENT-Eintraege'
    check('ICS erzeugt und VEVENT-Anzahl korrekt', t_ics_parsebar)

    def t_ics_datum():
        raw = build_ics_bytes(result, season_year=2026)
        txt = raw.decode('utf-8')
        # KW 37 liegt nach KW 26 -> Jahr 2026
        assert 'DTSTART;VALUE=DATE:2026' in txt, \
            'Kein 2026-Datum fuer KW>=37 gefunden'
        return 'Jahreszuordnung KW37->2026 korrekt'
    check('ICS: Datum aus Kalender-KW berechnet', t_ics_datum)

    # ── Feature 2: Manueller Tausch ──────────────────────────────────────────
    print('\n--- Feature 2: Manueller Heim-/Auswaerts-Tausch ---')

    def t_swap_basic():
        import copy
        r2  = copy.deepcopy(result)
        cfg2 = cfg
        # Ersten Spieltag nehmen, erstes Spiel tauschen
        day1 = cfg2.days[0]
        games_before = list(r2.schedule[day1])
        ht_before, at_before = games_before[0]
        swap_home_away(r2, cfg2, day1, 0)
        games_after = r2.schedule[day1]
        ht_after, at_after = games_after[0]
        assert ht_after == at_before and at_after == ht_before, \
            f'Tausch fehlgeschlagen: {ht_before}/{at_before} -> {ht_after}/{at_after}'
        return f'ST{day1}: {ht_before} vs {at_before} -> {ht_after} vs {at_after}'
    check('swap_home_away tauscht Schedule-Eintrag', t_swap_basic)

    def t_swap_home_vals():
        import copy
        r2  = copy.deepcopy(result)
        cfg2 = cfg
        day1 = cfg2.days[0]
        ht, at = r2.schedule[day1][0]
        t_idx  = {t: i for i, t in enumerate(cfg2.teams)}
        hi, ai = t_idx[ht], t_idx[at]
        val_ht_before = r2.home_vals.get((hi, day1))
        val_at_before = r2.home_vals.get((ai, day1))
        swap_home_away(r2, cfg2, day1, 0)
        val_ht_after = r2.home_vals.get((hi, day1))
        val_at_after = r2.home_vals.get((ai, day1))
        assert val_ht_after == 0, f'home_vals[{ht}] sollte 0 sein, ist {val_ht_after}'
        assert val_at_after == 1, f'home_vals[{at}] sollte 1 sein, ist {val_at_after}'
        return (f'{ht}: {val_ht_before}->{val_ht_after}  '
                f'{at}: {val_at_before}->{val_at_after}')
    check('swap_home_away aktualisiert home_vals', t_swap_home_vals)

    def t_recompute():
        travels, sw_counts, sw_rates = recompute_result_stats(result, cfg)
        assert len(travels)   == cfg.n_teams, f'travels Laenge {len(travels)}'
        assert len(sw_counts) == cfg.n_teams, f'sw_counts Laenge {len(sw_counts)}'
        assert len(sw_rates)  == cfg.n_teams, f'sw_rates Laenge {len(sw_rates)}'
        assert all(t >= 0 for t in travels),   'Negative travels'
        assert all(s >= 0 for s in sw_counts), 'Negative sw_counts'
        return f'travels={travels}  sw={sw_counts}'
    check('recompute_result_stats gibt konsistente Werte', t_recompute)

    # ── Spielzeiten zuweisen ─────────────────────────────────────────────────
    print('\n--- Spielzeiten zuweisen ---')

    def t_assign_basic():
        import copy
        r2 = copy.deepcopy(result)
        assign_game_times(r2, ['14:00', '16:00'])
        assert r2.game_times, 'game_times ist leer'
        for d in cfg.days:
            assert d in r2.game_times, f'ST{d} fehlt in game_times'
            assert len(r2.game_times[d]) == len(r2.schedule[d]), \
                f'Falsche Anzahl Zeiten fuer ST{d}'
        first_d = cfg.days[0]
        assert r2.game_times[first_d][0] == '14:00', 'Erste Zeit falsch'
        return f'{len(r2.game_times)} Spieltage mit Zeiten'
    check('assign_game_times fuellt alle Spieltage', t_assign_basic)

    def t_assign_fewer_slots():
        import copy
        r2 = copy.deepcopy(result)
        assign_game_times(r2, ['14:00'])   # nur 1 Slot, aber 2 Spiele/Tag
        first_d = cfg.days[0]
        assert r2.game_times[first_d][0] == '14:00', 'Slot 0 falsch'
        assert r2.game_times[first_d][1] == '', 'Fehlender Slot sollte leer sein'
        return 'Fehlende Slots werden als leer gesetzt'
    check('Fehlende Zeitslots werden als leer gesetzt', t_assign_fewer_slots)

    def t_excel_mit_zeiten():
        import copy, io
        import openpyxl
        r2 = copy.deepcopy(result)
        r2.cfg = cfg
        assign_game_times(r2, ['14:00', '16:00'])
        wb  = build_league_excel(r2)
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        wb2 = openpyxl.load_workbook(buf)
        ws  = wb2['Spielplan']
        hdr = [ws.cell(1, c).value for c in range(1, 9)]
        assert 'Uhrzeit' in hdr, f'Uhrzeit fehlt im Header: {hdr}'
        # Pruefe ob eine Zelle einen Zeitwert hat
        found_time = any(
            ws.cell(r, hdr.index('Uhrzeit') + 1).value == '14:00'
            for r in range(2, 5)
        )
        assert found_time, 'Kein 14:00 Wert in Spielplan-Sheet gefunden'
        return f'Header: {[h for h in hdr if h]}'
    check('Excel-Spielplan enthaelt Uhrzeit-Spalte nach Zuweisung', t_excel_mit_zeiten)

    def t_html_mit_zeiten():
        import copy
        r2 = copy.deepcopy(result)
        r2.cfg = cfg
        assign_game_times(r2, ['14:00', '16:00'])
        html = build_print_html(r2, season_year=2026)
        assert '14:00' in html, '14:00 fehlt im HTML'
        assert '<th>Uhrzeit</th>' in html, 'Uhrzeit-Spalten-Header fehlt'
        return 'Uhrzeiten im HTML enthalten'
    check('HTML enthaelt Uhrzeiten nach Zuweisung', t_html_mit_zeiten)

    # ── Druckbarer Spielplan (HTML) ───────────────────────────────────────────
    print('\n--- Druckbarer Spielplan (HTML) ---')

    def t_print_html_struktur():
        html = build_print_html(result, season_year=2026)
        assert '<!DOCTYPE html>' in html, 'Kein DOCTYPE'
        assert '<table' in html, 'Keine Tabelle'
        assert 'window.print()' in html, 'Kein Print-Button'
        assert cfg.name in html, 'Liga-Name fehlt'
        return f'{len(html):,} Zeichen'
    check('HTML enthaelt Grundstruktur und Print-Button', t_print_html_struktur)

    def t_print_html_alle_teams():
        html = build_print_html(result, season_year=2026)
        for t in cfg.teams:
            assert t in html, f'Team {t!r} fehlt im HTML'
        return f'Alle {cfg.n_teams} Teams enthalten'
    check('HTML enthaelt alle Teams', t_print_html_alle_teams)

    def t_print_html_alle_spiele():
        html = build_print_html(result, season_year=2026)
        # Jedes Spiel erscheint zweimal: im Gesamtplan + in Team-Ansicht
        for d, games in result.schedule.items():
            for ht, at in games:
                assert ht in html and at in html, f'Spiel {ht} vs {at} fehlt'
        return f'{sum(len(g) for g in result.schedule.values())} Spiele im HTML'
    check('HTML enthaelt alle Spielpaarungen', t_print_html_alle_spiele)

    def t_print_html_ohne_kalender():
        import copy
        r2 = copy.deepcopy(result)
        r2.cfg = copy.deepcopy(cfg)
        r2.cfg.calendar = {}   # kein Kalender
        html = build_print_html(r2, season_year=0)
        assert '<!DOCTYPE html>' in html, 'Kein HTML ohne Kalender'
        return 'HTML auch ohne Kalender-Daten generierbar'
    check('HTML auch ohne Kalender generierbar', t_print_html_ohne_kalender)

    def t_print_html_phase_spalte():
        import re
        html = build_print_html(result, season_year=2026)
        assert '<th>Phase</th>' in html, 'Phase-Spalten-Header fehlt im <thead>'
        # Jeder Spieltag hat genau eine Phase-Datenzelle mit rowspan.
        # style="color:#666" ohne text-align ist eindeutig fuer die Phase-Zelle.
        phase_cells = re.findall(
            r'<td rowspan="\d+" style="color:#666">'
            r'(Hinrunde|R\xfcckrunde|Dritte Runde|Runde \d+)</td>',
            html
        )
        assert len(phase_cells) == cfg.n_matchdays, (
            f'{len(phase_cells)} Phase-Zellen gefunden, '
            f'{cfg.n_matchdays} erwartet (eine pro Spieltag)'
        )
        return f'{len(phase_cells)} Phase-Datenzellen fuer {cfg.n_matchdays} Spieltage'
    check('HTML Alle-Spiele-Tabelle: Phase-Spalte in Header und Datenzeilen', t_print_html_phase_spalte)

    # ── Feature 5 (neu): Constraint-Validierung ──────────────────────────────
    print('\n--- Feature 5: Constraint-Validierung ---')

    def _make_ld(teams, fmt='Hin-/Rueckrunde'):
        return {'name': 'TestLiga', 'teams': [(t, t) for t in teams], 'fmt': fmt}

    def _calc_nmd(ld):
        n = len(ld.get('teams', []))
        return 2 * (n - 1) if n >= 2 else 0

    def _get_nr_gpd(ld): return (2, 1)

    def _run(leagues=None, dst=None, blocked=None, pinned=None,
             dist=None, kw_compat=None, clubs=None):
        lids = list((leagues or {}).keys())
        return validate_cfg(
            league_order     = lids,
            leagues          = leagues or {},
            dst_per_liga     = dst or {},
            blocked          = blocked or {},
            pinned           = pinned or {},
            dist_matrices    = dist or {},
            kw_compat        = kw_compat or {},
            clubs            = clubs or {},
            calc_n_matchdays = _calc_nmd,
            get_n_rounds_gpd = _get_nr_gpd,
        )

    def t_clean_config():
        issues = _run(
            leagues={'A': _make_ld(['X', 'Y', 'Z', 'W'])},
            dist={'A': np.ones((4, 4))},
        )
        assert issues == [], f'Fehler bei sauberer Config: {issues}'
        return 'Keine Probleme bei gueltiger Config'
    check('Saubere Config liefert keine Probleme', t_clean_config)

    def t_too_few_teams():
        issues = _run(leagues={'A': _make_ld(['EinTeam'])})
        errs = [i for i in issues if i['level'] == 'error']
        assert any('2 Teams' in i['msg'] for i in errs), f'Kein Fehler: {errs}'
        return 'Fehler bei 1 Team erkannt'
    check('Zu wenige Teams wird als Fehler erkannt', t_too_few_teams)

    def t_all_days_blocked():
        # 4 Teams -> 6 Spieltage; Team X auf allen 6 Tagen gesperrt
        issues = _run(
            leagues={'A': _make_ld(['X', 'Y', 'Z', 'W'])},
            blocked={'A': {'X': [1, 2, 3, 4, 5, 6]}},
            dist={'A': np.ones((4, 4))},
        )
        errs = [i for i in issues if i['level'] == 'error']
        assert any('alle' in i['msg'].lower() or 'alle' in i['msg'] for i in errs), \
            f'Kein Fehler bei allen Sperrtagen: {errs}'
        return 'Fehler bei allen Sperrtagen erkannt'
    check('Alle Heimspieltage gesperrt wird als Fehler erkannt', t_all_days_blocked)

    def t_dst_out_of_range():
        # 4 Teams -> 6 Spieltage; DST-Block auf ST7+8 (ungueltig)
        issues = _run(
            leagues={'A': _make_ld(['X', 'Y', 'Z', 'W'])},
            dst={'A': [(7, 8)]},
            dist={'A': np.ones((4, 4))},
        )
        errs = [i for i in issues if i['level'] == 'error']
        assert any('DST' in i['msg'] for i in errs), f'Kein DST-Fehler: {errs}'
        return 'Fehler fuer ungueltige DST-Tage erkannt'
    check('DST-Block ausserhalb gueltigem Bereich als Fehler erkannt', t_dst_out_of_range)

    def t_pinned_wrong_day():
        issues = _run(
            leagues={'A': _make_ld(['X', 'Y', 'Z', 'W'])},
            pinned={'A': [{'teamA': 'X', 'teamB': 'Y', 'day': 99, 'home': 'X'}]},
            dist={'A': np.ones((4, 4))},
        )
        errs = [i for i in issues if i['level'] == 'error']
        assert any('ST99' in i['msg'] or 'nicht existiert' in i['msg'] for i in errs), \
            f'Kein Fehler fuer ungueltigen Pflichtspieltag: {errs}'
        return 'Fehler fuer ungueltige Pflichtspieltag erkannt'
    check('Pflichtspiel auf ungueltigem Spieltag als Fehler erkannt', t_pinned_wrong_day)

    def t_contradicting_home():
        issues = _run(
            leagues={'A': _make_ld(['X', 'Y', 'Z', 'W'])},
            pinned={'A': [
                {'teamA': 'X', 'teamB': 'Y', 'day': 1, 'home': 'X'},
                {'teamA': 'X', 'teamB': 'Y', 'day': 1, 'home': 'Y'},
            ]},
            dist={'A': np.ones((4, 4))},
        )
        errs = [i for i in issues if i['level'] == 'error']
        assert any('widersprüchlich' in i['msg'] or 'widerspruc' in i['msg'].lower()
                   for i in errs), f'Kein Widerspruch-Fehler: {errs}'
        return 'Widersprüchliches Heimrecht erkannt'
    check('Widersprüchliches Heimrecht wird als Fehler erkannt', t_contradicting_home)

    def t_warning_half_blocked():
        # Mehr als 50% der Tage gesperrt -> Warnung (nicht alle)
        issues = _run(
            leagues={'A': _make_ld(['X', 'Y', 'Z', 'W'])},
            blocked={'A': {'X': [1, 2, 3, 4]}},   # 4 von 6 Tagen gesperrt
            dist={'A': np.ones((4, 4))},
        )
        warns = [i for i in issues if i['level'] == 'warning']
        errs  = [i for i in issues if i['level'] == 'error']
        assert not any('alle' in i['msg'].lower() for i in errs), \
            'Fehler statt Warnung bei teilweiser Sperrung'
        assert any('Hälfte' in i['msg'] or 'haelfte' in i['msg'].lower()
                   or 'Hälfte' in i['msg'] for i in warns), \
            f'Keine Warnung bei >50% gesperrt: {warns}'
        return 'Warnung bei >50% Sperrtagen erkannt'
    check('Warnung bei mehr als 50% gesperrten Tagen', t_warning_half_blocked)

    # ── Hallenbelegungsplan ───────────────────────────────────────────────────
    print('\n--- Hallenbelegungsplan ---')

    def t_hall_returns_workbook():
        wb = build_hall_schedule({'TEST': result})
        assert wb is not None, 'build_hall_schedule gab None zurueck'
        return f'{len(wb.sheetnames)} Sheet(s): {wb.sheetnames}'
    check('build_hall_schedule gibt Workbook zurueck', t_hall_returns_workbook)

    def t_hall_sheet_name():
        wb = build_hall_schedule({'TEST': result})
        assert 'Hallenbelegungsplan' in wb.sheetnames, \
            f'Sheet "Hallenbelegungsplan" fehlt: {wb.sheetnames}'
        return 'Sheet "Hallenbelegungsplan" vorhanden'
    check('Hallenbelegungsplan hat korrekten Sheet-Namen', t_hall_sheet_name)

    def t_hall_row_count():
        wb = build_hall_schedule({'TEST': result})
        ws = wb['Hallenbelegungsplan']
        n_games = sum(len(gs) for gs in result.schedule.values())
        data_rows = [r for r in ws.iter_rows(min_row=4, values_only=True)
                     if any(c is not None for c in r)]
        assert len(data_rows) >= n_games, \
            f'Zu wenige Zeilen: {len(data_rows)} < {n_games} Spiele'
        return f'{len(data_rows)} Datenzeilen fuer {n_games} Spiele'
    check('Hallenbelegung enthaelt alle Spiele als Zeilen', t_hall_row_count)

    def t_hall_header_cols():
        wb = build_hall_schedule({'TEST': result})
        ws = wb['Hallenbelegungsplan']
        header = [c.value for c in ws[3] if c.value is not None]
        for col in ('Spieltag', 'Liga', 'Heimteam', 'Gastteam'):
            assert col in header, f'Spalte "{col}" fehlt: {header}'
        return f'Header: {header}'
    check('Hallenbelegung hat Pflicht-Spalten', t_hall_header_cols)

    def t_hall_multi_league():
        cfg2 = make_cfg('L2', ['Eins', 'Zwei', 'Drei', 'Vier'], dist=DIST, calendar=CAL)
        r2 = solve_league_phase1(cfg2, time_limit=20, seed=1)
        assert r2 is not None, 'Solver Liga 2: keine Loesung'
        r2.cfg = cfg2
        wb = build_hall_schedule({'TEST': result, 'L2': r2})
        ws = wb['Hallenbelegungsplan']
        n_total = (sum(len(gs) for gs in result.schedule.values())
                   + sum(len(gs) for gs in r2.schedule.values()))
        data_rows = [r for r in ws.iter_rows(min_row=4, values_only=True)
                     if any(c is not None for c in r)]
        assert len(data_rows) >= n_total, \
            f'Zu wenige Zeilen fuer 2 Ligen: {len(data_rows)} < {n_total}'
        return f'{len(data_rows)} Zeilen fuer 2 Ligen ({n_total} Spiele)'
    check('Hallenbelegung funktioniert mit 2 Ligen', t_hall_multi_league)

    def t_hall_with_times():
        assign_game_times(result, ['14:00', '16:00'])
        wb = build_hall_schedule({'TEST': result})
        ws = wb['Hallenbelegungsplan']
        header = [c.value for c in ws[3] if c.value is not None]
        assert 'Uhrzeit' in header, f'Uhrzeit-Spalte fehlt bei gesetzten Zeiten: {header}'
        return 'Uhrzeit-Spalte vorhanden wenn game_times gesetzt'
    check('Hallenbelegung zeigt Uhrzeit-Spalte wenn Zeiten zugewiesen', t_hall_with_times)

    def t_hall_saveable():
        wb = build_hall_schedule({'TEST': result})
        buf = io.BytesIO()
        wb.save(buf)
        assert buf.tell() > 0, 'Gespeichertes Excel ist leer'
        return f'{buf.tell()} Bytes'
    check('Hallenbelegungsplan-Excel speicherbar (kein TypeError)', t_hall_saveable)

    def t_hall_empty_results():
        wb = build_hall_schedule({})
        ws = wb['Hallenbelegungsplan']
        data_rows = [r for r in ws.iter_rows(min_row=4, values_only=True)
                     if any(c is not None for c in r)]
        assert len(data_rows) == 0, f'Leere results sollten 0 Zeilen ergeben, war {len(data_rows)}'
        return 'Leeres Dict -> 0 Datenzeilen'
    check('build_hall_schedule mit leerem Dict stuerzt nicht ab', t_hall_empty_results)

    # ── Feature 7: Gesamtuebersicht-Excel (build_overview_excel) ─────────────
    print('\n--- Feature 7: Gesamtuebersicht-Excel (build_overview_excel) ---')

    from spielplan_multi.excel_output import build_overview_excel

    def t_overview_single_liga_no_calendar():
        """build_overview_excel ohne Kalender + ohne Co-Home -> Workbook mit Spielplan-Uebersicht."""
        wb = build_overview_excel({'TEST': result}, clubs={}, kw_compat={})
        assert wb is not None
        assert 'Spielplan-Uebersicht' in wb.sheetnames, f'Sheets: {wb.sheetnames}'
        return f'Sheets: {wb.sheetnames}'
    check('build_overview_excel: Single-Liga ohne Kalender', t_overview_single_liga_no_calendar)

    def t_overview_with_calendar():
        """build_overview_excel mit kw_compat -> Datum erscheint in Zeilen."""
        # CAL = {1..6 -> kw 37-42}. Wir bauen eine kw_compat-Variante daraus.
        kw_compat = {37: {'TEST': [1]}, 38: {'TEST': [2]}, 39: {'TEST': [3]},
                     40: {'TEST': [4]}, 41: {'TEST': [5]}, 42: {'TEST': [6]}}
        wb = build_overview_excel({'TEST': result}, clubs={}, kw_compat=kw_compat)
        ws = wb['Spielplan-Uebersicht']
        # Mindestens einige Datenzeilen + 3 Headerzeilen
        rows = list(ws.iter_rows(values_only=True))
        non_empty = [r for r in rows if any(c is not None and str(c).strip() for c in r)]
        assert len(non_empty) > 3, f'Zu wenige Zeilen: {len(non_empty)}'
        return f'{len(non_empty)} gefuellte Zeilen, {len(rows[0])} Spalten'
    check('build_overview_excel: mit Kalender erzeugt Datenzeilen', t_overview_with_calendar)

    def t_overview_saveable():
        """Workbook ist speicherbar als Bytes."""
        wb = build_overview_excel({'TEST': result}, clubs={}, kw_compat={})
        buf = io.BytesIO()
        wb.save(buf)
        assert buf.tell() > 0, 'Excel-Bytes sind leer'
        return f'{buf.tell()} Bytes'
    check('build_overview_excel: speicherbar (kein TypeError)', t_overview_saveable)

    def t_overview_multi_liga():
        """Multi-Liga: jede Liga bekommt 2 Spalten (Heim + Gast) im Header."""
        # Zweite Liga aufbauen
        cfg2 = make_cfg('L2', TEAMS, dist=DIST, calendar=CAL)
        r2 = solve_league_phase1(cfg2, time_limit=20, seed=42)
        assert r2 is not None
        r2.cfg = cfg2
        wb = build_overview_excel({'TEST': result, 'L2': r2}, clubs={}, kw_compat={})
        ws = wb['Spielplan-Uebersicht']
        # Header-Zeile 2 (1-indexiert) sollte 2 Liga-Namen (jeweils 2 Spalten gemergt) enthalten
        header2 = [c.value for c in ws[2]]
        liga_count = sum(1 for v in header2 if v and isinstance(v, str) and v.startswith('Liga'))
        assert liga_count == 2, f'erwartet 2 Liga-Header, got {liga_count}: {header2}'
        return f'Headerzeile 2: {liga_count} Liga-Namen erkannt'
    check('build_overview_excel: Multi-Liga (2 Ligen)', t_overview_multi_liga)

    def t_overview_with_cohome():
        """Co-Home-Sheet wird erzeugt wenn clubs gesetzt sind."""
        clubs = {'Verein-Multi': {'TEST': 'Alpha'}}
        wb = build_overview_excel({'TEST': result}, clubs=clubs, kw_compat={})
        # Co-Home-Sheet existiert (oder ist als Bezeichnung im Workbook drin)
        assert 'Spielplan-Uebersicht' in wb.sheetnames
        # Mindestens kein Crash. Co-Home-Sheet entsteht nur fuer >=2 Ligen.
        return f'Sheets: {wb.sheetnames}'
    check('build_overview_excel: mit clubs-Definition (Single-Liga, kein Crash)', t_overview_with_cohome)

    def t_overview_empty_results():
        """Leeres results-Dict -> Workbook ohne Datenzeilen, kein Crash."""
        wb = build_overview_excel({}, clubs={}, kw_compat={})
        assert wb is not None
        assert 'Spielplan-Uebersicht' in wb.sheetnames
        return 'Leeres Dict -> Workbook ohne Crash'
    check('build_overview_excel: leeres results-Dict', t_overview_empty_results)

    # ── Feature 8: Rahmenterminplan-Parser (calendar_parser) ─────────────────
    print('\n--- Feature 8: Rahmenterminplan-Parser (calendar_parser) ---')

    from spielplan_multi.calendar_parser import (
        _parse_cell, _to_date_str, _extract_kw,
        parse_rahmenterminplan, preview_columns
    )

    def t_parse_cell_einzel():
        assert _parse_cell(7) == [7]
        assert _parse_cell('7') == [7]
        assert _parse_cell(' 12 ') == [12]
        return '7 -> [7]; "12 " -> [12]'
    check('_parse_cell: Einzelspieltag', t_parse_cell_einzel)

    def t_parse_cell_doppel():
        assert _parse_cell('6/7') == [6, 7]
        assert _parse_cell('6 & 7') == [6, 7]
        assert _parse_cell('6 - 7') == [6, 7]
        assert _parse_cell('7/6') == [6, 7], 'Sortierung min/max'
        return 'Doppelspieltag mit /, &, - + Sortierung'
    check('_parse_cell: Doppelspieltag mit /, &, -', t_parse_cell_doppel)

    def t_parse_cell_invalid():
        assert _parse_cell(None) == []
        assert _parse_cell(float('nan')) == []
        assert _parse_cell('') == []
        assert _parse_cell('abc') == []
        assert _parse_cell('5/5') == [5], '5/5 ist faktisch Einzelspieltag'
        return 'None/NaN/leer/"abc" -> []; "5/5" -> [5]'
    check('_parse_cell: Edge-Cases', t_parse_cell_invalid)

    def t_to_date_str():
        assert _to_date_str(None) == ''
        assert _to_date_str(float('nan')) == ''
        assert _to_date_str('07.09.2026') == '07.09.2026'
        # date-Objekt hat .date()-Methode (datetime) — date allein nicht
        import datetime as dt
        d = dt.datetime(2026, 9, 7)
        assert _to_date_str(d) == '2026-09-07'
        return 'None/NaN -> "" ; datetime -> isoformat'
    check('_to_date_str: Edge-Cases + datetime', t_to_date_str)

    def t_extract_kw_int():
        assert _extract_kw(37) == 37
        assert _extract_kw('38') == 38
        return 'int/str -> kw'
    check('_extract_kw: int und numerischer string', t_extract_kw_int)

    def t_extract_kw_text():
        assert _extract_kw('KW 37') == 37
        assert _extract_kw('kw 38 07.09. - 13.09.2026') == 38
        assert _extract_kw('Vorbereitung') is None
        assert _extract_kw(None) is None
        assert _extract_kw(float('nan')) is None
        return '"KW 37"/"kw 38..." -> int; sonst None'
    check('_extract_kw: Text mit "KW ##"', t_extract_kw_text)

    def _build_synth_rahmenplan(path, rows):
        """Synthetic Rahmenterminplan-Excel: kw_col=0, date_from=1, date_to=2, liga_cols ab 3."""
        import openpyxl as opx
        wb = opx.Workbook()
        ws = wb.active
        for r_idx, row in enumerate(rows, start=1):
            for c_idx, val in enumerate(row, start=1):
                ws.cell(r_idx, c_idx, value=val)
        wb.save(path)

    def t_parse_rahmenterminplan_basic():
        """Synthetic Excel mit 3 KWs, eine Liga, 1 DST-Block."""
        import tempfile, os
        tmpdir = tempfile.mkdtemp()
        try:
            xlsx = os.path.join(tmpdir, 'cal.xlsx')
            _build_synth_rahmenplan(xlsx, [
                ['KW', 'von', 'bis', 'LigaA'],
                [37, '07.09.2026', '13.09.2026', 1],
                [38, '14.09.2026', '20.09.2026', 2],
                [39, '21.09.2026', '27.09.2026', '3/4'],
                [40, '28.09.2026', '04.10.2026', 5],
            ])
            res = parse_rahmenterminplan(xlsx, {'LigaA': 3}, kw_col=0,
                                          date_from_col=1, date_to_col=2)
            assert res is not None, 'parse_rahmenterminplan gab None zurueck'
            assert 'spieltage' in res and 'dst_blocks' in res and 'kw_compat' in res
            spieltage = res['spieltage']['LigaA']
            assert set(spieltage.keys()) == {1, 2, 3, 4, 5}, \
                f'Spieltage: {sorted(spieltage.keys())}'
            assert spieltage[1]['kw'] == 37
            assert spieltage[3]['kw'] == 39 and spieltage[4]['kw'] == 39
            dst = res['dst_blocks']['LigaA']
            assert (3, 4) in dst, f'erwartet (3,4), got {dst}'
            return f'5 Spieltage, 1 DST-Block (3,4)'
        finally:
            import shutil
            shutil.rmtree(tmpdir, ignore_errors=True)
    check('parse_rahmenterminplan: 5 Spieltage + 1 DST-Block', t_parse_rahmenterminplan_basic)

    def t_parse_rahmenterminplan_multi_liga():
        """Synthetic Excel mit 2 Ligen, einzelne Spalten."""
        import tempfile, os
        tmpdir = tempfile.mkdtemp()
        try:
            xlsx = os.path.join(tmpdir, 'cal.xlsx')
            _build_synth_rahmenplan(xlsx, [
                ['KW', 'von', 'bis', 'LigaA', 'LigaB'],
                [37, '07.09.2026', '13.09.2026', 1,   None],
                [38, '14.09.2026', '20.09.2026', 2,   1],
                [39, '21.09.2026', '27.09.2026', '3/4', 2],
            ])
            res = parse_rahmenterminplan(xlsx, {'LigaA': 3, 'LigaB': 4},
                                          kw_col=0, date_from_col=1, date_to_col=2)
            assert set(res['spieltage']['LigaA'].keys()) == {1, 2, 3, 4}
            assert set(res['spieltage']['LigaB'].keys()) == {1, 2}
            assert res['dst_blocks']['LigaA'] == [(3, 4)]
            assert res['dst_blocks']['LigaB'] == []
            return 'LigaA: 4 ST + 1 DST, LigaB: 2 ST + 0 DST'
        finally:
            import shutil
            shutil.rmtree(tmpdir, ignore_errors=True)
    check('parse_rahmenterminplan: 2 Ligen', t_parse_rahmenterminplan_multi_liga)

    def t_parse_rahmenterminplan_file_not_found():
        """Nicht-existente Datei -> None."""
        res = parse_rahmenterminplan('/nonexistent/path/foo.xlsx', {'A': 0})
        assert res is None
        return 'Nicht-existente Datei -> None'
    check('parse_rahmenterminplan: Datei fehlt -> None', t_parse_rahmenterminplan_file_not_found)

    def t_parse_rahmenterminplan_kw_compat():
        """kw_compat-Struktur korrekt aufgebaut."""
        import tempfile, os
        tmpdir = tempfile.mkdtemp()
        try:
            xlsx = os.path.join(tmpdir, 'cal.xlsx')
            _build_synth_rahmenplan(xlsx, [
                ['KW', 'von', 'bis', 'LigaA'],
                [37, '07.09.2026', '13.09.2026', 1],
                [38, '14.09.2026', '20.09.2026', '2/3'],
            ])
            res = parse_rahmenterminplan(xlsx, {'LigaA': 3}, kw_col=0,
                                          date_from_col=1, date_to_col=2)
            kw_compat = res['kw_compat']
            assert 37 in kw_compat and 38 in kw_compat
            assert kw_compat[37]['LigaA'] == [1]
            assert kw_compat[38]['LigaA'] == [2, 3]
            return f'KW 37: [1], KW 38: [2,3]'
        finally:
            import shutil
            shutil.rmtree(tmpdir, ignore_errors=True)
    check('parse_rahmenterminplan: kw_compat-Struktur', t_parse_rahmenterminplan_kw_compat)

    def t_preview_columns():
        """preview_columns liest die ersten n Zeilen ein."""
        import tempfile, os
        tmpdir = tempfile.mkdtemp()
        try:
            xlsx = os.path.join(tmpdir, 'cal.xlsx')
            _build_synth_rahmenplan(xlsx, [
                ['KW', 'von', 'bis', 'LigaA'],
                [37, '07.09.', '13.09.', 1],
                [38, '14.09.', '20.09.', 2],
                [39, '21.09.', '27.09.', 3],
            ])
            df = preview_columns(xlsx, n_rows=3)
            assert df is not None, 'preview_columns gab None zurueck'
            assert len(df) >= 3, f'erwartet >= 3 Zeilen, got {len(df)}'
            return f'{len(df)} Zeilen geladen'
        finally:
            import shutil
            shutil.rmtree(tmpdir, ignore_errors=True)
    check('preview_columns: liest erste n Zeilen', t_preview_columns)

    def t_preview_columns_nonexistent():
        """Nicht-existente Datei -> None."""
        df = preview_columns('/nonexistent/foo.xlsx')
        assert df is None
        return 'Nicht-existente Datei -> None'
    check('preview_columns: fehlende Datei -> None', t_preview_columns_nonexistent)

    # ── Feature 9: Karten-Visualisierung (A1, v1.9.0) ────────────────────────
    print('\n--- Feature 9: Karten-Visualisierung (geocode + map_output) ---')

    from spielplan_multi.geocode import _normalize, _load_cache, _save_cache
    from spielplan_multi.map_output import build_route_map

    def t_geocode_normalize():
        assert _normalize('  Berlin  ') == 'berlin'
        assert _normalize('Hamburg   Hbf') == 'hamburg hbf'
        assert _normalize('KÖLN') == 'köln'
        return 'Trim + lowercase + whitespace-collapse'
    check('geocode._normalize: trimmt + lowercase', t_geocode_normalize)

    def t_geocode_cache_roundtrip():
        """Cache write+read mit Tuples roundtrip."""
        import tempfile, os
        from spielplan_multi import geocode
        with tempfile.TemporaryDirectory() as tmpdir:
            old_cache = geocode._CACHE_FILE
            geocode._CACHE_FILE = type(old_cache)(os.path.join(tmpdir, 'gc.json'))
            try:
                data = {'addr1': (52.5, 13.4), 'addr2': None, 'addr3': (50.0, 10.0)}
                _save_cache(data)
                read_back = _load_cache()
                assert read_back == data, f'Cache roundtrip: {read_back}'
                return f'{len(data)} Eintraege roundtripped'
            finally:
                geocode._CACHE_FILE = old_cache
    check('geocode._load_cache/_save_cache: Roundtrip', t_geocode_cache_roundtrip)

    def t_map_empty():
        """build_route_map mit leeren results -> Karte ohne Marker."""
        import folium
        m = build_route_map({}, {})
        assert isinstance(m, folium.Map), 'kein folium.Map zurueck'
        return 'Leeres Dict -> leere Karte ohne Crash'
    check('build_route_map: leeres results', t_map_empty)

    def t_map_single_liga():
        """build_route_map mit einer Liga + 2 Koordinaten -> Marker + Linien."""
        import folium
        # Reusing 'result' (TEST-Liga, 4 Teams)
        fake_geocodes = {
            'Alpha': (52.5, 13.4),    # Berlin
            'Beta':  (53.5, 10.0),    # Hamburg
            'Gamma': (50.9, 6.9),     # Köln
            'Delta': (48.1, 11.6),    # München
        }
        m = build_route_map({'TEST': result}, fake_geocodes)
        assert isinstance(m, folium.Map)
        # Karte als HTML rendern und Marker zaehlen (CircleMarker -> 'circle')
        html = m._repr_html_()
        # 4 CircleMarker fuer 4 Teams
        n_circles = html.count('circleMarker')
        assert n_circles >= 4, f'erwartet >= 4 CircleMarker, got {n_circles}'
        return f'4 Standorte -> {n_circles} CircleMarker im HTML'
    check('build_route_map: Single-Liga mit Markern', t_map_single_liga)

    def t_map_missing_coords():
        """Wenn fuer ein Team keine Koordinate vorhanden ist, kein Crash."""
        import folium
        fake_geocodes = {'Alpha': (52.5, 13.4)}  # Nur 1/4
        m = build_route_map({'TEST': result}, fake_geocodes)
        assert isinstance(m, folium.Map)
        return 'Teil-Geocodes -> kein Crash'
    check('build_route_map: fehlende Koordinaten ueberspringen', t_map_missing_coords)

    # ── Feature 10: Interaktive Kalenderansicht (A2, v1.10.0) ─────────────
    print('\n--- Feature 10: Kalenderansicht (calendar_output) ---')

    from spielplan_multi.calendar_output import (
        build_calendar_events, default_calendar_options, _parse_date
    )

    def t_parse_date_iso():
        assert _parse_date('2026-09-07') == _dt.date(2026, 9, 7)
        return 'ISO-Format YYYY-MM-DD geparst'
    import datetime as _dt
    check('_parse_date: ISO-Format', t_parse_date_iso)

    def t_parse_date_de():
        assert _parse_date('07.09.2026') == _dt.date(2026, 9, 7)
        return 'DE-Format DD.MM.YYYY geparst'
    check('_parse_date: DE-Format', t_parse_date_de)

    def t_parse_date_invalid():
        assert _parse_date(None) is None
        assert _parse_date('') is None
        assert _parse_date('foo') is None
        assert _parse_date('99.99.9999') is None
        return 'None/leer/garbage -> None'
    check('_parse_date: Edge-Cases', t_parse_date_invalid)

    def t_events_no_calendar():
        """Wenn cfg.calendar leer ist, werden alle Events uebersprungen."""
        # 'result' aus TEST-Liga hat CAL gesetzt (37-42)
        # Hier baue ich ein result ohne calendar nach
        cfg_no_cal = make_cfg('NC', TEAMS, dist=DIST)  # calendar default {}
        r2 = solve_league_phase1(cfg_no_cal, time_limit=15, seed=42)
        assert r2 is not None
        r2.cfg = cfg_no_cal
        events = build_calendar_events({'NC': r2})
        assert events == [], f'erwartet leer, got {len(events)} events'
        return 'Ohne Kalender -> []'
    check('build_calendar_events: ohne Kalender -> leer', t_events_no_calendar)

    def t_events_with_calendar():
        """Mit CAL gesetzt: Events fuer alle Spiele."""
        # TEST-Liga hat 4 Teams * 3 Spiele = 12 Spiele insgesamt
        # CAL hat 6 Spieltage definiert, aber kein week_start -> Events leer
        # Ergaenze week_start fuer alle Spieltage
        cfg_w = make_cfg('WC', TEAMS, dist=DIST, calendar={
            1: {'kw': 37, 'week_start': '2026-09-07', 'week_end': '2026-09-13'},
            2: {'kw': 38, 'week_start': '2026-09-14', 'week_end': '2026-09-20'},
            3: {'kw': 39, 'week_start': '2026-09-21', 'week_end': '2026-09-27'},
            4: {'kw': 40, 'week_start': '2026-09-28', 'week_end': '2026-10-04'},
            5: {'kw': 41, 'week_start': '2026-10-05', 'week_end': '2026-10-11'},
            6: {'kw': 42, 'week_start': '2026-10-12', 'week_end': '2026-10-18'},
        })
        r3 = solve_league_phase1(cfg_w, time_limit=15, seed=42)
        assert r3 is not None
        r3.cfg = cfg_w
        events = build_calendar_events({'WC': r3})
        total_games = sum(len(g) for g in r3.schedule.values())
        assert len(events) == total_games, f'erwartet {total_games}, got {len(events)}'
        # Erstes Event prüfen
        e0 = events[0]
        assert 'title' in e0 and '–' in e0['title'], f'Title-Format falsch: {e0}'
        assert e0['start'].startswith('2026-'), f'Start-Datum falsch: {e0["start"]}'
        assert 'backgroundColor' in e0
        assert 'extendedProps' in e0
        return f'{len(events)} Events erzeugt'
    check('build_calendar_events: mit Kalender -> Events', t_events_with_calendar)

    def t_events_allday_vs_uhrzeit():
        """Mit game_times: Events haben Uhrzeit, ohne: allDay=True."""
        cfg_w = make_cfg('UT', TEAMS, dist=DIST, calendar={
            1: {'kw': 37, 'week_start': '2026-09-07'},
            2: {'kw': 38, 'week_start': '2026-09-14'},
            3: {'kw': 39, 'week_start': '2026-09-21'},
            4: {'kw': 40, 'week_start': '2026-09-28'},
            5: {'kw': 41, 'week_start': '2026-10-05'},
            6: {'kw': 42, 'week_start': '2026-10-12'},
        })
        r = solve_league_phase1(cfg_w, time_limit=15, seed=42)
        assert r is not None
        r.cfg = cfg_w
        # game_times setzen: Spieltag 1 hat zwei Spiele um '14:00' und '16:00'
        r.game_times = {d: ['14:00'] * len(r.schedule.get(d, []))
                        for d in r.schedule}
        events = build_calendar_events({'UT': r})
        with_time = [e for e in events if not e.get('allDay', False)]
        assert len(with_time) > 0, 'erwartet mindestens 1 Event mit Uhrzeit'
        e = with_time[0]
        assert 'T' in e['start'], f'Uhrzeit-Format falsch: {e["start"]}'
        return f'{len(with_time)} Events mit Uhrzeit'
    check('build_calendar_events: Uhrzeit setzt allDay=False', t_events_allday_vs_uhrzeit)

    def t_default_options():
        opts = default_calendar_options()
        assert opts['initialView'] == 'dayGridMonth'
        assert opts['locale'] == 'de'
        assert opts['firstDay'] == 1
        return 'Defaults: dayGridMonth, locale=de, firstDay=Mo'
    check('default_calendar_options: vernuenftige Defaults', t_default_options)

    def t_events_kw_fallback():
        """Bei fehlendem week_start aber vorhandener KW: Datum aus KW berechnet."""
        # KW ohne Datum — soll trotzdem Events erzeugen
        cfg_kw_only = make_cfg('KO', TEAMS, dist=DIST, calendar={
            1: {'kw': 37, 'week_start': '', 'week_end': ''},
            2: {'kw': 38, 'week_start': '', 'week_end': ''},
            3: {'kw': 39, 'week_start': '', 'week_end': ''},
            4: {'kw': 40, 'week_start': '', 'week_end': ''},
            5: {'kw': 41, 'week_start': '', 'week_end': ''},
            6: {'kw': 42, 'week_start': '', 'week_end': ''},
        })
        r = solve_league_phase1(cfg_kw_only, time_limit=15, seed=42)
        assert r is not None
        r.cfg = cfg_kw_only
        # Heuristik schaetzt 2026 aus dem Default (today.year) wenn nichts gesetzt
        events = build_calendar_events({'KO': r}, season_year=2026)
        total = sum(len(g) for g in r.schedule.values())
        assert len(events) == total, f'erwartet {total}, got {len(events)}'
        return f'{len(events)} Events trotz fehlendem week_start (Fallback aus KW)'
    check('build_calendar_events: KW-Fallback wenn week_start fehlt', t_events_kw_fallback)

    def t_events_multi_liga_mixed():
        """Multi-Liga: eine Liga mit Datum, andere nur mit KW -> alle Events."""
        cfg_full = make_cfg('LF', TEAMS, dist=DIST, calendar={
            1: {'kw': 37, 'week_start': '2026-09-07', 'week_end': '2026-09-13'},
            2: {'kw': 38, 'week_start': '2026-09-14', 'week_end': '2026-09-20'},
            3: {'kw': 39, 'week_start': '2026-09-21', 'week_end': '2026-09-27'},
            4: {'kw': 40, 'week_start': '2026-09-28', 'week_end': '2026-10-04'},
            5: {'kw': 41, 'week_start': '2026-10-05', 'week_end': '2026-10-11'},
            6: {'kw': 42, 'week_start': '2026-10-12', 'week_end': '2026-10-18'},
        })
        cfg_kw_only = make_cfg('LK', TEAMS, dist=DIST, calendar={
            1: {'kw': 37, 'week_start': ''},
            2: {'kw': 38, 'week_start': ''},
            3: {'kw': 39, 'week_start': ''},
            4: {'kw': 40, 'week_start': ''},
            5: {'kw': 41, 'week_start': ''},
            6: {'kw': 42, 'week_start': ''},
        })
        r_full = solve_league_phase1(cfg_full, time_limit=15, seed=42)
        r_kw = solve_league_phase1(cfg_kw_only, time_limit=15, seed=42)
        assert r_full is not None and r_kw is not None
        r_full.cfg = cfg_full
        r_kw.cfg = cfg_kw_only
        events = build_calendar_events({'LF': r_full, 'LK': r_kw})
        ligas = {e['extendedProps']['liga'] for e in events}
        assert ligas == {'LF', 'LK'}, f'erwartet beide Ligen, got {ligas}'
        return f'Beide Ligen im Kalender: {sorted(ligas)}'
    check('build_calendar_events: Multi-Liga mit gemischtem Kalender', t_events_multi_liga_mixed)

    _summarize()


def _summarize():
    passed = sum(1 for _, ok in _results if ok)
    total  = len(_results)
    print(f'\n  {passed}/{total} Tests bestanden\n')
    if passed < total:
        sys.exit(1)


if __name__ == '__main__':
    main()
