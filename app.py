"""Spielplan-Optimierer – Streamlit GUI

Wizard-basierte Oberfläche zur Spielplanerstellung für Floorball-Ligen.
Orchestriert die 3-Phasen-Pipeline aus spielplan_multi/.

Starten:
    streamlit run app.py
    oder: start.bat doppelklicken
"""
from __future__ import annotations

import io
import json
import math
import multiprocessing
import queue
import re
import sys
import tempfile
import threading
import time
import zipfile
from pathlib import Path
from typing import Dict, List, Optional

import numpy as np
import pandas as pd
import streamlit as st

# ── Pfad-Setup ────────────────────────────────────────────────────────────────
_HERE = Path(__file__).resolve().parent
if str(_HERE) not in sys.path:
    sys.path.insert(0, str(_HERE))

from spielplan_multi.config import WEIGHT_SCALES
from spielplan_multi.calendar_parser import build_weekends
from spielplan_multi.league_types import LeagueConfig, LeagueResult
from spielplan_multi.multi_solver import solve_all
from spielplan_multi.schedule_utils import (
    recompute_result_stats  as _recompute_result_stats_fn,
    swap_home_away          as _swap_home_away_fn,
    build_ics_bytes         as _build_ics_bytes_fn,
    build_print_html        as _build_print_html_fn,
    assign_game_times       as _assign_game_times_fn,
    move_game               as _move_game_fn,
    cancel_game             as _cancel_game_fn,
    reschedule_game         as _reschedule_game_fn,
    find_free_days          as _find_free_days_fn,
    find_schedule_warnings  as _find_schedule_warnings_fn,
)
from spielplan_multi.config_validator import validate as _validate_cfg


# ── Vereinsdatenbank ──────────────────────────────────────────────────────────
def _records_from_df(df: 'pd.DataFrame') -> List[dict]:
    """Konvertiert DataFrame → [{liga, verein, teamname, adresse}].
    Unterstützt neues Format (Spalte Teamname) und altes Format (Verein+Stadt)."""
    low = {c.lower(): c for c in df.columns}

    def _val(row, *names) -> str:
        for n in names:
            if n in low:
                v = str(row[low[n]]).strip()
                if v and v.lower() != 'nan':
                    return v
        return ''

    if 'teamname' in low:
        records = []
        for _, row in df.iterrows():
            tm = _val(row, 'teamname')
            if not tm:
                continue
            records.append({
                'liga':     _val(row, 'liga', 'ligabezeichnung'),
                'verein':   _val(row, 'verein', 'club'),
                'teamname': tm,
                'adresse':  _val(row, 'adresse', 'adress', 'stadt', 'city', 'standort', 'ort'),
            })
        return records

    if 'verein' in low:
        records = []
        for _, row in df.iterrows():
            v = _val(row, 'verein')
            if not v:
                continue
            records.append({
                'liga': '', 'verein': v, 'teamname': v,
                'adresse': _val(row, 'adresse', 'stadt', 'city', 'standort', 'ort'),
            })
        return records

    return []


@st.cache_data
def _load_club_db_file(path: str) -> List[dict]:
    """Lädt Vereinsdatenbank (CSV oder Excel) → [{liga, verein, teamname, adresse}, ...]"""
    try:
        p = Path(path)
        if p.suffix.lower() in ('.xlsx', '.xls'):
            df = pd.read_excel(path, dtype=str).fillna('')
        else:
            df = pd.read_csv(path, encoding='utf-8-sig', dtype=str).fillna('')
        df.columns = [c.strip() for c in df.columns]
        return _records_from_df(df)
    except Exception:
        pass
    return []


def load_club_db() -> List[dict]:
    """Gibt die aktive Vereinsdatenbank zurück (Bundle + Session-Ergänzungen)."""
    base  = _load_club_db_file(str(_HERE / 'clubs_db.csv'))
    extra = st.session_state.get('extra_clubs', [])
    if isinstance(extra, dict):
        extra = [{'liga': '', 'verein': k, 'teamname': k, 'adresse': v}
                 for k, v in extra.items()]
    return base + extra


def _load_excel_safe(
    uploaded_file,
    sheet_name: Optional[str] = None,
    nrows: Optional[int] = None,
    label: str = 'Excel-Datei',
) -> Optional[pd.DataFrame]:
    """Liest eine Excel-Datei ein und zeigt bei Fehler st.error(). Gibt None bei Fehler zurück."""
    try:
        df = pd.read_excel(uploaded_file, sheet_name=sheet_name, dtype=str, nrows=nrows).fillna('')
        df.columns = [c.strip() for c in df.columns]
        return df
    except zipfile.BadZipFile:
        st.error(f'{label}: Datei ist beschädigt oder kein gültiges Excel-Format (.xlsx).')
    except ValueError as exc:
        st.error(f'{label}: Sheet nicht gefunden oder ungültiges Format: {exc}')
    except Exception as exc:
        st.error(f'{label}: Lesefehler: {exc}')
    return None


def _parse_club_upload(uploaded_file) -> List[dict]:
    """Liest Excel (Liga/Verein/Teamname/Adresse) oder CSV (Verein/Stadt) → Datensatz-Liste."""
    fname = getattr(uploaded_file, 'name', '')
    df: Optional[pd.DataFrame] = None
    if fname.lower().endswith(('.xlsx', '.xls')):
        df = _load_excel_safe(uploaded_file, nrows=2000, label='Vereinsdatei')
    else:
        try:
            df = pd.read_csv(uploaded_file, encoding='utf-8-sig', dtype=str, nrows=2000).fillna('')
            df.columns = [c.strip() for c in df.columns]
        except Exception as exc:
            st.error(f'Fehler beim Lesen der Vereinsdatei: {exc}')
            return []
    if df is None:
        return []
    return _records_from_df(df)

# ── Seitenkonfiguration ───────────────────────────────────────────────────────
st.set_page_config(
    page_title='Spielplan-Optimierer',
    page_icon=str(_HERE / 'assets' / 'floorball_icon2.png'),
    layout='wide',
    initial_sidebar_state='expanded',
)

# ── Konstanten ────────────────────────────────────────────────────────────────
STEPS = [
    'Ligen & Teams',
    'Distanzmatrizen',
    'Kalender & DST',
    'Routing & Gewichte',
    'Pflichtspiele',
    'Sperrtage',
    'Co-Home',
    'Solver',
    'Optimierung & Ergebnisse',
]

FMT_OPTIONS = ['Hin-/Rückrunde', 'Einfachrunde', 'Dreifachrunde', 'Turniertag']
FMT_MAP = {
    'Hin-/Rückrunde': (2, 1),
    'Einfachrunde':   (1, 1),
    'Dreifachrunde':  (3, 1),
    # Turniertag: n_rounds und gpd werden pro Liga aus dem State gelesen
}
# Rückwärtskompatibilität: alter Formatname
_FMT_ALIAS = {'Turniertag (2 Spiele/Tag)': 'Turniertag'}


def _get_n_rounds_gpd(ld: dict) -> tuple:
    """Gibt (n_rounds, gpd) für ein Liga-Dict zurück."""
    fmt = _FMT_ALIAS.get(ld.get('fmt', ''), ld.get('fmt', 'Hin-/Rückrunde'))
    if fmt in FMT_MAP:
        return FMT_MAP[fmt]
    # Turniertag: aus Liga-Dict lesen
    return int(ld.get('tt_rounds', 2)), int(ld.get('gpd', 2))


def _valid_K_values(n_teams: int, gpd: int) -> List[int]:
    """Gültige Gruppengrößen K für n_teams und gpd:
    K|n, (K-1)%gpd==0, K*gpd gerade, gpd <= K-1."""
    return [K for K in range(gpd + 1, n_teams + 1)
            if n_teams % K == 0 and (K - 1) % gpd == 0
            and K * gpd % 2 == 0]


def _gpd_compatible_team_counts(gpd: int) -> List[int]:
    """Teamanzahlen die mit gpd Spielen/Tag funktionieren: (n-1)%gpd==0 und n*gpd gerade."""
    return [n for n in range(gpd + 1, 30)
            if (n - 1) % gpd == 0 and (n * gpd) % 2 == 0]


def _calc_n_matchdays(ld: dict) -> int:
    """Berechnet n_matchdays aus dem Liga-Dict (für Anzeige vor Config-Bau)."""
    n_rounds, gpd = _get_n_rounds_gpd(ld)
    n = len(ld.get('teams', []))
    if n < 2:
        return 0
    if ld.get('fmt') == 'Turniertag':
        apd = int(ld.get('active_per_day', n))
        if 0 < apd < n and apd >= gpd + 1 and apd * gpd % 2 == 0:
            total_matches = n_rounds * n * (n - 1) // 2
            games_per_day = apd * gpd // 2
            return total_matches // games_per_day if games_per_day > 0 else 0
    return n_rounds * (n - 1) // max(1, gpd)

def _detect_dst_blocks(rows: list) -> list:
    """DST-Blöcke aus Kalender-Tabelle ableiten: aufeinanderfolgende Spieltag-Paare in gleicher KW."""
    kw_to_sts: dict = {}
    for row in rows:
        kw = row.get('kw')
        if kw is not None:
            kw_to_sts.setdefault(kw, []).append(row['spieltag'])
    blocks = []
    for sts in kw_to_sts.values():
        if len(sts) == 2:
            d1, d2 = sorted(sts)
            if d2 == d1 + 1:
                blocks.append((d1, d2))
    return sorted(blocks)


WEIGHT_LABELS = {
    'switch':    ('Heimrechtswechsel',   'Wie oft wechselt ein Team zwischen Heim- und Auswärtsspielen. Höherer Wert = abwechslungsreichere Spielfolge (z.B. Heim–Auswärts–Heim statt drei Heimspiele hintereinander).'),
    'sw_fair':   ('Wechsel-Fairness',    'Wie gleichmäßig die Wechselhäufigkeit über alle Teams verteilt ist. Höherer Wert = kein Team hat deutlich mehr oder weniger Wechsel als die anderen.'),
    'travel':    ('Reisedistanz',        'Gesamte Fahrtstrecke aller Teams über die Saison. Höherer Wert = kürzere Gesamtkilometer werden stärker bevorzugt.'),
    'trav_fair': ('Reise-Fairness',      'Wie gleichmäßig die Reisebelastung auf alle Teams verteilt ist. Höherer Wert = kein Team muss deutlich mehr fahren als die anderen.'),
    'dst_eff':   ('DST-Reiseeffizienz',  'Bevorzugt Doppelwochenenden, bei denen beide Auswärtsspiele räumlich nah beieinander liegen. Teams aus Randlagen (z.B. Hamburg, München) profitieren am meisten: ihre weit entfernten Auswärtsspiele werden in einem DST gebündelt statt getrennt angesetzt. 0 = aus · 5 = empfohlen'),
}

# ── Session-State initialisieren ──────────────────────────────────────────────
_DEFAULTS: dict = dict(
    step=0,
    max_step=0,
    _wizard_started=False,
    league_order=[],      # [lid, ...]
    leagues={},           # {lid: {name, teams:[(nm,loc)], fmt, hw}}
    dist_method='manual',
    api_key='',
    dist_matrices={},     # {lid: np.ndarray}
    use_calendar=False,
    cal_path=None,
    dst_per_liga={},      # {lid: [(d1,d2),...]}
    kw_compat={},
    routing={},           # {lid: (apply:bool, pct:int)}
    same_weights=True,
    weights={},           # {lid: {key: float}}
    w_cohome=5.0,
    pinned={},            # {lid: [{'teamA','teamB','day','home'}]}
    blocked={},           # {lid: {team:[days]}}
    forced_home={},       # {lid: {team:[days]}}
    clubs={},             # {club_name: {lid: team}}
    team_verein_map={},   # {teamname: verein} – beim Hinzufügen aus DB befüllt
    solver=dict(p1=900, p2=5400, nm=False, seeds=2, sa=120),
    opt_running=False,
    opt_done=False,
    opt_log=[],
    opt_warnings=[],      # [!!]-Zeilen aus dem Log, persistent nach Abschluss
    results=None,
    excel_bytes={},       # {lid: bytes}
    cohome_bytes=None,
    hall_bytes=None,      # bytes – ligaübergreifender Hallenbelegungsplan
    opt_queue=None,
    opt_process=None,
    opt_result_holder=None,
    time_templates={},      # {lid: 'HH:MM, HH:MM, ...'} – Uhrzeiten-Vorlage je Liga
    opt_best={},            # {lid: {'obj': float, 'elapsed': str, 'count': int}}
    opt_start_time=None,    # float – time.time() beim Start der Optimierung
    move_pending=None,      # {'lid','day','idx','ht','at'} – Spiel zum Verschieben gewählt
    cancel_pending=None,    # {'lid','ht','at'} – ausgefallenes Spiel, Nachholtermin offen
    cal_table={},           # {lid: [{'spieltag':int,'kw':int|None,'date':str}]}
)

for _k, _v in _DEFAULTS.items():
    if _k not in st.session_state:
        st.session_state[_k] = _v

S = st.session_state


# ── Sitzung-laden-Dialog ──────────────────────────────────────────────────────
@st.dialog('Gespeicherte Sitzung laden')
def _show_load_session_dialog():
    st.caption(
        'Lade eine zuvor gespeicherte Sitzungs-Datei (.json), um direkt zur '
        'Ergebnisansicht zu springen – inkl. Spielzeiten zuweisen, '
        'Spiele verschieben und Absagen & Nachholspiele.'
    )
    _up = st.file_uploader(
        'Sitzungs-Datei (.json)',
        type=['json'],
        key='sess_dlg_upload',
        label_visibility='collapsed',
    )
    if _up is not None:
        with st.spinner('Sitzung wird geladen…'):
            _err = _session_from_json(_up.getvalue())
        if _err:
            st.error(f'Fehler beim Laden: {_err}')
        else:
            S.step     = 8
            S.max_step = 8
            st.rerun()


# ── Backlog-Dialog ────────────────────────────────────────────────────────────
@st.dialog('Funktionswunsch / Fehler melden')
def _show_backlog_dialog():
    import urllib.parse as _up

    # Bereits vorbereitete E-Mail anzeigen
    if st.session_state.get('_backlog_mailto'):
        st.success('E-Mail vorbereitet – klicke auf den Button, um sie in deinem E-Mail-Programm zu öffnen.')
        st.link_button('📧 E-Mail jetzt senden', st.session_state['_backlog_mailto'],
                       type='primary', width='stretch')
        st.caption('Die Nachricht ist vorausgefüllt und geht direkt an das FD-Team.')
        if st.button('Neue Meldung eingeben'):
            del st.session_state['_backlog_mailto']
            st.rerun()
        return

    st.caption(
        'Nach dem Ausfüllen wird eine vorbereitete E-Mail an das FD-Team geöffnet. '
        'Du musst sie nur noch absenden.'
    )
    with st.form('backlog_form', clear_on_submit=True):
        _bl_typ = st.selectbox(
            'Typ *',
            ['Neue Funktion', 'Verbesserung', 'Fehler / Bug'],
        )
        _bl_bereich = st.selectbox(
            'Bereich *',
            ['Spielplan-Optimierung', 'Excel-Export', 'Konfiguration (Wizard)',
             'Co-Home / Mehrspartenverein', 'Kalender / Termine',
             'Distanzen / Karte', 'Sonstiges'],
        )
        _bl_wichtig = st.selectbox(
            'Wichtigkeit *',
            ['Kleiner Wunsch', 'Wichtig für den Alltag', 'Blocker – kann nicht arbeiten'],
        )
        _bl_titel = st.text_input(
            'Titel * (ein Satz)',
            max_chars=120,
            placeholder='z. B. "Spielplan als PDF exportieren"',
        )
        _bl_beschr = st.text_area(
            'Beschreibung * – Was soll passieren? Was ist das Problem?',
            height=120,
            placeholder=(
                'Schritt in dem das Problem auftritt, was erwartet wird, '
                'was stattdessen passiert …'
            ),
        )
        _bl_kontakt = st.text_input(
            'Kontakt (optional) – E-Mail für Rückfragen',
            placeholder='name@verein.de',
        )
        _bl_submit = st.form_submit_button('E-Mail vorbereiten', type='primary')

    if _bl_submit:
        if not _bl_titel.strip() or not _bl_beschr.strip():
            st.error('Bitte Titel und Beschreibung ausfüllen.')
        else:
            _subject = f'[Spielplan-Optimierer] {_bl_typ} – {_bl_titel.strip()}'
            _body_parts = [
                f'Typ: {_bl_typ}',
                f'Bereich: {_bl_bereich}',
                f'Wichtigkeit: {_bl_wichtig}',
                '',
                f'Titel: {_bl_titel.strip()}',
                '',
                'Beschreibung:',
                _bl_beschr.strip(),
            ]
            if _bl_kontakt.strip():
                _body_parts += ['', f'Kontakt: {_bl_kontakt.strip()}']
            _body = '\r\n'.join(_body_parts)
            st.session_state['_backlog_mailto'] = (
                f'mailto:it@floorball.de'
                f'?subject={_up.quote(_subject)}'
                f'&body={_up.quote(_body)}'
            )
            st.rerun()


# ── Sidebar ───────────────────────────────────────────────────────────────────
_LOGO_PATH = str(_HERE / 'assets' / 'floorball_logo.png')

def _clubs_excel_bytes(clubs: list) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    wb = Workbook()
    ws = wb.active
    ws.title = 'Vereinsdatenbank'
    headers = ['Liga', 'Verein', 'Teamname', 'Adresse']
    hdr_fill = PatternFill('solid', fgColor='1F4E79')
    hdr_font = Font(bold=True, color='FFFFFF')
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(1, ci, h)
        cell.fill = hdr_fill
        cell.font = hdr_font
    for r in clubs:
        ws.append([
            r.get('liga', ''),
            r.get('verein', r.get('teamname', '')),
            r.get('teamname', ''),
            r.get('adresse', ''),
        ])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = max(
            (len(str(c.value or '')) for c in col), default=10) + 4
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _sidebar():
    with st.sidebar:
        st.image(_LOGO_PATH, width='content')
        st.markdown('### Spielplan-Optimierer')
        st.caption('Automatische Spielplanerstellung')

        # Übersicht-Link + Wizard-Menü
        if S._wizard_started:
            st.divider()
            if st.button('← Zur Übersicht', key='nav_intro', width='stretch'):
                S._wizard_started = False
                st.rerun()
            # max_step mitführen: höchster je erreichter Schritt
            if S.step > S.max_step:
                S.max_step = S.step
            st.divider()
            for i, label in enumerate(STEPS):
                if i == S.step:
                    # Aktueller Schritt: fett, nicht klickbar
                    st.markdown(f'**→ {i+1}. {label}**')
                elif i <= S.max_step:
                    # Bereits besucht (vor oder nach aktuellem Schritt): klickbar
                    _icon = '✓' if i < S.step else '○'
                    _btn_key = f'nav_{i}'
                    if not S.opt_running:
                        if st.button(f'{_icon} {i+1}. {label}', key=_btn_key,
                                     width='stretch'):
                            S.step = i
                            st.rerun()
                    else:
                        st.markdown(f'{_icon} {i+1}. {label}')
                else:
                    # Noch nicht erreicht: grau, nicht klickbar
                    st.markdown(f'○ {i+1}. {label}')

        st.divider()
        with st.expander('🗂 Vereinsdatenbank', expanded=False):
            _all_clubs = load_club_db()
            st.caption(f'{len(_all_clubs)} Einträge geladen.')

            # ── Datei hochladen ──────────────────────────────────────────────
            up = st.file_uploader('Eigene Teamliste (Excel/CSV)',
                type=['xlsx', 'xls', 'csv'],
                key='club_db_upload',
                help='Excel mit Spalten: Liga, Verein, Teamname, Adresse\n'
                     '(CSV-Altformat: Verein, Stadt)')
            if up:
                extra = _parse_club_upload(up)
                if extra:
                    st.session_state.extra_clubs = extra
                    st.success(f'{len(extra)} Teams importiert.')
                    _load_club_db_file.clear()
                    st.rerun()
                else:
                    st.error('Datei konnte nicht gelesen werden – benötigt Spalten "Teamname" und "Adresse".')

            # ── Eintrag manuell hinzufügen ───────────────────────────────────
            with st.form('add_club_form', clear_on_submit=True):
                st.caption('Eintrag manuell hinzufügen')
                _c1, _c2 = st.columns(2)
                _new_verein = _c1.text_input('Verein', key='new_club_verein')
                _new_stadt  = _c2.text_input('Stadt', key='new_club_stadt')
                _new_addr   = st.text_input('Adresse (optional)', key='new_club_addr')
                if st.form_submit_button('Hinzufügen'):
                    if _new_verein.strip():
                        _extra_now = st.session_state.get('extra_clubs', [])
                        if isinstance(_extra_now, dict):
                            _extra_now = [{'liga': '', 'verein': k, 'teamname': k,
                                           'adresse': v} for k, v in _extra_now.items()]
                        _extra_now.append({
                            'liga': '', 'verein': _new_verein.strip(),
                            'teamname': _new_verein.strip(),
                            'adresse': f'{_new_stadt.strip()}  {_new_addr.strip()}'.strip(),
                        })
                        st.session_state.extra_clubs = _extra_now
                        st.rerun()

            # ── Aktuelle Liste als Excel herunterladen ───────────────────────
            st.download_button(
                'Aktuelle Liste als Excel herunterladen',
                data=_clubs_excel_bytes(_all_clubs),
                file_name='clubs_db.xlsx',
                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                key='dl_clubs_xlsx',
                help='Excel-Datei mit allen Einträgen (Liga / Verein / Teamname / Adresse). '
                     'Kann direkt wieder hochgeladen werden.',
            )

            _extra = st.session_state.get('extra_clubs', [])
            if _extra:
                n_extra = len(_extra) if isinstance(_extra, list) else len(_extra)
                st.caption(f'+ {n_extra} eigene Einträge aktiv.')
                if st.button('Eigene Einträge entfernen', key='clear_extra'):
                    st.session_state.extra_clubs = []
                    st.rerun()
        st.caption('v1.1 · Spielplan-Optimierer')
        if st.button('💾 Gespeicherte Sitzung laden', width='stretch'):
            _show_load_session_dialog()
        if st.button('📋 Funktionswunsch / Fehler melden', width='stretch'):
            _show_backlog_dialog()


# ── Navigations-Buttons ───────────────────────────────────────────────────────
def _nav(back=True, fwd_label='Weiter →', fwd_disabled=False):
    c1, c2, c3 = st.columns([2, 4, 2])
    go_back = go_fwd = False
    with c1:
        if back and S.step > 0 and not S.opt_running:
            if st.button('← Zurück', key=f'back_{S.step}'):
                go_back = True
    with c2:
        if S.league_order:
            _xl = _full_config_excel_bytes()
            st.download_button(
                '⬇ Konfiguration speichern',
                data=_xl,
                file_name='Spielplan_Konfiguration.xlsx',
                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                key=f'dl_cfg_{S.step}',
                width='stretch',
            )
    with c3:
        if st.button(fwd_label, key=f'fwd_{S.step}', type='primary',
                     disabled=fwd_disabled):
            go_fwd = True
    return go_back, go_fwd


# ═══════════════════════════════════════════════════════════════════════════════
# SCHRITT 0 – Ligen & Teams
# ═══════════════════════════════════════════════════════════════════════════════
def _teams_excel_bytes(leagues: dict, league_order: list) -> bytes:
    """Erstellt eine formatierte Excel-Vorlage mit den aktuellen Ligen/Teams."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = 'Ligen & Teams'

    # Farben
    HDR_FILL  = PatternFill('solid', fgColor='1F4E79')
    ALT_FILL  = PatternFill('solid', fgColor='D6E4F0')
    NONE_FILL = PatternFill('solid', fgColor='FFFFFF')
    thin = Side(style='thin', color='AAAAAA')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ['Liga-ID', 'Ligabezeichnung', 'Spielformat', 'Teamname', 'Standort']
    col_w   = [14, 28, 30, 30, 30]

    # Kopfzeile
    for col, (h, w) in enumerate(zip(headers, col_w), 1):
        cell = ws.cell(1, col, h)
        cell.font      = Font(bold=True, color='FFFFFF', size=11)
        cell.fill      = HDR_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border    = border
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = 'A2'

    # Daten
    row_num = 2
    for liga_idx, lid in enumerate(league_order):
        ld     = leagues.get(lid, {})
        teams  = ld.get('teams', []) or [('', '')]
        fill   = ALT_FILL if liga_idx % 2 == 0 else NONE_FILL
        for tname, city in teams:
            for col, val in enumerate([
                lid,
                ld.get('name', lid),
                ld.get('fmt', FMT_OPTIONS[0]),
                tname,
                city,
            ], 1):
                cell = ws.cell(row_num, col, val)
                cell.fill      = fill
                cell.border    = border
                cell.alignment = Alignment(vertical='center')
            row_num += 1

    # Dropdown-Validierung für Spielformat (Spalte C, ab Zeile 2)
    dv = DataValidation(
        type='list',
        formula1=f'"{",".join(FMT_OPTIONS)}"',
        allow_blank=False,
        showDropDown=False,
        showErrorMessage=True,
        errorTitle='Ungültiges Format',
        error=f'Bitte wähle: {", ".join(FMT_OPTIONS)}',
    )
    ws.add_data_validation(dv)
    dv.sqref = f'C2:C{max(row_num, 200)}'

    # Hinweis-Blatt
    ws2 = wb.create_sheet('Hinweise')
    ws2.column_dimensions['A'].width = 60
    hints = [
        ('Hinweise zur Vorlage', True),
        ('', False),
        ('Liga-ID', True),
        ('  Kurzkürzel ohne Leerzeichen, z. B. BL1, BL2, REGL', False),
        ('  Alle Zeilen einer Liga müssen dieselbe Liga-ID haben.', False),
        ('', False),
        ('Ligabezeichnung', True),
        ('  Vollständiger Name, z. B. 1. Bundesliga Herren', False),
        ('', False),
        ('Spielformat', True),
        *[(f'  • {f}', False) for f in FMT_OPTIONS],
        ('  Dropdown in Spalte C auswählen.', False),
        ('', False),
        ('Teamname', True),
        ('  Name des Teams im Spielplan, z. B. ETV Hamburg Damen', False),
        ('', False),
        ('Standort', True),
        ('  Stadt oder vollständige Adresse (für Distanzberechnung).', False),
        ('  Beispiel: Hamburg  oder  Tarpenbekstr. 25, 20251 Hamburg', False),
    ]
    for r, (text, bold) in enumerate(hints, 1):
        cell = ws2.cell(r, 1, text)
        if bold:
            cell.font = Font(bold=True, size=11, color='1F4E79')
        else:
            cell.font = Font(size=10)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _load_teams_excel(uploaded_file) -> Optional[dict]:
    """Liest die Excel-Vorlage und gibt {league_order, leagues} zurück."""
    df = _load_excel_safe(uploaded_file, sheet_name='Ligen & Teams', nrows=500, label='Konfigurationsdatei')
    if df is None:
        return None
    try:
        if 'Spielformat' in df.columns and 'Format' not in df.columns:
            df.rename(columns={'Spielformat': 'Format'}, inplace=True)
        required = {'Liga-ID', 'Teamname'}
        missing = required - set(df.columns)
        if missing:
            st.error(f'Fehlende Spalten in der Excel-Datei: {", ".join(sorted(missing))}')
            return None
        league_order: List[str] = []
        leagues: Dict[str, dict] = {}
        for _, row in df.iterrows():
            lid = str(row['Liga-ID']).strip()
            if not lid or lid.lower() == 'nan':
                continue
            if lid not in leagues:
                league_order.append(lid)
                fmt = str(row.get('Format', FMT_OPTIONS[0])).strip()
                fmt = _FMT_ALIAS.get(fmt, fmt)
                if fmt not in FMT_OPTIONS:
                    fmt = FMT_OPTIONS[0]
                try:
                    hw = float(str(row.get('Hierarchiegewicht', 1.0) or 1.0))
                except Exception:
                    hw = 1.0
                ld_new: dict = dict(name=str(row.get('Ligabezeichnung', lid)).strip() or lid,
                                    teams=[], fmt=fmt, hw=hw)
                for col, key in [('TT-Runden', 'tt_rounds'), ('Spiele/Tag', 'gpd'),
                                  ('Teams/Gruppe', 'k_group'), ('Anwesende/Tag', 'active_per_day')]:
                    if col in df.columns:
                        v = str(row.get(col, '')).strip()
                        if v and v.lower() != 'nan':
                            try:
                                ld_new[key] = int(float(v))
                            except Exception:
                                pass
                leagues[lid] = ld_new
            tname = str(row['Teamname']).strip()
            city  = str(row.get('Standort', '')).strip()
            if tname and tname.lower() != 'nan':
                leagues[lid]['teams'].append((tname, city if city.lower() != 'nan' else ''))
        if not league_order:
            st.error('Keine Ligen gefunden. Bitte prüfen, ob die Liga-ID-Spalte korrekt befüllt ist.')
            return None
        return {'league_order': league_order, 'leagues': leagues}
    except Exception as exc:
        st.error(f'Fehler beim Verarbeiten der Excel-Datei: {exc}')
        return None


def _full_config_excel_bytes() -> bytes:
    """Exportiert die komplette Optimizer-Konfiguration als Excel (alle Schritte)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = Workbook()
    HDR  = PatternFill('solid', fgColor='1F4E79')
    HDR2 = PatternFill('solid', fgColor='2E75B6')
    ALT  = PatternFill('solid', fgColor='D6E4F0')
    WHT  = PatternFill('solid', fgColor='FFFFFF')
    thin   = Side(style='thin', color='AAAAAA')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _h(ws, r, c, val, fill=None):
        cell = ws.cell(r, c, val)
        cell.font      = Font(bold=True, color='FFFFFF', size=10)
        cell.fill      = fill or HDR
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border    = border
        return cell

    def _d(ws, r, c, val, fill=None):
        cell = ws.cell(r, c, val)
        cell.fill      = fill or WHT
        cell.border    = border
        cell.alignment = Alignment(vertical='center')
        return cell

    def _set_col_w(ws, widths):
        for col, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = w

    # ── Sheet 1: Ligen & Teams ────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = 'Ligen & Teams'
    hdrs1 = ['Liga-ID', 'Ligabezeichnung', 'Spielformat', 'Teamname', 'Standort',
             'Hierarchiegewicht', 'TT-Runden', 'Spiele/Tag', 'Teams/Gruppe', 'Anwesende/Tag']
    _set_col_w(ws1, [14, 28, 30, 30, 30, 18, 12, 12, 14, 14])
    for col, h in enumerate(hdrs1, 1):
        _h(ws1, 1, col, h)
    ws1.row_dimensions[1].height = 20
    ws1.freeze_panes = 'A2'
    row = 2
    for li, lid in enumerate(S.league_order):
        ld    = S.leagues.get(lid, {})
        teams = ld.get('teams', []) or [('', '')]
        fill  = ALT if li % 2 == 0 else WHT
        for tname, city in teams:
            for col, val in enumerate([
                lid, ld.get('name', lid), ld.get('fmt', FMT_OPTIONS[0]),
                tname, city, ld.get('hw', 1.0),
                ld.get('tt_rounds', '') or '', ld.get('gpd', '') or '', ld.get('k_group', '') or '',
                ld.get('active_per_day', '') or '',
            ], 1):
                _d(ws1, row, col, val, fill)
            row += 1
    dv = DataValidation(type='list', formula1=f'"{",".join(FMT_OPTIONS)}"',
                        allow_blank=False, showDropDown=False, showErrorMessage=True,
                        errorTitle='Ungültiges Format', error=f'Bitte wähle: {", ".join(FMT_OPTIONS)}')
    ws1.add_data_validation(dv)
    dv.sqref = f'C2:C{max(row, 200)}'

    # ── Sheet 2: Einstellungen ────────────────────────────────────────────────
    ws_e = wb.create_sheet('Einstellungen')
    _set_col_w(ws_e, [28, 20])
    _h(ws_e, 1, 1, 'Einstellung'); _h(ws_e, 1, 2, 'Wert')
    for r, (k, v) in enumerate([
        ('dist_method',   S.dist_method),
        ('same_weights',  'J' if S.same_weights else 'N'),
        ('w_cohome',      S.w_cohome),
        ('solver_p1',     S.solver.get('p1',    900)),
        ('solver_p2',     S.solver.get('p2',  5400)),
        ('solver_sa',     S.solver.get('sa',   120)),
        ('solver_seeds',  S.solver.get('seeds',  2)),
    ], 2):
        _d(ws_e, r, 1, k); _d(ws_e, r, 2, v)

    # ── Sheet 3: Distanzmatrizen ──────────────────────────────────────────────
    ws_d = wb.create_sheet('Distanzmatrizen')
    d_row = 1
    for lid in S.league_order:
        mat  = S.dist_matrices.get(lid)
        ld   = S.leagues.get(lid, {})
        if mat is None or np.sum(mat) == 0:
            continue
        tms = [t for t, _ in ld.get('teams', [])]
        n   = len(tms)
        if n == 0:
            continue
        c = ws_d.cell(d_row, 1, f'Liga: {lid} – {ld.get("name", "")}')
        c.font = Font(bold=True, size=11, color='1F4E79')
        d_row += 1
        for j, t in enumerate(tms, 2):
            _h(ws_d, d_row, j, t, HDR2)
        for i, trow in enumerate(tms):
            r = d_row + 1 + i
            _h(ws_d, r, 1, trow, HDR2)
            fill = ALT if i % 2 == 0 else WHT
            for j, val in enumerate(mat[i], 2):
                _d(ws_d, r, j, round(float(val), 1), fill)
        ws_d.column_dimensions[get_column_letter(1)].width = 26
        for j in range(2, n + 2):
            ws_d.column_dimensions[get_column_letter(j)].width = 12
        d_row += n + 3

    # ── Sheet 4: Gewichte ─────────────────────────────────────────────────────
    ws_w = wb.create_sheet('Gewichte')
    hdrs_w = ['Liga-ID', 'Heimrechtwechsel', 'Wechsel-Fairness', 'Reisedistanz', 'Reise-Fairness', 'DST-Reiseeffizienz']
    _set_col_w(ws_w, [18, 20, 20, 16, 16, 20])
    for col, h in enumerate(hdrs_w, 1):
        _h(ws_w, 1, col, h)
    wr = 2
    if S.same_weights:
        common = S.weights.get('__common__', {})
        _d(ws_w, wr, 1, '__common__')
        for col, k in enumerate(['switch', 'sw_fair', 'travel', 'trav_fair', 'dst_eff'], 2):
            _d(ws_w, wr, col, common.get(k, 5.0) if k != 'dst_eff' else common.get(k, 0.0))
        wr += 1
    else:
        for lid in S.league_order:
            w = S.weights.get(lid, {})
            _d(ws_w, wr, 1, lid)
            for col, k in enumerate(['switch', 'sw_fair', 'travel', 'trav_fair', 'dst_eff'], 2):
                _d(ws_w, wr, col, w.get(k, 5.0) if k != 'dst_eff' else w.get(k, 0.0))
            wr += 1

    # ── Sheet 5: Kalender ─────────────────────────────────────────────────────
    ws_cal = wb.create_sheet('Kalender')
    _set_col_w(ws_cal, [16, 12, 8, 20])
    for col, h in enumerate(['Liga-ID', 'Spieltag', 'KW', 'Datum'], 1):
        _h(ws_cal, 1, col, h)
    cal_r = 2
    for li, lid in enumerate(S.league_order):
        rows_cal = S.cal_table.get(lid, [])
        for row_cal in rows_cal:
            kw_val   = row_cal.get('kw')
            date_val = row_cal.get('date', '')
            if kw_val is None and not date_val:
                continue
            fill = ALT if li % 2 == 0 else WHT
            _d(ws_cal, cal_r, 1, lid,                    fill)
            _d(ws_cal, cal_r, 2, row_cal['spieltag'],    fill)
            _d(ws_cal, cal_r, 3, kw_val if kw_val is not None else '', fill)
            _d(ws_cal, cal_r, 4, date_val,               fill)
            cal_r += 1

    # ── Sheet 7: Routing ──────────────────────────────────────────────────────
    ws_rt = wb.create_sheet('Routing')
    _set_col_w(ws_rt, [16, 22, 20])
    for col, h in enumerate(['Liga-ID', 'Routing aktiv (J/N)', 'Mehrkilometer (%)'], 1):
        _h(ws_rt, 1, col, h)
    rr = 2
    for lid in S.league_order:
        apply_, pct = S.routing.get(lid, (False, 25))
        _d(ws_rt, rr, 1, lid)
        _d(ws_rt, rr, 2, 'J' if apply_ else 'N')
        _d(ws_rt, rr, 3, pct)
        rr += 1

    # ── Sheet 7: Pflichtspiele ────────────────────────────────────────────────
    ws_p = wb.create_sheet('Pflichtspiele')
    _set_col_w(ws_p, [16, 26, 26, 12, 26])
    for col, h in enumerate(['Liga-ID', 'Team A', 'Team B', 'Spieltag', 'Heimrecht'], 1):
        _h(ws_p, 1, col, h)
    pr = 2
    for lid in S.league_order:
        for pm in S.pinned.get(lid, []):
            _d(ws_p, pr, 1, lid)
            _d(ws_p, pr, 2, pm.get('teamA', ''))
            _d(ws_p, pr, 3, pm.get('teamB', ''))
            _d(ws_p, pr, 4, pm.get('day', ''))
            _d(ws_p, pr, 5, pm.get('home', '') or 'beliebig')
            pr += 1

    # ── Sheet 8: Sperrtage ────────────────────────────────────────────────────
    ws_b = wb.create_sheet('Sperrtage')
    _set_col_w(ws_b, [16, 30, 30])
    for col, h in enumerate(['Liga-ID', 'Team', 'Spieltage'], 1):
        _h(ws_b, 1, col, h)
    br = 2
    for lid in S.league_order:
        for team, days in S.blocked.get(lid, {}).items():
            _d(ws_b, br, 1, lid)
            _d(ws_b, br, 2, team)
            _d(ws_b, br, 3, ', '.join(str(d) for d in sorted(days)))
            br += 1

    # ── Sheet 9: Pflichtheim ──────────────────────────────────────────────────
    ws_fh = wb.create_sheet('Pflichtheim')
    _set_col_w(ws_fh, [16, 30, 30])
    for col, h in enumerate(['Liga-ID', 'Team', 'Spieltage'], 1):
        _h(ws_fh, 1, col, h)
    fhr = 2
    for lid in S.league_order:
        for team, days in S.forced_home.get(lid, {}).items():
            _d(ws_fh, fhr, 1, lid)
            _d(ws_fh, fhr, 2, team)
            _d(ws_fh, fhr, 3, ', '.join(str(d) for d in sorted(days)))
            fhr += 1

    # ── Sheet 10: Co-Home ─────────────────────────────────────────────────────
    # Automatisch erkannte Vereine als Basis, explizit konfigurierte haben Vorrang
    _auto_clubs = _autodetect_cohome(S.league_order, S.leagues, load_club_db(), S.team_verein_map)
    _all_clubs: Dict[str, Dict[str, str]] = dict(_auto_clubs)
    _all_clubs.update(S.clubs)

    ws_c = wb.create_sheet('Co-Home')
    _set_col_w(ws_c, [28, 16, 30])
    for col, h in enumerate(['Vereinsname', 'Liga-ID', 'Teamname'], 1):
        _h(ws_c, 1, col, h)
    cr = 2
    for club_name, liga_map in _all_clubs.items():
        for lid, team in liga_map.items():
            _d(ws_c, cr, 1, club_name); _d(ws_c, cr, 2, lid); _d(ws_c, cr, 3, team)
            cr += 1

    # ── Sheet 11: Hinweise ────────────────────────────────────────────────────
    ws_h = wb.create_sheet('Hinweise')
    ws_h.column_dimensions['A'].width = 64
    for r, (text, bold) in enumerate([
        ('Spielplan-Optimierer – Vollständige Konfiguration', True),
        ('', False),
        ('Ligen & Teams', True),
        ('  Liga-ID: Kurzkürzel ohne Leerzeichen, z. B. BL1', False),
        ('  Spielformat: ' + ', '.join(FMT_OPTIONS), False),
        ('  Hierarchiegewicht: Priorität bei Co-Home-Konflikten (0.1–5.0)', False),
        ('  TT-Runden / Spiele/Tag / Teams/Gruppe: nur für Turniertag-Format', False),
        ('', False),
        ('Einstellungen', True),
        ('  dist_method: manual / file / maps', False),
        ('  same_weights: J = alle Ligen gleich gewichtet, N = pro Liga separat', False),
        ('  w_cohome: Co-Home-Bonus 0–10', False),
        ('  solver_p1/p2/sa: Zeitlimits in Sekunden; solver_seeds: Starts', False),
        ('', False),
        ('Gewichte', True),
        ('  Werte 0 (ignorieren) – 10 (höchste Priorität)', False),
        ('  Bei same_weights=J: Liga-ID = __common__', False),
        ('', False),
        ('Kalender', True),
        ('  Kalenderwoche (KW) und optionales Datum je Spieltag und Liga.', False),
        ('  Zwei aufeinanderfolgende Spieltage in gleicher KW → DST-Block.', False),
        ('', False),
        ('Routing', True),
        ('  J = Umwegbegrenzung aktiv; Mehrkilometer = erlaubter Aufschlag in %.', False),
        ('', False),
        ('Pflichtspiele', True),
        ('  Heimrecht "beliebig" = kein festes Heimrecht.', False),
        ('', False),
        ('Sperrtage', True),
        ('  Spieltage kommagetrennt, z. B.: 3, 7, 12', False),
        ('', False),
        ('Co-Home', True),
        ('  Eine Zeile pro Verein-Liga-Kombination.', False),
    ], 1):
        cell = ws_h.cell(r, 1, text)
        cell.font = Font(bold=bold, size=10, color='1F4E79' if bold else '000000')

    # ── Sheet: TT-Spielreihenfolge ────────────────────────────────────────────
    import json as _json
    _tt_lids = [l for l in S.league_order
                if S.leagues.get(l, {}).get('fmt') == 'Turniertag'
                and S.leagues.get(l, {}).get('tt_settings')]
    if _tt_lids:
        ws_tt = wb.create_sheet('TT-Spielreihenfolge')
        _tt_hdrs = ['Liga-ID', 'Ausrichterposition (J/N)', 'Mindestpause',
                    'Maximalpause', 'Ausrichter-Modus',
                    'Ausrichter-Zaehler (JSON)', 'Ausrichter-pro-Tag (JSON)']
        _set_col_w(ws_tt, [14, 24, 14, 14, 20, 45, 45])
        for _ttc, _tth in enumerate(_tt_hdrs, 1):
            _h(ws_tt, 1, _ttc, _tth)
        ws_tt.row_dimensions[1].height = 20
        _ttr = 2
        for _li, _lid in enumerate(_tt_lids):
            _ldd  = S.leagues.get(_lid, {})
            _tt   = _ldd.get('tt_settings', {})
            _fil  = ALT if _li % 2 == 0 else WHT
            _hcj  = _json.dumps(_tt.get('host_counts', {}), ensure_ascii=False) if _tt.get('host_counts') else ''
            _hpj  = _json.dumps({str(k): v for k, v in _tt.get('host_per_day', {}).items()},
                                 ensure_ascii=False) if _tt.get('host_per_day') else ''
            for _ttc, _ttv in enumerate([
                _lid,
                'J' if _tt.get('host_position', False) else 'N',
                _tt.get('min_gap', 0),
                _tt.get('max_gap', 3),
                _tt.get('host_mode', 'per_team'),
                _hcj, _hpj,
            ], 1):
                _d(ws_tt, _ttr, _ttc, _ttv, _fil)
            _ttr += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _load_full_config_excel(uploaded_file) -> Optional[dict]:
    """Liest eine vollständige Konfigurationsdatei (alle Sheets).
    Fällt auf Ligen-only zurück, wenn kein 'Einstellungen'-Sheet vorhanden."""
    try:
        fb = io.BytesIO(uploaded_file.getvalue())
        xl = pd.ExcelFile(fb)
        sn = set(xl.sheet_names)

        if 'Ligen & Teams' not in sn:
            st.error('Sheet "Ligen & Teams" nicht gefunden.')
            return None

        # Ligen & Teams (nutzt bestehende Funktion über BytesIO)
        fb.seek(0)
        teams_data = _load_teams_excel(fb)
        if teams_data is None:
            return None
        result = dict(teams_data)

        if 'Einstellungen' not in sn:
            return result  # altes Format – nur Teams zurückgeben

        # Einstellungen
        df_e = xl.parse('Einstellungen', dtype=str, header=0).fillna('')
        settings: dict = {}
        for _, row in df_e.iterrows():
            k = str(row.iloc[0]).strip()
            v = str(row.iloc[1]).strip()
            if k and k.lower() != 'nan':
                settings[k] = v
        result['settings'] = settings

        # Distanzmatrizen
        if 'Distanzmatrizen' in sn:
            df_d = xl.parse('Distanzmatrizen', header=None, dtype=str).fillna('')
            dist_matrices: dict = {}
            rows_arr = df_d.values
            i = 0
            while i < len(rows_arr):
                cell0 = str(rows_arr[i][0]).strip()
                if cell0.startswith('Liga:'):
                    lid_part = cell0[5:].strip()
                    # Robust gegen verschiedene Gedankenstrich-Varianten (–, —, ‒, U+FFFD)
                    _parts = re.split(r'\s*[–—‒�]\s*', lid_part, maxsplit=1)
                    lid = _parts[0].strip() if len(_parts) > 1 else lid_part.strip()
                    i += 1
                    if i >= len(rows_arr):
                        break
                    _all_hdrs = [str(c).strip() for c in rows_arr[i]]
                    _last_non_empty = max(
                        (j for j, h in enumerate(_all_hdrs) if h and h.lower() != 'nan'),
                        default=-1)
                    col_hdrs = [h for h in _all_hdrs[:_last_non_empty + 1]
                                if h and h.lower() != 'nan']
                    # n = Anzahl Teams (nur nicht-leere Header-Zellen)
                    n = len(col_hdrs)
                    i += 1
                    mat_rows = []
                    for _ in range(n):
                        if i >= len(rows_arr):
                            break
                        row_vals = list(rows_arr[i])
                        vals = []
                        for c in row_vals[1:n + 1]:
                            try:
                                vals.append(float(str(c) if str(c).lower() != 'nan' else '0'))
                            except Exception:
                                vals.append(0.0)
                        # Zeilen überspringen die leer sind (Trennzeilen zwischen Ligen)
                        if not any(v != 0.0 for v in vals) and not str(row_vals[0]).strip():
                            i += 1
                            break
                        mat_rows.append(vals)
                        i += 1
                    if len(mat_rows) == n > 0:
                        dist_matrices[lid] = np.array(mat_rows, dtype=float)
                else:
                    i += 1
            result['dist_matrices'] = dist_matrices

        # Gewichte
        if 'Gewichte' in sn:
            df_w = xl.parse('Gewichte', dtype=str).fillna('')
            weights: dict = {}
            for _, row in df_w.iterrows():
                lid = str(row.get('Liga-ID', '')).strip()
                if not lid or lid.lower() == 'nan':
                    continue
                try:
                    weights[lid] = {
                        'switch':    float(row.get('Heimrechtwechsel',    5.0) or 5.0),
                        'sw_fair':   float(row.get('Wechsel-Fairness',    5.0) or 5.0),
                        'travel':    float(row.get('Reisedistanz',        5.0) or 5.0),
                        'trav_fair': float(row.get('Reise-Fairness',      5.0) or 5.0),
                        'dst_eff':   float(row.get('DST-Reiseeffizienz',  0.0) or 0.0),
                    }
                except Exception:
                    pass
            result['weights'] = weights

        # Kalender
        if 'Kalender' in sn:
            df_kal = xl.parse('Kalender', dtype=str).fillna('')
            cal_raw: dict = {}
            for _, row in df_kal.iterrows():
                _lid = str(row.get('Liga-ID', '')).strip()
                if not _lid or _lid.lower() == 'nan':
                    continue
                try:
                    _st = int(float(str(row.get('Spieltag', ''))))
                except (ValueError, TypeError):
                    continue
                _kw_s = str(row.get('KW', '')).strip()
                _kw = None
                if _kw_s and _kw_s.lower() not in ('nan', ''):
                    try:
                        _kw = int(float(_kw_s))
                    except (ValueError, TypeError):
                        pass
                _date = str(row.get('Datum', '')).strip()
                if _date.lower() == 'nan':
                    _date = ''
                cal_raw.setdefault(_lid, {})[_st] = {'kw': _kw, 'date': _date}
            _cal_table: dict = {}
            for _lid in result.get('league_order', []):
                _ld  = result.get('leagues', {}).get(_lid, {})
                _n   = _calc_n_matchdays(_ld)
                _raw = cal_raw.get(_lid, {})
                _cal_table[_lid] = [
                    {'spieltag': i + 1,
                     'kw':   _raw.get(i + 1, {}).get('kw'),
                     'date': _raw.get(i + 1, {}).get('date', '')}
                    for i in range(_n)
                ]
            result['cal_table'] = _cal_table

        # DST-Blöcke
        if 'DST-Blöcke' in sn:
            df_dst = xl.parse('DST-Blöcke', dtype=str).fillna('')
            dst: dict = {}
            for _, row in df_dst.iterrows():
                lid = str(row.get('Liga-ID', '')).strip()
                if not lid or lid.lower() == 'nan':
                    continue
                try:
                    d1 = int(float(str(row.get('Spieltag 1', ''))))
                    d2 = int(float(str(row.get('Spieltag 2', ''))))
                    dst.setdefault(lid, []).append((min(d1, d2), max(d1, d2)))
                except Exception:
                    pass
            result['dst_per_liga'] = dst

        # Routing
        if 'Routing' in sn:
            df_rt = xl.parse('Routing', dtype=str).fillna('')
            routing: dict = {}
            for _, row in df_rt.iterrows():
                lid = str(row.get('Liga-ID', '')).strip()
                if not lid or lid.lower() == 'nan':
                    continue
                try:
                    apply_ = str(row.get('Routing aktiv (J/N)', 'N')).strip().upper() == 'J'
                    pct    = int(float(str(row.get('Mehrkilometer (%)', 25) or 25)))
                    routing[lid] = (apply_, pct)
                except Exception:
                    pass
            result['routing'] = routing

        # Pflichtspiele
        if 'Pflichtspiele' in sn:
            df_p = xl.parse('Pflichtspiele', dtype=str).fillna('')
            pinned: dict = {}
            for _, row in df_p.iterrows():
                lid = str(row.get('Liga-ID', '')).strip()
                if not lid or lid.lower() == 'nan':
                    continue
                try:
                    home = str(row.get('Heimrecht', 'beliebig')).strip()
                    pinned.setdefault(lid, []).append({
                        'teamA': str(row.get('Team A',    '')).strip(),
                        'teamB': str(row.get('Team B',    '')).strip(),
                        'day':   int(float(str(row.get('Spieltag', 1) or 1))),
                        'home':  None if home in ('beliebig', '', 'nan') else home,
                    })
                except Exception:
                    pass
            result['pinned'] = pinned

        # Sperrtage
        if 'Sperrtage' in sn:
            df_b = xl.parse('Sperrtage', dtype=str).fillna('')
            blocked: dict = {}
            for _, row in df_b.iterrows():
                lid  = str(row.get('Liga-ID', '')).strip()
                team = str(row.get('Team',    '')).strip()
                if not lid or lid.lower() == 'nan' or not team or team.lower() == 'nan':
                    continue
                try:
                    days_str = str(row.get('Spieltage', '') or '').strip()
                    days = sorted({int(x.strip()) for x in days_str.split(',')
                                   if x.strip() and x.strip() != 'nan'})
                    if days:
                        blocked.setdefault(lid, {})[team] = days
                except Exception:
                    pass
            result['blocked'] = blocked

        # Pflichtheim
        if 'Pflichtheim' in sn:
            df_fh = xl.parse('Pflichtheim', dtype=str).fillna('')
            forced_home_loaded: dict = {}
            for _, row in df_fh.iterrows():
                lid  = str(row.get('Liga-ID', '')).strip()
                team = str(row.get('Team',    '')).strip()
                if not lid or lid.lower() == 'nan' or not team or team.lower() == 'nan':
                    continue
                try:
                    days_str = str(row.get('Spieltage', '') or '').strip()
                    days = sorted({int(x.strip()) for x in days_str.split(',')
                                   if x.strip() and x.strip() != 'nan'})
                    if days:
                        forced_home_loaded.setdefault(lid, {})[team] = days
                except Exception:
                    pass
            result['forced_home'] = forced_home_loaded

        # Co-Home
        if 'Co-Home' in sn:
            df_c = xl.parse('Co-Home', dtype=str).fillna('')
            clubs: dict = {}
            for _, row in df_c.iterrows():
                club = str(row.get('Vereinsname', '')).strip()
                lid  = str(row.get('Liga-ID',    '')).strip()
                team = str(row.get('Teamname',   '')).strip()
                if not club or club.lower() == 'nan':
                    continue
                if not lid or lid.lower() == 'nan' or not team or team.lower() == 'nan':
                    continue
                clubs.setdefault(club, {})[lid] = team
            result['clubs'] = clubs

        # TT-Spielreihenfolge
        if 'TT-Spielreihenfolge' in sn:
            import json as _json
            df_tt = xl.parse('TT-Spielreihenfolge', dtype=str).fillna('')
            for _, row in df_tt.iterrows():
                lid = str(row.get('Liga-ID', '')).strip()
                if not lid or lid.lower() == 'nan':
                    continue
                if lid not in result['leagues']:
                    continue
                tt = result['leagues'][lid].setdefault('tt_settings', {})
                try:
                    tt['host_position'] = str(row.get('Ausrichterposition (J/N)', 'N')).strip().upper() == 'J'
                except Exception:
                    pass
                for _col, _key, _default in [('Mindestpause', 'min_gap', 0),
                                              ('Maximalpause', 'max_gap', 3)]:
                    try:
                        tt[_key] = int(float(str(row.get(_col, _default) or _default)))
                    except Exception:
                        pass
                hmode = str(row.get('Ausrichter-Modus', 'per_team')).strip()
                tt['host_mode'] = 'per_day' if hmode == 'per_day' else 'per_team'
                for _col, _key, _conv in [
                    ('Ausrichter-Zaehler (JSON)', 'host_counts',
                     lambda s: _json.loads(s)),
                    ('Ausrichter-pro-Tag (JSON)', 'host_per_day',
                     lambda s: {int(k): v for k, v in _json.loads(s).items()}),
                ]:
                    _v = str(row.get(_col, '') or '').strip()
                    if _v and _v.lower() not in ('nan', '{}', ''):
                        try:
                            tt[_key] = _conv(_v)
                        except Exception:
                            pass

        return result

    except zipfile.BadZipFile:
        st.error('Die Datei ist beschädigt oder kein gültiges Excel-Format (.xlsx).')
        return None
    except ValueError as exc:
        st.error(f'Sheet nicht gefunden oder ungültiges Format: {exc}')
        return None
    except Exception as exc:
        st.error(f'Lesefehler beim Import: {exc}')
        return None


def _step0():
    st.header('1. Ligen & Teams konfigurieren')
    st.info('Definiere alle Ligen. Teams per Vereinssuche hinzufügen – '
            'Adresse wird automatisch aus der Datenbank übernommen.')

    # ── Anzahl Ligen (zuerst, damit Vorlage die richtige Zahl enthält) ────────
    # Session State immer VOR dem Widget-Rendering setzen, damit kein value=-Konflikt entsteht
    if '_pending_n_ligen' in st.session_state:
        st.session_state['n_ligen_input'] = st.session_state.pop('_pending_n_ligen')
    elif 'n_ligen_input' not in st.session_state:
        st.session_state['n_ligen_input'] = max(1, len(S.league_order) if S.league_order else 1)
    n = int(st.number_input('Anzahl Ligen', min_value=1, max_value=8, step=1, key='n_ligen_input'))

    # ── Template Import / Export ──────────────────────────────────────────────
    with st.expander('📋 Konfiguration laden / Leere Vorlage herunterladen', expanded=not bool(S.league_order)):
        c_dl, c_up = st.columns(2)
        with c_dl:
            st.markdown('**1. Leere Vorlage herunterladen**')
            st.caption(f'Leere Vorlage für {n} Liga(en) – befüllen und hochladen.')
            _ph_ords = [f'LIGA_{i}' for i in range(1, n + 1)]
            _ph_ligs = {lid: dict(name=f'Liga {i}', fmt=FMT_OPTIONS[0], hw=1.0,
                            teams=[('Teamname 1', 'Adresse 1'), ('Teamname 2', 'Adresse 2'),
                                   ('Teamname 3', 'Adresse 3'), ('Teamname 4', 'Adresse 4')])
                        for i, lid in enumerate(_ph_ords, 1)}
            xl_bytes = _teams_excel_bytes(_ph_ligs, _ph_ords)
            st.download_button('⬇ Leere Vorlage herunterladen',
                data=xl_bytes, file_name='Ligen_Teams_Vorlage.xlsx',
                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                key='dl_tmpl_empty', width='stretch')
        with c_up:
            st.markdown('**2. Konfiguration hochladen**')
            st.caption('Vollständige Konfigurationsdatei oder ausgefüllte Vorlage (Ligen & Teams) hochladen.')
            _ukey   = f'tmpl_upload_{st.session_state.get("_upload_key", 0)}'
            up_tmpl = st.file_uploader('Konfiguration / Vorlage', type=['xlsx', 'xls'], key=_ukey,
                label_visibility='collapsed')
            if up_tmpl:
                parsed = _load_full_config_excel(up_tmpl)
                st.session_state['_upload_key'] = st.session_state.get('_upload_key', 0) + 1
                if parsed:
                    S.league_order  = parsed['league_order']
                    S.leagues       = parsed['leagues']
                    _has_loaded_matrices = False
                    if 'dist_matrices' in parsed:
                        S.dist_matrices = parsed['dist_matrices']
                        for _lid in parsed['dist_matrices']:
                            st.session_state.pop(f'de_{_lid}', None)
                        _has_loaded_matrices = bool(parsed['dist_matrices'])
                    if 'settings' in parsed:
                        s = parsed['settings']
                        if 'dist_method' in s:
                            S.dist_method = s['dist_method']
                    # Matrizen sind bereits geladen → manuell anzeigen, kein API-Aufruf nötig
                    if _has_loaded_matrices:
                        S.dist_method = 'manual'
                        if 'same_weights' in s:
                            S.same_weights = s['same_weights'].strip().upper() == 'J'
                        if 'w_cohome' in s:
                            try:
                                S.w_cohome = float(s['w_cohome'])
                            except Exception:
                                pass
                        sol = dict(S.solver)
                        for sk, dk in [('solver_p1', 'p1'), ('solver_p2', 'p2'),
                                       ('solver_sa', 'sa'), ('solver_seeds', 'seeds')]:
                            if sk in s:
                                try:
                                    sol[dk] = int(float(s[sk]))
                                except Exception:
                                    pass
                        sol['nm'] = sol.get('p2', 5400) >= 28800
                        S.solver = sol
                    if 'weights' in parsed:
                        S.weights = parsed['weights']
                        if '__common__' in parsed['weights']:
                            S.same_weights = True
                    if 'dst_per_liga' in parsed:
                        S.dst_per_liga = parsed['dst_per_liga']
                    if 'cal_table' in parsed:
                        S.cal_table = parsed['cal_table']
                        for _lid in S.league_order:
                            st.session_state.pop(f'cal_editor_{_lid}', None)
                        _cal_table_to_kw_compat()
                    if 'routing' in parsed:
                        S.routing = parsed['routing']
                    if 'pinned' in parsed:
                        S.pinned = parsed['pinned']
                    if 'blocked' in parsed:
                        S.blocked = parsed['blocked']
                    if 'forced_home' in parsed:
                        S.forced_home = parsed['forced_home']
                    if 'clubs' in parsed:
                        S.clubs = parsed['clubs']
                    # team_verein_map aus DB für alle geladenen Teams aufbauen
                    _db_tv = {r['teamname']: (r['verein'] or r['teamname'])
                              for r in load_club_db() if r.get('verein')}
                    _tvm: dict = {}
                    for _lid in S.league_order:
                        for _tn, _ in S.leagues.get(_lid, {}).get('teams', []):
                            if _tn in _db_tv:
                                _tvm[_tn] = _db_tv[_tn]
                            else:
                                _norm = _normalize_club_name(_tn)
                                if _norm in _db_tv:
                                    _tvm[_tn] = _db_tv[_norm]
                    S.team_verein_map = _tvm
                    n_loaded = len(parsed['league_order'])
                    n_teams  = sum(len(v['teams']) for v in parsed['leagues'].values())
                    st.session_state['_pending_n_ligen'] = n_loaded
                    extra_sheets = [k for k in parsed if k not in ('league_order', 'leagues')]
                    st.session_state['_tmpl_msg'] = (
                        f'{n_loaded} Ligen · {n_teams} Teams geladen'
                        + (f' · inkl. {len(extra_sheets)} weitere Einstellungen' if extra_sheets else '') + '.'
                    )
                    # Widget-Keys für geladene Ligen explizit setzen – Session State hat
                    # immer Vorrang vor Browser-Rückgabewerten, daher SET statt pop().
                    for _i, _lid in enumerate(parsed['league_order']):
                        _ld = parsed['leagues'].get(_lid, {})
                        st.session_state[f'lid_{_i}'] = _lid
                        st.session_state[f'lnm_{_i}'] = _ld.get('name', _lid)
                        _fmt = _FMT_ALIAS.get(
                            _ld.get('fmt', FMT_OPTIONS[0]),
                            _ld.get('fmt', FMT_OPTIONS[0]))
                        st.session_state[f'fmt_{_i}'] = (
                            _fmt if _fmt in FMT_OPTIONS else FMT_OPTIONS[0])
                        try:
                            st.session_state[f'hw_{_i}'] = max(0.1, min(5.0, float(_ld.get('hw', 1.0))))
                        except Exception:
                            st.session_state[f'hw_{_i}'] = 1.0
                        # Turniertag-Widget-Keys
                        if _ld.get('fmt') == 'Turniertag':
                            _apd = _ld.get('active_per_day')
                            if _apd:
                                st.session_state[f'apd_{_i}'] = int(_apd)
                            _tt = _ld.get('tt_settings', {})
                            if _tt:
                                st.session_state[f'tt_hmode_{_lid}'] = (
                                    'Pro Spieltag' if _tt.get('host_mode') == 'per_day' else 'Anzahl pro Team')
                                st.session_state[f'tt_mingap_{_lid}'] = int(_tt.get('min_gap', 0))
                                st.session_state[f'tt_maxgap_{_lid}'] = int(_tt.get('max_gap', 3))
                                _loaded_slots = _tt.get('host_slots', [])
                                _ld_gpd = int(_ld.get('gpd', 2) or 2)
                                for _gi in range(_ld_gpd):
                                    _sv = int(_loaded_slots[_gi]) if _gi < len(_loaded_slots) else 0
                                    st.session_state[f'tt_hslot_{_lid}_{_gi}'] = _sv
                                for _t, _cnt in _tt.get('host_counts', {}).items():
                                    st.session_state[f'tt_hcnt_{_lid}_{_t}'] = int(_cnt)
                                for _dd, _hh in _tt.get('host_per_day', {}).items():
                                    st.session_state[f'tt_hday_{_lid}_{_dd}'] = _hh
                    # Keys jenseits der geladenen Ligen-Anzahl löschen
                    for _i in range(len(parsed['league_order']), 20):
                        for _k in (f'lid_{_i}', f'lnm_{_i}', f'fmt_{_i}', f'hw_{_i}',
                                   f'ttr_{_i}', f'gpd_{_i}', f'kg_{_i}', f'apd_{_i}'):
                            st.session_state.pop(_k, None)
                    st.rerun()

    if '_tmpl_msg' in st.session_state:
        st.success(st.session_state.pop('_tmpl_msg'))

    # ── Anzahl anpassen: letzte Liga entfernen / neue leere Liga anhängen ──────
    while len(S.league_order) < n:
        # Freie Liga-ID finden, die noch nicht vergeben ist
        _i = len(S.league_order) + 1
        while f'LIGA_{_i}' in S.leagues:
            _i += 1
        lid = f'LIGA_{_i}'
        S.league_order.append(lid)
        S.leagues[lid] = dict(name=f'Liga {_i}', teams=[], fmt=FMT_OPTIONS[0], hw=1.0)
    while len(S.league_order) > n:
        _removed = S.league_order.pop()
        S.leagues.pop(_removed, None)
        for _sd in (S.dist_matrices, S.dst_per_liga, S.routing,
                    S.weights, S.pinned, S.blocked, S.forced_home):
            _sd.pop(_removed, None)
        for _club_dict in S.clubs.values():
            _club_dict.pop(_removed, None)

    club_records = load_club_db()
    club_adresse  = {(r['verein'] or r['teamname']): r['adresse']  for r in club_records}
    club_teamname = {(r['verein'] or r['teamname']): r['teamname'] for r in club_records}
    SEARCH_PH  = '– Verein suchen –'
    MANUAL_OPT = '(Manuell eingeben)'

    errors = []
    for i, lid in enumerate(S.league_order):
        ld  = S.leagues[lid]
        n_t = len(ld.get('teams', []))
        # Club-DB für diese Liga filtern (nach Ligabezeichnung)
        _liga_nm = ld.get('name', '').lower()
        _filtered = [r for r in club_records if r['liga'] and r['liga'].lower() == _liga_nm]
        club_opts = [SEARCH_PH, MANUAL_OPT] + sorted(set(
            r['verein'] or r['teamname'] for r in (_filtered if _filtered else club_records)
        ))
        _exp_key = f'_exp_{lid}'
        if _exp_key not in st.session_state:
            st.session_state[_exp_key] = (n_t < 4)
        with st.expander(f'Liga {i+1}: {ld.get("name", lid)}  ({n_t} Teams)',
                         expanded=st.session_state[_exp_key]):

            # ── Liga-Metadaten ────────────────────────────────────────────────
            c1, c2, c3 = st.columns([2, 3, 2])
            with c1:
                if f'lid_{i}' not in st.session_state:
                    st.session_state[f'lid_{i}'] = lid
                new_lid = st.text_input('Liga-ID', key=f'lid_{i}',
                    help='Kurzkürzel ohne Leerzeichen, z. B. BL1').strip().upper()
            with c2:
                if f'lnm_{i}' not in st.session_state:
                    st.session_state[f'lnm_{i}'] = ld.get('name', lid)
                ld['name'] = st.text_input('Ligabezeichnung', key=f'lnm_{i}')
            with c3:
                # Rückwärtskompatibilität: alten Formatnamen migrieren
                cur_fmt = _FMT_ALIAS.get(ld.get('fmt', FMT_OPTIONS[0]), ld.get('fmt', FMT_OPTIONS[0]))
                if cur_fmt not in FMT_OPTIONS:
                    cur_fmt = FMT_OPTIONS[0]
                ld['fmt'] = st.selectbox('Spielformat', FMT_OPTIONS,
                    index=FMT_OPTIONS.index(cur_fmt), key=f'fmt_{i}',
                    help='Einfachrunde: jeder gegen jeden einmal · Hin-/Rückrunde: zweimal (Hin- und Rückspiel) · Dreifachrunde: dreimal · Turniertag: alle Teams spielen gemeinsam an einem Ort')
            if n > 1:
                if f'hw_{i}' not in st.session_state:
                    st.session_state[f'hw_{i}'] = float(ld.get('hw', 1.0))
                ld['hw'] = st.slider('Priorität bei Heimspiel-Koordination', 0.1, 5.0,
                    step=0.1, key=f'hw_{i}',
                    help='Wenn Heimspiele mehrerer Ligen nicht alle in dieselbe Woche passen, wird die Liga mit dem höheren Wert bevorzugt.')

            # ── Turniertag-Optionen ───────────────────────────────────────────
            if ld.get('fmt') == 'Turniertag':
                n_t = len(ld.get('teams', []))
                ta, tb, tc = st.columns(3)
                with ta:
                    tt_rounds_opts = ['Einfachrunde (1 Runde)', 'Hin-/Rückrunde (2 Runden)', 'Dreifachrunde (3 Runden)']
                    tt_rounds_val  = int(ld.get('tt_rounds', 2))
                    st.selectbox('Turniertag – Runden',
                        tt_rounds_opts, index=max(0, min(tt_rounds_val - 1, 2)),
                        key=f'ttr_{i}',
                        help='Wie viele Runden soll die Liga haben?',
                    )
                    # Einfachrunde→1, Hin/Rück→2, Dreifach→3
                    ld['tt_rounds'] = [1, 2, 3][tt_rounds_opts.index(
                        st.session_state.get(f'ttr_{i}', tt_rounds_opts[tt_rounds_val - 1]))]
                with tb:
                    _gpd_opts = [1, 2, 3]
                    cur_gpd   = int(ld.get('gpd', 2))
                    if cur_gpd not in _gpd_opts:
                        cur_gpd = 2
                    ld['gpd'] = st.selectbox(
                        'Spiele pro Team pro Spieltag', _gpd_opts,
                        index=_gpd_opts.index(cur_gpd), key=f'gpd_{i}',
                        help='1 = ein Spiel pro Turniertag · 2 = zwei Spiele · 3 = drei Spiele.',
                    )
                    _gpd_v   = ld['gpd']
                    _apd_min = _gpd_v + 1
                    _cur_apd = int(ld.get('active_per_day', n_t))
                    _cur_apd = min(max(_cur_apd, _apd_min), n_t) if n_t >= _apd_min else n_t
                    if n_t > _apd_min:
                        ld['active_per_day'] = st.number_input(
                            'Anwesende Teams/Spieltag',
                            min_value=_apd_min, max_value=n_t,
                            value=_cur_apd, step=1, key=f'apd_{i}',
                            help=(
                                f'Wie viele der {n_t} Teams nehmen pro Turniertag teil? '
                                'Teams, die nicht alle Spieltage dabei sind, haben an manchen Terminen spielfrei.'
                            ),
                        )
                    else:
                        ld['active_per_day'] = n_t
                    _apd = int(ld.get('active_per_day', n_t))
                    _n_rounds_v, _ = _get_n_rounds_gpd(ld)
                    _total_m_v = _n_rounds_v * n_t * (n_t - 1) // 2
                    _apd_ok = (_apd >= _gpd_v + 1
                               and _apd * _gpd_v % 2 == 0
                               and (_total_m_v * 2) % max(1, _apd * _gpd_v) == 0)
                    if n_t < 2:
                        st.caption('Mindestens 2 Teams konfigurieren.')
                    elif not _apd_ok:
                        _compat = [a for a in range(_gpd_v + 1, n_t + 1)
                                   if a * _gpd_v % 2 == 0
                                   and (_total_m_v * 2) % max(1, a * _gpd_v) == 0]
                        _subj = (f'{_apd} anwesende Teams' if _apd < n_t
                                 else f'{n_t} Teams')
                        st.warning(
                            f'**{_subj}** lässt sich nicht gleichmäßig auf **{_gpd_v} Spiele/Tag** aufteilen. '
                            'Passende Teamanzahlen wären: '
                            + ', '.join(str(a) for a in _compat[:5])
                        )
                    elif _apd < n_t:
                        _n_days_bye = _total_m_v * 2 // max(1, _apd * _gpd_v)
                        st.info(
                            f'Pro Spieltag sind {_apd} von {n_t} Teams dabei – '
                            f'{n_t - _apd} Team(s) haben jeweils Spielfrei. '
                            f'Insgesamt {_n_days_bye} Spieltage. Die Spielfrei-Tage werden fair auf alle Teams verteilt.'
                        )
                with tc:
                    cur_gpd_val = int(ld.get('gpd', 2))
                    _apd_tc = int(ld.get('active_per_day', n_t))
                    if _apd_tc < n_t:
                        # Spielfrei-Modus: Gruppe = alle anwesenden Teams
                        ld['k_group'] = _apd_tc
                        n_days = _calc_n_matchdays(ld)
                        if n_days > 0:
                            st.caption(f'→ {n_days} Spieltage · 1 Gruppe ({_apd_tc} Teams/Tag, {n_t - _apd_tc} Spielfrei)')
                    else:
                        valid_k = _valid_K_values(n_t, cur_gpd_val) if n_t >= cur_gpd_val + 1 else []
                        if valid_k:
                            cur_k = int(ld.get('k_group', n_t))
                            if cur_k not in valid_k:
                                cur_k = valid_k[-1]
                            k_labels = [
                                f'Alle ({n_t}) – 1 Gruppe' if K == n_t
                                else f'{K} pro Gruppe ({math.ceil(n_t / K)} Gruppen)'
                                for K in valid_k
                            ]
                            sel_lbl = st.selectbox(
                                'Teams pro Gruppe', k_labels,
                                index=valid_k.index(cur_k) if cur_k in valid_k else len(valid_k) - 1,
                                key=f'kg_{i}',
                                help='Alle Teams gemeinsam = ein einziger Austragungsort. Weniger Teams pro Gruppe = mehrere Gruppen, die an verschiedenen Orten spielen.')
                            ld['k_group'] = valid_k[k_labels.index(sel_lbl)]
                            K_val  = ld['k_group']
                            n_days = _calc_n_matchdays(ld)
                            G_val  = math.ceil(n_t / K_val)
                            if K_val < n_t:
                                st.caption(f'→ {n_days} Spieltage · {G_val} Gruppen à {K_val}')
                            else:
                                _gpd_day = n_t * cur_gpd_val
                                st.caption(
                                    f'→ {n_days} Spieltage · '
                                    + (f'{_gpd_day // 2} Spiele/Tag' if _gpd_day % 2 == 0
                                       else f'{n_t} Teams × {cur_gpd_val}/2 Spiele/Tag')
                                )
                        else:
                            ld['k_group'] = n_t
                            if n_t >= 2:
                                n_days = _calc_n_matchdays(ld)
                                st.caption(f'→ {n_days} Spieltage (alle {n_t} Teams gemeinsam)')

                # ── Turniertag-Spielreihenfolge ───────────────────────────────
                _gpd_val = int(ld.get('gpd', 1))
                if n_t >= 4 and _gpd_val > 1:
                    tt_s = ld.setdefault('tt_settings', {})
                    n_days_tt  = _calc_n_matchdays(ld)
                    teams_list = [t for t, _ in ld.get('teams', [])]
                    _apd_tt = int(ld.get('active_per_day', 0) or 0)
                    _n_active_tt = _apd_tt if _apd_tt > 0 else n_t
                    _N_day_tt = _n_active_tt * _gpd_val // 2  # Spiele pro Turniertag

                    # Expander nur öffnen wenn explizit Ausrichter eingetragen
                    _hc_vals    = tt_s.get('host_counts', {})
                    _has_tt_cfg = (bool(tt_s.get('host_per_day'))
                                   or any(v > 0 for v in _hc_vals.values()))
                    _ekey_tt = f'_exp_tt_{lid}'
                    if _has_tt_cfg:
                        st.session_state[_ekey_tt] = True
                    elif _ekey_tt not in st.session_state:
                        st.session_state[_ekey_tt] = False

                    with st.expander('Spielreihenfolge am Turniertag (optional)',
                                     expanded=st.session_state[_ekey_tt]):
                        st.caption(
                            'Hier kannst du festlegen, in welcher Reihenfolge '
                            'die Spiele innerhalb eines Turniertags stattfinden sollen.'
                        )

                        # Anzahl Spiele pro Turniertag anzeigen
                        st.caption(f'**{_N_day_tt} Spiele pro Turniertag** '
                                   f'({_n_active_tt} Teams × {_gpd_val} Spiele je Team ÷ 2)')

                        # Ausrichter-Slot-Eingaben
                        # Rückwärtskompatibilität: host_position → host_slots
                        if 'host_slots' not in tt_s and tt_s.get('host_position', False):
                            tt_s['host_slots'] = [2, max(2, _N_day_tt - 1)] if _N_day_tt >= 4 else []

                        _existing_slots = tt_s.get('host_slots', [])
                        _slot_cols = st.columns(max(1, _gpd_val))
                        _new_slots = []
                        for _gi in range(_gpd_val):
                            with _slot_cols[_gi % len(_slot_cols)]:
                                try:
                                    _default_s = int(_existing_slots[_gi]) if _gi < len(_existing_slots) else 0
                                except (ValueError, TypeError):
                                    _default_s = 0
                                _sv = st.number_input(
                                    f'Ausrichter-Spiel {_gi + 1}: Position (0 = beliebig)',
                                    min_value=0, max_value=_N_day_tt,
                                    value=_default_s, step=1,
                                    key=f'tt_hslot_{lid}_{_gi}',
                                    help=(f'An welcher Stelle im Turniertag soll das {_gi + 1}. Ausrichter-Spiel stehen? '
                                          f'1 = erstes Spiel des Tages, {_N_day_tt} = letztes. '
                                          '0 = keine Vorgabe.'),
                                )
                                if _sv > 0:
                                    _new_slots.append(_sv)
                        tt_s['host_slots'] = _new_slots
                        if _new_slots:
                            st.caption(f'Gewünschte Positionen der Ausrichter-Spiele: {_new_slots} '
                                       f'(1 = erstes Spiel des Tages)')

                        with st.container():
                            _hmode_labels = ['Anzahl pro Team', 'Pro Spieltag']
                            _hmode_cur    = tt_s.get('host_mode', 'per_team')
                            _hmode_idx    = 1 if _hmode_cur == 'per_day' else 0
                            _hmode_sel    = st.radio(
                                'Ausrichter festlegen',
                                _hmode_labels,
                                index=_hmode_idx,
                                key=f'tt_hmode_{lid}',
                                horizontal=True,
                                help=(
                                    '**Anzahl pro Team**: Jedes Team bekommt eine bestimmte Anzahl Spieltage als Ausrichter zugewiesen – der Optimierer verteilt sie automatisch.  \n'
                                    '**Pro Spieltag**: Du legst konkret fest, welches Team welchen Spieltag ausrichtet.'
                                ),
                            )
                            tt_s['host_mode'] = 'per_day' if _hmode_sel == 'Pro Spieltag' else 'per_team'

                        _tts_c, _tts_d = st.columns(2)
                        with _tts_c:
                            _mg = int(tt_s.get('min_gap', 0))
                            tt_s['min_gap'] = st.slider(
                                'Mindestpause zwischen zwei Auftritten (Spiele)',
                                0, 3, _mg,
                                key=f'tt_mingap_{lid}',
                                help=(
                                    'Wie viele Spiele anderer Teams müssen mindestens zwischen zwei Auftritten eines Teams liegen? '
                                    '0 = kein Mindestabstand (sinnvoll bei kleinen Turnieren).'
                                ),
                            )
                        with _tts_d:
                            _min_maxg = max(1, int(tt_s['min_gap']) + 1)
                            _xg = max(int(tt_s.get('max_gap', 3)), _min_maxg)
                            tt_s['max_gap'] = st.slider(
                                'Maximalpause zwischen zwei Auftritten (Spiele)',
                                _min_maxg, 5, _xg,
                                key=f'tt_maxgap_{lid}',
                                help='Wie viele Spiele anderer Teams dürfen höchstens zwischen zwei Auftritten liegen?',
                            )

                        if teams_list and n_days_tt > 0:
                            st.divider()
                            if tt_s['host_mode'] == 'per_team':
                                _hc_head, _hc_btn = st.columns([3, 1])
                                _hc_head.markdown('**Wie viele Turniertage richtet jedes Team aus?**')
                                if _hc_btn.button('Gleichmäßig', key=f'tt_dist_{lid}',
                                                  help='Verteilt Ausrichter-Spieltage automatisch gleich auf alle Teams'):
                                    _base = n_days_tt // len(teams_list)
                                    _rem  = n_days_tt % len(teams_list)
                                    for _ti, _t in enumerate(teams_list):
                                        tt_s.setdefault('host_counts', {})[_t] = (
                                            _base + (1 if _ti < _rem else 0)
                                        )
                                        st.session_state[f'tt_hcnt_{lid}_{_t}'] = (
                                            _base + (1 if _ti < _rem else 0)
                                        )
                                    st.session_state[_ekey_tt] = True
                                    st.rerun()

                                host_counts = tt_s.setdefault('host_counts', {})
                                _total   = 0
                                _hc_cols = st.columns(min(4, len(teams_list)))
                                for _ti, _t in enumerate(teams_list):
                                    with _hc_cols[_ti % len(_hc_cols)]:
                                        _cnt = st.number_input(
                                            _t, 0, n_days_tt,
                                            int(host_counts.get(_t, 0)), 1,
                                            key=f'tt_hcnt_{lid}_{_t}',
                                        )
                                    host_counts[_t] = _cnt
                                    _total += _cnt

                                tt_s['host_counts'] = host_counts
                                _diff = _total - n_days_tt
                                if _total == 0:
                                    st.caption(f'{n_days_tt} Spieltage · noch kein Ausrichter eingetragen')
                                elif _diff != 0:
                                    st.warning(
                                        f'Summe: {_total} Spieltage zugewiesen, '
                                        f'{n_days_tt} vorhanden – Differenz: {_diff:+d}'
                                    )
                                else:
                                    st.success(f'Summe stimmt: {_total} = {n_days_tt} Spieltage')

                            else:  # per_day
                                st.markdown('**Wer richtet welchen Spieltag aus?**')
                                host_per_day = tt_s.setdefault('host_per_day', {})
                                _team_opts   = ['(Kein Ausrichter)'] + teams_list
                                _hpd_cols    = st.columns(min(4, n_days_tt))
                                for _d in range(1, n_days_tt + 1):
                                    _cur_h = host_per_day.get(_d, _team_opts[0])
                                    if _cur_h not in _team_opts:
                                        _cur_h = _team_opts[0]
                                    with _hpd_cols[(_d - 1) % len(_hpd_cols)]:
                                        _sel = st.selectbox(
                                            f'ST {_d}', _team_opts,
                                            index=_team_opts.index(_cur_h),
                                            key=f'tt_hday_{lid}_{_d}',
                                        )
                                    if _sel == '(Kein Ausrichter)':
                                        host_per_day.pop(_d, None)
                                    else:
                                        host_per_day[_d] = _sel
                                tt_s['host_per_day'] = host_per_day

            st.divider()

            # ── Team-Liste ────────────────────────────────────────────────────
            teams: List = ld.get('teams', [])
            st.markdown(f'**Teams** ({len(teams)} eingetragen)')

            if teams:
                header = st.columns([1, 4, 4, 1])
                header[0].caption('Nr.')
                header[1].caption('Teamname')
                header[2].caption('Standort / Adresse')
                header[3].caption('')
                for idx_t, (tname, city) in enumerate(teams):
                    tc0, tc1, tc2, tc3 = st.columns([1, 4, 4, 1])
                    tc0.write(str(idx_t + 1))
                    tc1.write(tname)
                    tc2.caption(city)
                    if tc3.button('✕', key=f'dt_{lid}_{idx_t}',
                                  help='Team entfernen'):
                        teams.pop(idx_t)
                        ld['teams'] = teams
                        st.rerun()
            else:
                st.caption('Noch keine Teams. Verein unten suchen und hinzufügen.')

            # ── Team hinzufügen ───────────────────────────────────────────────
            st.markdown('**Team hinzufügen**')

            # Reset-Flag: Keys löschen bevor Widgets instantiiert werden
            if st.session_state.pop(f'_reset_{lid}', False):
                for _k in (f'cs_{lid}', f'atn_{lid}', f'acy_{lid}'):
                    st.session_state.pop(_k, None)

            # on_change: Felder automatisch befüllen wenn Team aus DB gewählt
            def _on_club_select(k_sel=f'cs_{lid}', k_name=f'atn_{lid}',
                                k_city=f'acy_{lid}', _adr=club_adresse, _tnm=club_teamname):
                sel = st.session_state.get(k_sel, SEARCH_PH)
                if sel not in (SEARCH_PH, MANUAL_OPT):
                    st.session_state[k_name] = _tnm.get(sel, sel)
                    st.session_state[k_city] = _adr.get(sel, '')
                elif sel == MANUAL_OPT:
                    st.session_state[k_name] = ''
                    st.session_state[k_city] = ''

            fa, fb, fc, fd = st.columns([4, 4, 4, 1])
            with fa:
                sel = st.selectbox(
                    'Verein suchen',
                    club_opts,
                    key=f'cs_{lid}',
                    on_change=_on_club_select,
                    label_visibility='visible',
                    help=f'{len(club_opts) - 2} Teams in Datenbank. '
                         'Tippen zum Filtern. Nicht gefunden → "(Manuell eingeben)" wählen.',
                )
            sel = st.session_state.get(f'cs_{lid}', SEARCH_PH)
            waiting = (sel == SEARCH_PH)
            manual  = (sel == MANUAL_OPT)

            # Standort: bei DB-Treffer vorausgefüllt und editierbar (für Google-Maps-Adressen)
            with fb:
                tname_val = st.text_input(
                    'Teamname im Spielplan',
                    key=f'atn_{lid}',
                    placeholder='z. B. ETV Hamburg Damen',
                    disabled=waiting,
                    help='Kann vom Vereinsnamen abweichen (z. B. Damen/Herren/U23)',
                )
            with fc:
                city_val = st.text_input(
                    'Standort / Adresse',
                    key=f'acy_{lid}',
                    placeholder='Stadt oder vollständige Adresse',
                    disabled=waiting,
                    help='Für Google Maps: vollständige Adresse empfohlen. '
                         'Wird automatisch aus der Vereinsdatenbank übernommen.',
                )
            with fd:
                st.write('')  # vertikaler Abstand
                can_add = (not waiting
                           and bool(str(st.session_state.get(f'atn_{lid}', '')).strip()))
                add_clicked = st.button('＋', key=f'ab_{lid}',
                                        disabled=not can_add, type='primary',
                                        help='Team zur Liste hinzufügen')

            if add_clicked:
                t = str(st.session_state.get(f'atn_{lid}', '')).strip()
                c = str(st.session_state.get(f'acy_{lid}', '')).strip() or t
                if t:
                    teams.append((t, c))
                    ld['teams'] = teams
                    # Verein-Mapping speichern wenn Team aus DB gewählt wurde
                    _sel_v = st.session_state.get(f'cs_{lid}', SEARCH_PH)
                    if _sel_v not in (SEARCH_PH, MANUAL_OPT):
                        S.team_verein_map[t] = _sel_v
                    st.session_state[f'_exp_{lid}'] = True
                    st.session_state[f'_reset_{lid}'] = True
                    st.rerun()

            ld['teams'] = teams

            # ── Liga-ID umbenennen ────────────────────────────────────────────
            if new_lid and new_lid != lid and new_lid not in S.league_order:
                idx = S.league_order.index(lid)
                S.league_order[idx] = new_lid
                S.leagues[new_lid]  = S.leagues.pop(lid)
                for _state_dict in (S.dist_matrices, S.dst_per_liga, S.routing,
                                    S.weights, S.pinned, S.blocked, S.forced_home):
                    if lid in _state_dict:
                        _state_dict[new_lid] = _state_dict.pop(lid)
                for _club in S.clubs.values():
                    if lid in _club:
                        _club[new_lid] = _club.pop(lid)
                for _exp_key in (f'_exp_{lid}', f'_reset_{lid}'):
                    if _exp_key in st.session_state:
                        st.session_state[f'_exp_{new_lid}' if '_exp_' in _exp_key
                                         else f'_reset_{new_lid}'] = st.session_state.pop(_exp_key)
                st.session_state[f'lid_{i}'] = new_lid
                st.rerun()

            if len(teams) < 4:
                errors.append(f'**{ld["name"]}**: Mindestens 4 Teams erforderlich '
                               f'(aktuell: {len(teams)}).')

    total   = sum(len(S.leagues[l]['teams']) for l in S.league_order)
    n_ligen = len(S.league_order)
    if total > 48:
        st.warning(f'⚠ {total} Teams gesamt – Phase 2 kann mehrere Stunden dauern.')
    elif total > 0:
        lbl = 'Liga' if n_ligen == 1 else 'Ligen'
        st.caption(f'{total} Teams über {n_ligen} {lbl}.')

    for e in errors:
        st.error(e)

    go_back, go_fwd = _nav(back=False, fwd_disabled=bool(errors))
    if go_fwd:
        S.step = 1
        st.rerun()


# ═══════════════════════════════════════════════════════════════════════════════
# SCHRITT 1 – Distanzmatrizen
# ═══════════════════════════════════════════════════════════════════════════════
def _step1():
    st.header('2. Distanzmatrizen')
    st.info('Der Optimierer braucht die Fahrtstrecken (km) zwischen allen Teamstandorten jeder Liga. Die Werte können manuell eingetragen, aus einer Datei geladen oder automatisch über Google Maps berechnet werden.')

    method = st.radio('Methode zur Entfernungsermittlung',
        ['Manuell eingeben', 'Datei laden (CSV / Excel)', 'Google Maps API (automatisch)'],
        index={'manual': 0, 'file': 1, 'maps': 2}.get(S.dist_method, 0))
    S.dist_method = ['manual', 'file', 'maps'][
        ['Manuell eingeben', 'Datei laden (CSV / Excel)', 'Google Maps API (automatisch)'].index(method)]

    if S.dist_method == 'maps':
        S.api_key = st.text_input('Google Maps API-Key', S.api_key, type='password',
            help='Google Cloud Console → APIs & Dienste → „Distance Matrix API" aktivieren. '
                 'Unter Anmeldedaten einen API-Key erstellen und hier eintragen.')

    cache_dir = _HERE / '.cache'
    cache_dir.mkdir(exist_ok=True)
    errors = []

    for lid in S.league_order:
        ld    = S.leagues[lid]
        teams = [t for t, _ in ld['teams']]
        locs  = [loc for _, loc in ld['teams']]
        n     = len(teams)
        if n < 4:
            continue

        st.subheader(f'{ld["name"]}  ({n} Teams)')

        if S.dist_method == 'manual':
            # Bestehende Matrix holen oder neu anlegen
            mat = S.dist_matrices.get(lid)
            if mat is None or mat.shape[0] != n:
                mat = np.zeros((n, n), dtype=float)
            df = pd.DataFrame(mat.astype(float), index=teams, columns=teams)
            edited = st.data_editor(df, key=f'de_{lid}', width='stretch',
                column_config={t: st.column_config.NumberColumn(t, min_value=0,
                    format='%.0f') for t in teams})
            mat2 = edited.to_numpy(dtype=float).copy()
            # Obere Dreiecksmatrix auf untere spiegeln
            for r in range(n):
                for c in range(r + 1, n):
                    if mat2[r, c] > 0 and mat2[c, r] == 0:
                        mat2[c, r] = mat2[r, c]
                    elif mat2[c, r] > 0 and mat2[r, c] == 0:
                        mat2[r, c] = mat2[c, r]
            np.fill_diagonal(mat2, 0)
            S.dist_matrices[lid] = mat2
            st.caption('Tipp: Nur obere Hälfte ausfüllen – Spiegelung erfolgt automatisch.')

        elif S.dist_method == 'file':
            uploaded = st.file_uploader(f'Datei für {ld["name"]}', type=['csv', 'xlsx', 'xls'],
                key=f'df_{lid}', help='CSV oder Excel mit NxN-Matrix (Teams als Zeilen- und Spaltenköpfe)')
            if uploaded:
                tmp = Path(tempfile.gettempdir()) / uploaded.name
                tmp.write_bytes(uploaded.getvalue())
                from spielplan_multi.distances import load_distances_from_file
                mat = load_distances_from_file(str(tmp), teams)
                if mat is not None:
                    S.dist_matrices[lid] = mat
                    st.success(f'✓ {n}×{n}-Matrix geladen.')
                else:
                    st.error('Datei konnte nicht verarbeitet werden. Bitte Format prüfen.')
                    if lid not in S.dist_matrices:
                        errors.append(lid)
            if lid in S.dist_matrices:
                with st.expander('Matrix anzeigen'):
                    st.dataframe(pd.DataFrame(S.dist_matrices[lid], index=teams, columns=teams))
            elif not uploaded:
                errors.append(lid)

        else:  # maps
            if not S.api_key:
                st.warning('API-Key erforderlich.')
                errors.append(lid)
            elif lid in S.dist_matrices and np.sum(S.dist_matrices[lid]) > 0:
                st.success('✓ Bereits berechnet.')
                with st.expander('Matrix anzeigen'):
                    st.dataframe(pd.DataFrame(S.dist_matrices[lid], index=teams, columns=teams))
                if st.button(f'Neu berechnen', key=f'recalc_{lid}'):
                    del S.dist_matrices[lid]
                    st.rerun()
            else:
                # Adressen-Vorschau und Validierung
                empty_teams = [teams[i] for i, loc in enumerate(locs) if not loc.strip()]
                if empty_teams:
                    st.error(f'Fehlende Standort-Adresse für: {", ".join(empty_teams)}. '
                             f'Bitte in Schritt 1 (Ligen & Teams) ergänzen.')
                    errors.append(lid)
                else:
                    with st.expander('Adressen prüfen (werden an Google Maps gesendet)'):
                        for t, loc in zip(teams, locs):
                            st.caption(f'**{t}** → {loc}')
                    if st.button(f'Distanzen berechnen (Google Maps)', key=f'calc_{lid}', type='primary'):
                        _buf = io.StringIO()
                        _old_stdout = sys.stdout
                        sys.stdout = _buf
                        mat = None
                        try:
                            from spielplan_multi.distances import calculate_distance_matrix
                            with st.spinner('Google Maps API wird abgefragt…'):
                                mat = calculate_distance_matrix(locs, S.api_key,
                                    cache_dir / f'dist_{lid}.json')
                        finally:
                            sys.stdout = _old_stdout
                        _output = _buf.getvalue()
                        if mat is not None:
                            S.dist_matrices[lid] = mat
                            st.rerun()
                        else:
                            # Fehlerdetails aus captured stdout extrahieren
                            _details = '\n'.join(
                                l.strip() for l in _output.splitlines()
                                if '[XX]' in l or '[!!]' in l
                            )
                            st.error(
                                'Google Maps API-Abfrage fehlgeschlagen.\n\n' +
                                (_details if _details else
                                 'Mögliche Ursachen: falscher oder gesperrter API-Key, '
                                 '"Distance Matrix API" nicht aktiviert, kein Guthaben '
                                 'auf dem Google Cloud-Konto, oder Netzwerkproblem. '
                                 'Prüfen unter: Google Cloud Console → APIs & Dienste.')
                            )
                            errors.append(lid)
                    else:
                        errors.append(lid)

    if errors:
        st.error(f'Fehlende Distanzmatrix für: {", ".join(errors)}')

    go_back, go_fwd = _nav(fwd_disabled=bool(errors))
    if go_back:
        S.step = 0; st.rerun()
    if go_fwd:
        S.step = 2; st.rerun()


# ═══════════════════════════════════════════════════════════════════════════════
# SCHRITT 2 – Kalender & DST
# ═══════════════════════════════════════════════════════════════════════════════
def _rahmenterminplan_vorlage_bytes() -> bytes:
    """Erstellt eine Excel-Vorlage für den Rahmenterminplan."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    import datetime

    wb = Workbook()
    ws = wb.active
    ws.title = 'Rahmenterminplan'

    lids  = S.league_order if S.league_order else ['LIGA_1', 'LIGA_2']
    names = [S.leagues.get(l, {}).get('name', l) for l in lids]
    n_lig = len(lids)

    hdr_fill = PatternFill('solid', fgColor='1F4E79')
    hdr_font = Font(color='FFFFFF', bold=True)
    dst_fill = PatternFill('solid', fgColor='E2EFDA')
    tip_font = Font(italic=True, color='7F6000')

    # Zeile 1: Spaltenindex-Hinweis
    ws.cell(row=1, column=1, value='Tipp:').font = Font(bold=True, color='7F6000')
    idx_hints = ['Kalenderwochen-Spalte: 1 (Spalte B)']
    for i, nm in enumerate(names):
        idx_hints.append(f'{nm}: {2 + i} (Spalte {get_column_letter(3 + i)})')
    ws.cell(row=1, column=2, value='  |  '.join(idx_hints)).font = tip_font
    if n_lig >= 1:
        ws.merge_cells(f'B1:{get_column_letter(2 + n_lig)}1')

    # Zeile 2: Header
    for c, h in enumerate(['Nr.', 'Kalenderwoche'] + list(names), 1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal='center')

    # Beispieldaten: 16 Wochen ab KW 36 2026
    base_monday = datetime.date(2026, 8, 31)

    def _liga_sched(liga_idx: int, total: int = 16) -> list:
        skip = {1, 4, 7, 11, 13, 14}
        dst  = {2, 8} if liga_idx % 2 == 0 else {3, 9}
        result, md = [], 1
        for wi in range(total):
            if wi in skip:
                result.append([])
            elif wi in dst:
                result.append([md, md + 1]); md += 2
            else:
                result.append([md]); md += 1
        return result

    schedules = [_liga_sched(i) for i in range(n_lig)]

    for wi in range(16):
        row  = wi + 3
        mon  = base_monday + datetime.timedelta(weeks=wi)
        sun  = mon + datetime.timedelta(days=6)
        kw   = int(mon.isocalendar()[1])
        ws.cell(row=row, column=1, value=wi + 1).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=2, value=f'KW {kw} {mon.strftime("%d.%m.")} - {sun.strftime("%d.%m.%Y")}')
        for li, sched in enumerate(schedules):
            sts = sched[wi] if wi < len(sched) else []
            col  = li + 3
            if len(sts) == 2:
                cell = ws.cell(row=row, column=col, value=f'{sts[0]} & {sts[1]}')
                cell.fill = dst_fill
            else:
                cell = ws.cell(row=row, column=col, value=str(sts[0]) if sts else '')
            cell.alignment = Alignment(horizontal='center')

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 30
    for i in range(n_lig):
        ws.column_dimensions[get_column_letter(3 + i)].width = max(18, len(names[i]) + 4)
    ws.freeze_panes = 'A3'

    # Hinweis-Sheet
    ws2 = wb.create_sheet('Hinweise')
    ws2['A1'] = 'Hinweise zur Rahmenterminplan-Vorlage'
    ws2['A1'].font = Font(bold=True, size=13)
    lines = [(2, 'Spaltenindizes im Tool eintragen:'),
             (3, '  • Kalenderwochen-Spalte (Spalte B): Index 1')]
    for i, (lid, nm) in enumerate(zip(lids, names)):
        lines.append((4 + i, f'  • {nm} (Spalte {get_column_letter(3 + i)}): Index {2 + i}'))
    r = 4 + n_lig + 1
    lines += [
        (r,      ''),
        (r + 1,  'Erlaubte Inhalte in den Spieltag-Spalten:'),
        (r + 2,  '  • Einzelspieltag:  7'),
        (r + 3,  '  • Doppelspieltag:  6 & 7   oder   6/7'),
        (r + 4,  '  • Leer oder sonstiger Text = kein Spieltag (wird ignoriert)'),
        (r + 5,  ''),
        (r + 6,  'Format der Kalenderwochen-Spalte:'),
        (r + 7,  '  • Nur Zahl:   36'),
        (r + 8,  '  • Mit Datum:  KW 36 02.09. - 08.09.2026'),
        (r + 9,  ''),
        (r + 10, 'Doppelspieltage (grün hinterlegt) = beide Spieltage desselben Wochenendes.'),
        (r + 11, 'Das Heimrecht gilt dann für das gesamte Wochenende.'),
    ]
    for row_n, text in lines:
        ws2.cell(row=row_n, column=1, value=text)
    ws2.column_dimensions['A'].width = 65

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _weekend_dates_for_kw(week_start: str, week_end: str):
    """Gibt (samstag_str, sonntag_str, bereich_str) für eine KW zurück.

    Erwartet week_end als 'DD.MM.YYYY' oder 'YYYY-MM-DD' (ISO-Woche endet Sonntag).
    Samstag = Sonntag - 1 Tag.
    Bereich z. B. '07.-13.09.2026' oder '29.08.-04.09.2026'.
    """
    from datetime import datetime, timedelta
    sun = None
    for fmt in ('%d.%m.%Y', '%Y-%m-%d', '%d.%m.%y'):
        for src in (week_end, week_start):
            s = str(src).strip()
            if not s or s == 'nan':
                continue
            try:
                sun = datetime.strptime(s, fmt)
                # Wenn week_start übergeben: auf Sonntag vorwärts rechnen
                if src == week_start and sun.weekday() != 6:
                    sun = sun + timedelta(days=(6 - sun.weekday()))
                break
            except ValueError:
                continue
        if sun:
            break
    if sun is None:
        return '', '', ''
    sat = sun - timedelta(days=1)
    sat_s = sat.strftime('%d.%m.%Y')
    sun_s = sun.strftime('%d.%m.%Y')
    # Bereichsstring: gleicher Monat → '07.-13.09.2026', sonst '29.08.-04.09.2026'
    ws_dt = sun - timedelta(days=6)  # Montag der Woche
    if ws_dt.month == sun.month:
        bereich = f'{ws_dt.day:02d}.-{sun.day:02d}.{sun.month:02d}.{sun.year}'
    else:
        bereich = f'{ws_dt.day:02d}.{ws_dt.month:02d}.-{sun.day:02d}.{sun.month:02d}.{sun.year}'
    return sat_s, sun_s, bereich


def _apply_weekend_dates(rows: list, spieltage_info: dict) -> None:
    """Setzt Datum-Felder in rows basierend auf KW-Wochenend-Logik (in-place).

    DST-Paar (zwei aufeinanderfolgende Spieltage in gleicher KW):
      → erster Spieltag = Samstag, zweiter = Sonntag
    Einzelspieltag:
      → Datum = Wochenbereich (z. B. '07.-13.09.2026')
    """
    kw_to_sts: dict = {}
    for row in rows:
        kw = row.get('kw')
        if kw is not None:
            kw_to_sts.setdefault(kw, []).append(row['spieltag'])
    for kw, sts in kw_to_sts.items():
        info = spieltage_info.get(sts[0], {})
        sat_s, sun_s, bereich = _weekend_dates_for_kw(
            info.get('week_start', ''), info.get('week_end', ''))
        sorted_sts = sorted(sts)
        if (len(sorted_sts) == 2
                and sorted_sts[1] == sorted_sts[0] + 1
                and sat_s):
            rows[sorted_sts[0] - 1]['date'] = sat_s
            rows[sorted_sts[1] - 1]['date'] = sun_s
        else:
            for st_nr in sorted_sts:
                rows[st_nr - 1]['date'] = bereich if bereich else ''


def _cal_table_to_kw_compat() -> None:
    """Leitet S.kw_compat und S.dst_per_liga aus S.cal_table ab."""
    kw_compat: dict = {}
    for lid in S.league_order:
        rows = S.cal_table.get(lid, [])
        ld = S.leagues.get(lid, {})
        _, gpd = _get_n_rounds_gpd(ld)
        kw_to_sts: dict = {}
        for row in rows:
            kw = row.get('kw')
            if kw is not None:
                kw_to_sts.setdefault(kw, []).append(row['spieltag'])
        for kw, sts in kw_to_sts.items():
            kw_compat.setdefault(kw, {})[lid] = sorted(sts)
        S.dst_per_liga[lid] = [] if gpd > 1 else _detect_dst_blocks(rows)
    S.kw_compat = kw_compat


def _migrate_cal_table_from_kw_compat() -> None:
    """Migriert alte S.kw_compat-Daten in S.cal_table (für ältere gespeicherte Sitzungen)."""
    for lid in S.league_order:
        if lid in S.cal_table:
            continue
        ld = S.leagues.get(lid, {})
        n_md = _calc_n_matchdays(ld)
        rows = [{'spieltag': i + 1, 'kw': None, 'date': ''} for i in range(n_md)]
        for kw, kw_data in (S.kw_compat or {}).items():
            if lid in kw_data:
                for st_nr in kw_data[lid]:
                    if 1 <= st_nr <= n_md:
                        rows[st_nr - 1]['kw'] = int(kw)
        S.cal_table[lid] = rows


def _step2():
    st.header('3. Kalender & Doppelspieltage (DST)')
    st.info(
        'Ein **Doppelspieltag (DST)** sind zwei direkt aufeinanderfolgende Spieltage '
        '(typisch: Samstag + Sonntag), bei denen jedes Team das **gleiche Heimrecht** '
        'an beiden Tagen hat – also entweder beide Tage zuhause oder beide auswärts. '
        'Das vermeidet unnötige Fahrten zwischen den Spielen. '
        'Beim Turniertag-Format entfallen DST-Blöcke.'
    )

    # ── Excel-Import (optional, klappt Tabellen vor) ──────────────────────────
    with st.expander('Rahmenterminplan aus Excel laden (optional)'):
        st.caption(
            'Importiert Spieltag-Kalenderwochen aus einer Excel-Datei und befüllt die '
            'Tabellen unten automatisch. Ohne Import können die KW-Spalten manuell eingetragen werden.'
        )
        _col_tpl, _col_upl = st.columns([1, 2])
        with _col_tpl:
            st.markdown('**Vorlage herunterladen**')
            st.caption('Noch keine Datei? Vorlage herunterladen, befüllen und rechts hochladen.')
            st.download_button(
                '⬇ Rahmenterminplan-Vorlage',
                data=_rahmenterminplan_vorlage_bytes(),
                file_name='Rahmenterminplan_Vorlage.xlsx',
                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                width='stretch',
                key='dl_cal_tmpl',
            )
        with _col_upl:
            st.markdown('**Datei hochladen**')
            uploaded = st.file_uploader(
                'Rahmenplan-Excel', type=['xlsx', 'xls'], key='cal_upload',
                label_visibility='collapsed',
            )
        if uploaded:
            tmp = Path(tempfile.gettempdir()) / uploaded.name
            tmp.write_bytes(uploaded.getvalue())
            S.cal_path = str(tmp)
        if S.cal_path:
            from spielplan_multi.calendar_parser import preview_columns
            prev = preview_columns(S.cal_path)
            if prev is not None:
                with st.expander('Spaltenvorschau (erste 8 Zeilen)'):
                    st.dataframe(prev.head(8))
                st.caption(
                    'Spalten werden von links gezählt: A = 0, B = 1, C = 2, … '
                    'Trage für jede Liga die Spaltennummer ein, in der die Spieltag-Nummern stehen. '
                    'Gültige Zellinhalte: **"7"** (einzelner Spieltag) oder **"6 & 7"** / **"6/7"** '
                    '(Doppelspieltag – wird automatisch als DST erkannt).'
                )
                col_mapping = {}
                _cols = st.columns(max(1, len(S.league_order)))
                for i, lid in enumerate(S.league_order):
                    with _cols[i]:
                        col_mapping[lid] = int(st.number_input(
                            S.leagues[lid]['name'], 0, 50,
                            int(st.session_state.get(f'col_{lid}', 2)),
                            key=f'col_{lid}',
                            help='Spalte A = 0, B = 1, C = 2 usw.'))
                kw_col = int(st.number_input(
                    'Spalte mit Kalenderwochen-Nummern (z. B. "KW 37" oder "37")',
                    0, 50, int(st.session_state.get('kw_col', 1)),
                    key='kw_col',
                    help='Spalte A = 0, B = 1, C = 2 usw. Meistens Spalte B (= 1)'))

                if st.button('Kalender laden', type='primary', key='load_cal'):
                    from spielplan_multi.calendar_parser import parse_rahmenterminplan
                    cal = parse_rahmenterminplan(S.cal_path, col_mapping,
                        kw_col=kw_col,
                        date_from_col=kw_col,
                        date_to_col=kw_col)
                    if cal:
                        spieltage_from_cal = cal.get('spieltage', {})
                        for lid in S.league_order:
                            ld = S.leagues.get(lid, {})
                            n_md = _calc_n_matchdays(ld)
                            rows = [{'spieltag': i + 1, 'kw': None, 'date': ''}
                                    for i in range(n_md)]
                            for st_nr, info in spieltage_from_cal.get(lid, {}).items():
                                if 1 <= st_nr <= n_md:
                                    rows[st_nr - 1]['kw'] = int(info['kw'])
                            # Wochenend-Datums-Logik: DST → Sa/So, Einzelspiel → Bereich
                            _apply_weekend_dates(rows, spieltage_from_cal.get(lid, {}))
                            S.cal_table[lid] = rows
                            # Tabellen-Widget zurücksetzen, damit neue Daten angezeigt werden
                            editor_key = f'cal_editor_{lid}'
                            if editor_key in st.session_state:
                                del st.session_state[editor_key]
                        _cal_table_to_kw_compat()
                        n_st = sum(len(v) for v in spieltage_from_cal.values())
                        n_dst = sum(len(b) for b in S.dst_per_liga.values())
                        st.success(
                            f'Kalender geladen: {n_st} Spieltage, {n_dst} DST-Blöcke erkannt.')
                        st.rerun()
                    else:
                        st.error(
                            'Kalender konnte nicht gelesen werden. '
                            'Bitte prüfen: Ist es eine gültige Excel-Datei (.xlsx/.xls)? '
                            'Stimmen die Spaltennummern?'
                        )

    # ── Kalender-Tabellen je Liga ─────────────────────────────────────────────
    _migrate_cal_table_from_kw_compat()

    st.subheader('Spieltage & Kalenderwochen')
    st.caption(
        'Trage für jeden Spieltag die **Kalenderwoche (KW)** ein. '
        'Zwei aufeinanderfolgende Spieltage in derselben KW werden automatisch als '
        '**DST-Block** erkannt. '
        'Optional: **Datum** eintragen (z. B. 20.09.2025) für einen fixen Termin – '
        'leer lassen = Team wählt Termin frei innerhalb der KW.'
    )

    _any_changed = False
    for lid in S.league_order:
        ld = S.leagues[lid]
        n_t = len(ld['teams'])
        if n_t < 4:
            continue
        n_rounds, gpd = _get_n_rounds_gpd(ld)
        n_md = _calc_n_matchdays(ld)

        if gpd > 1:
            S.dst_per_liga[lid] = []
            n_days = n_md
            st.info(f'**{ld["name"]}** (Turniertag, {gpd} Spiele/Tag): '
                    f'{n_days} Turniertage – keine DST-Blöcke erforderlich.')
            continue

        # Tabelle initialisieren oder bei Größenänderung anpassen
        existing = S.cal_table.get(lid, [])
        if len(existing) != n_md:
            new_rows = [{'spieltag': i + 1, 'kw': None, 'date': ''} for i in range(n_md)]
            for old_row in existing:
                st_nr = old_row.get('spieltag', 0)
                if 1 <= st_nr <= n_md:
                    new_rows[st_nr - 1] = old_row
            S.cal_table[lid] = new_rows
            existing = new_rows

        # DST-Erkennung für Anzeige-Spalte
        kw_count: dict = {}
        for row in existing:
            kw = row.get('kw')
            if kw is not None:
                kw_count[kw] = kw_count.get(kw, 0) + 1

        display_rows = []
        for row in existing:
            kw = row.get('kw')
            if kw is None:
                dst_col = ''
            elif kw_count.get(kw, 0) == 2:
                dst_col = 'DST'
            elif kw_count.get(kw, 0) > 2:
                dst_col = '> 2 ⚠'
            else:
                dst_col = ''
            display_rows.append({
                'ST': row['spieltag'],
                'KW': pd.NA if kw is None else kw,
                'Datum': row.get('date', ''),
                'DST': dst_col,
            })

        df = pd.DataFrame(display_rows)
        df['KW'] = df['KW'].astype(pd.Int64Dtype())

        assigned = sum(1 for r in existing if r.get('kw') is not None)
        dst_blocks = _detect_dst_blocks(existing)
        header_parts = [f'{assigned}/{n_md} KW vergeben']
        if dst_blocks:
            header_parts.append(f'{len(dst_blocks)} DST erkannt')

        with st.expander(
                f'{ld["name"]}  ({n_md} Spieltage · {" · ".join(header_parts)})',
                expanded=True):
            edited_df = st.data_editor(
                df,
                column_config={
                    'ST': st.column_config.NumberColumn(
                        'ST', disabled=True, width='small'),
                    'KW': st.column_config.NumberColumn(
                        'KW', min_value=1, max_value=53, step=1, width='small',
                        help='Kalenderwoche 1–53. Gleiche KW bei zwei aufeinanderfolgenden '
                             'Spieltagen → automatisch DST-Block.'),
                    'Datum': st.column_config.TextColumn(
                        'Datum (opt.)',
                        help='Konkretes Datum des Spieltags, z. B. 20.09.2025. '
                             'Leer lassen = Termin frei wählbar innerhalb der KW.',
                        width='medium'),
                    'DST': st.column_config.TextColumn(
                        'DST', disabled=True, width='small'),
                },
                hide_index=True,
                width='stretch',
                key=f'cal_editor_{lid}',
                num_rows='fixed',
            )

            # Geänderte Werte in Session speichern
            new_rows = []
            for _, erow in edited_df.iterrows():
                kw_val = erow['KW']
                try:
                    kw_val = int(kw_val) if pd.notna(kw_val) else None
                except (ValueError, TypeError):
                    kw_val = None
                date_raw = erow['Datum']
                date_val = str(date_raw).strip() if pd.notna(date_raw) and str(date_raw).strip() else ''
                new_rows.append({
                    'spieltag': int(erow['ST']),
                    'kw': kw_val,
                    'date': date_val,
                })
            if new_rows != existing:
                S.cal_table[lid] = new_rows
                _any_changed = True

            if dst_blocks:
                st.caption('Erkannte DST-Blöcke: ' +
                           ', '.join(f'ST {d1}+{d2}' for d1, d2 in dst_blocks))
            missing = n_md - assigned
            if missing > 0:
                st.caption(f'ℹ {missing} Spieltage ohne KW – Termin frei wählbar.')

    if _any_changed:
        _cal_table_to_kw_compat()

    go_back, go_fwd = _nav()
    if go_back:
        S.step = 1; st.rerun()
    if go_fwd:
        S.step = 3; st.rerun()


# ═══════════════════════════════════════════════════════════════════════════════
# SCHRITT 3 – Routing & Gewichte
# ═══════════════════════════════════════════════════════════════════════════════
def _step3():
    st.header('4. Routing & Optimierungs-Gewichte')

    _routing_active = any(v[0] for v in S.routing.values())
    if _routing_active:
        st.session_state.setdefault('_exp_routing', True)
    elif '_exp_routing' not in st.session_state:
        st.session_state['_exp_routing'] = False
    with st.expander('Fahrtwege bei Doppelspieltagen begrenzen (optional)', expanded=st.session_state['_exp_routing']):
        st.caption('Verhindert, dass Teams zwischen den zwei Spielen eines Doppelwochenendes unnötig weite Umwege fahren.')
        for lid in S.league_order:
            ld = S.leagues[lid]
            if not S.dst_per_liga.get(lid):
                continue
            cur_apply, cur_pct = S.routing.get(lid, (False, 25))
            enabled = st.checkbox(f'Routing für {ld["name"]}', cur_apply, key=f'rt_{lid}')
            pct     = cur_pct
            if enabled:
                pct = st.slider(f'Erlaubter Mehraufwand (%)', 1, 100, max(1, cur_pct), 5,
                    key=f'rp_{lid}', help='Wie viel Prozent Mehrkilometer darf ein Team bei einem Doppelwochenende in Kauf nehmen? 1 % = fast kein Umweg erlaubt · 25 % = bis zu 25 % mehr als der direkte Weg')
            S.routing[lid] = (enabled, pct)

    st.subheader('Optimierungsziele gewichten')
    st.caption('Lege fest, wie wichtig dir jedes Ziel ist. 0 = wird nicht berücksichtigt · 5 = normal · 10 = höchste Priorität')

    same = st.checkbox('Gleiche Gewichte für alle Ligen', S.same_weights, key='samew')
    S.same_weights = same

    _W_DEFAULTS = {'dst_eff': 0.0}

    def _weight_inputs(key_prefix: str, existing: dict) -> dict:
        vals = {}
        for key, (label, tip) in WEIGHT_LABELS.items():
            default = float(existing.get(key, _W_DEFAULTS.get(key, 5.0)))
            vals[key] = st.slider(label, 0.0, 10.0, default,
                0.5, key=f'w_{key_prefix}_{key}', help=tip)
        return vals

    if same:
        common = _weight_inputs('all', S.weights.get('__common__', {}))
        S.weights['__common__'] = common
        for lid in S.league_order:
            S.weights[lid] = common.copy()
    else:
        for lid in S.league_order:
            with st.expander(S.leagues[lid]['name']):
                S.weights[lid] = _weight_inputs(lid, S.weights.get(lid, {}))

    if len(S.league_order) >= 2:
        st.subheader('Heimspiel-Koordination (Co-Home)')
        st.caption('Wie wichtig ist es, dass Vereine mit Teams in mehreren Ligen ihre Heimspiele in denselben Kalenderwochen haben?')
        S.w_cohome = st.slider('Co-Home-Gewicht', 0.0, 10.0, float(S.w_cohome), 0.5, key='wch')

    go_back, go_fwd = _nav()
    if go_back:
        S.step = 2; st.rerun()
    if go_fwd:
        S.step = 4; st.rerun()


# ═══════════════════════════════════════════════════════════════════════════════
# SCHRITT 4 – Pflichtspiele
# ═══════════════════════════════════════════════════════════════════════════════
def _step4():
    st.header('5. Pflichtspiele (optional)')
    st.info('Bestimmte Begegnungen können auf einen festen Spieltag und/oder ein Heimrecht festgelegt werden.')

    has_any = any(S.pinned.get(lid) for lid in S.league_order)
    if not has_any:
        if st.checkbox('Keine Pflichtspiele – Schritt überspringen', True, key='skip_pin'):
            go_back, go_fwd = _nav()
            if go_back:
                S.step = 3; st.rerun()
            if go_fwd:
                S.step = 5; st.rerun()
            return

    for lid in S.league_order:
        ld    = S.leagues[lid]
        teams = [t for t, _ in ld['teams']]
        if len(teams) < 4:
            continue
        n_rounds, _ = _get_n_rounds_gpd(ld)
        n_md  = _calc_n_matchdays(ld)
        pinned: List[dict] = S.pinned.get(lid, [])

        _ekey = f'_exp_pin_{lid}'
        # Immer offen wenn Einträge vorhanden (z. B. nach Import), sonst letzten Stand halten
        if len(pinned) > 0:
            st.session_state[_ekey] = True
        elif _ekey not in st.session_state:
            st.session_state[_ekey] = False
        with st.expander(f'{ld["name"]}  ({len(pinned)} Pflichtspiele)',
                         expanded=st.session_state[_ekey]):
            for idx_p, pm in enumerate(pinned):
                c1, c2, c3, c4 = st.columns([3, 2, 2, 1])
                c1.write(f'{pm["teamA"]} – {pm["teamB"]}')
                c2.write(f'Spieltag {pm["day"]}')
                c3.write(f'Heim: {pm.get("home") or "beliebig"}')
                if c4.button('✕', key=f'delp_{lid}_{idx_p}'):
                    pinned.pop(idx_p)
                    S.pinned[lid] = pinned
                    st.session_state[_ekey] = True
                    st.rerun()

            st.caption('Neues Pflichtspiel:')
            pa, pb, pc, pd_ = st.columns([3, 3, 2, 1])
            ta = pa.selectbox('Team A', teams, key=f'pa_{lid}')
            tb = pb.selectbox('Team B', [t for t in teams if t != ta], key=f'pb_{lid}')
            day = pc.number_input('Spieltag', 1, n_md, 1, key=f'pd_{lid}')
            home_sel = pd_.selectbox('Heim', [ta, tb, 'beliebig'], key=f'ph_{lid}')
            if st.button('Hinzufügen', key=f'addp_{lid}'):
                _new_pin = {'teamA': ta, 'teamB': tb, 'day': int(day),
                            'home': None if home_sel == 'beliebig' else home_sel}
                _is_dup = any(
                    p.get('teamA') == ta and p.get('teamB') == tb and int(p.get('day', 0)) == int(day)
                    for p in pinned
                )
                if _is_dup:
                    st.warning(f'Pflichtspiel {ta} – {tb} an Spieltag {int(day)} ist bereits eingetragen.')
                else:
                    pinned.append(_new_pin)
                    S.pinned[lid] = pinned
                    st.session_state[_ekey] = True
                    st.rerun()
        S.pinned[lid] = pinned

    go_back, go_fwd = _nav()
    if go_back:
        S.step = 3; st.rerun()
    if go_fwd:
        S.step = 5; st.rerun()


# ═══════════════════════════════════════════════════════════════════════════════
# SCHRITT 5 – Sperrtage & Pflichttage
# ═══════════════════════════════════════════════════════════════════════════════
def _step5():
    st.header('6. Heim-Einschränkungen (optional)')

    has_any_blk = any(S.blocked.get(lid) for lid in S.league_order)
    has_any_frc = any(S.forced_home.get(lid) for lid in S.league_order)
    if not has_any_blk and not has_any_frc:
        if st.checkbox('Keine Heim-Einschränkungen – Schritt überspringen', True, key='skip_blk'):
            go_back, go_fwd = _nav()
            if go_back:
                S.step = 4; st.rerun()
            if go_fwd:
                S.step = 6; st.rerun()
            return

    # ── Sperrtage ────────────────────────────────────────────────────────────
    st.subheader('Heimspiel-Sperrtage')
    st.info('Spieltage, an denen ein Team kein Heimspiel austragen darf (z. B. Hallensperrung).')
    for lid in S.league_order:
        ld    = S.leagues[lid]
        teams = [t for t, _ in ld['teams']]
        if len(teams) < 4:
            continue
        n_rounds, _ = _get_n_rounds_gpd(ld)
        n_md    = _calc_n_matchdays(ld)
        blocked = S.blocked.get(lid, {})

        _ekey = f'_exp_blk_{lid}'
        n_blk = sum(len(v) for v in blocked.values())
        if n_blk > 0:
            st.session_state[_ekey] = True
        elif _ekey not in st.session_state:
            st.session_state[_ekey] = False
        with st.expander(f'{ld["name"]}  ({n_blk} Sperrtage)', expanded=st.session_state[_ekey]):
            for team, days_list in list(blocked.items()):
                c1, c2 = st.columns([5, 1])
                c1.write(f'**{team}**: Spieltag(e) {days_list}')
                if c2.button('✕', key=f'delb_{lid}_{team}'):
                    del blocked[team]
                    S.blocked[lid] = blocked
                    st.session_state[_ekey] = True
                    st.rerun()

            ba, bb, bc = st.columns([3, 4, 1])
            t_sel  = ba.selectbox('Team', teams, key=f'bt_{lid}')
            d_raw  = bb.text_input('Spieltage (kommagetrennt)', '', key=f'bd_{lid}',
                placeholder='z. B.  3, 7, 12')
            if bc.button('+ Sperren', key=f'addb_{lid}'):
                try:
                    new_days = sorted({int(x.strip()) for x in d_raw.split(',') if x.strip()})
                    if not new_days or not all(1 <= d <= n_md for d in new_days):
                        st.error(f'Bitte gültige Spieltagnummern zwischen 1 und {n_md} eingeben.')
                    else:
                        blocked[t_sel] = sorted(set(blocked.get(t_sel, [])) | set(new_days))
                        S.blocked[lid] = blocked
                        st.session_state[_ekey] = True
                        st.rerun()
                except Exception:
                    st.error('Bitte nur Zahlen eingeben, getrennt durch Kommas – zum Beispiel: 3, 7, 12')
        S.blocked[lid] = blocked

    st.divider()

    # ── Pflichtheim ──────────────────────────────────────────────────────────
    st.subheader('Heimspiel-Pflichttage')
    st.info('Spieltage, an denen ein Team zwingend Heimrecht haben muss '
            '(z. B. weil die Halle nur an bestimmten Terminen verfügbar ist).')
    for lid in S.league_order:
        ld    = S.leagues[lid]
        teams = [t for t, _ in ld['teams']]
        if len(teams) < 4:
            continue
        n_rounds, _ = _get_n_rounds_gpd(ld)
        n_md    = _calc_n_matchdays(ld)
        forced  = S.forced_home.get(lid, {})

        _ekey_f = f'_exp_frc_{lid}'
        n_frc = sum(len(v) for v in forced.values())
        if n_frc > 0:
            st.session_state[_ekey_f] = True
        elif _ekey_f not in st.session_state:
            st.session_state[_ekey_f] = False
        with st.expander(f'{ld["name"]}  ({n_frc} Pflichttage)', expanded=st.session_state[_ekey_f]):
            for team, days_list in list(forced.items()):
                c1, c2 = st.columns([5, 1])
                c1.write(f'**{team}**: Spieltag(e) {days_list}')
                if c2.button('✕', key=f'delf_{lid}_{team}'):
                    del forced[team]
                    S.forced_home[lid] = forced
                    st.session_state[_ekey_f] = True
                    st.rerun()

            fa, fb, fc = st.columns([3, 4, 1])
            ft_sel = fa.selectbox('Team', teams, key=f'ft_{lid}')
            fd_raw = fb.text_input('Spieltage (kommagetrennt)', '', key=f'fd_{lid}',
                placeholder='z. B.  1, 5, 9')
            if fc.button('+ Pflichtheim', key=f'addf_{lid}'):
                try:
                    new_days = sorted({int(x.strip()) for x in fd_raw.split(',') if x.strip()})
                    if not new_days or not all(1 <= d <= n_md for d in new_days):
                        st.error(f'Bitte gültige Spieltagnummern zwischen 1 und {n_md} eingeben.')
                    else:
                        forced[ft_sel] = sorted(set(forced.get(ft_sel, [])) | set(new_days))
                        S.forced_home[lid] = forced
                        st.session_state[_ekey_f] = True
                        st.rerun()
                except Exception:
                    st.error('Bitte nur Zahlen eingeben, getrennt durch Kommas – zum Beispiel: 1, 5, 9')
        S.forced_home[lid] = forced

    # Widerspruchs-Check: gleicher Spieltag in blocked UND forced_home
    for lid in S.league_order:
        blk = S.blocked.get(lid, {})
        frc = S.forced_home.get(lid, {})
        for team in set(blk) & set(frc):
            conflict = set(blk[team]) & set(frc[team])
            if conflict:
                _lname = S.leagues[lid].get('name', lid)
                st.warning(
                    f'**{_lname} – {team}**: '
                    f'Spieltag(e) {sorted(conflict)} sind gleichzeitig als '
                    f'Sperrtag und Pflichtheim eingetragen – das ist ein Widerspruch '
                    f'und führt zu einer unlösbaren Konfiguration.'
                )

    go_back, go_fwd = _nav()
    if go_back:
        S.step = 4; st.rerun()
    if go_fwd:
        S.step = 6; st.rerun()


# ═══════════════════════════════════════════════════════════════════════════════
# SCHRITT 6 – Co-Home
# ═══════════════════════════════════════════════════════════════════════════════

def _normalize_club_name(name: str) -> str:
    """Entfernt typische Nummerierungs- und Gendersuffixe aus Teamnamen."""
    s = name.strip()
    s = re.sub(r'\s+(?:I{1,4}|VI{0,3}|IX|IV|\d{1,2})\s*$', '', s, flags=re.IGNORECASE)
    s = re.sub(r'\s+(?:Damen|Herren|Mixed)\s*$', '', s, flags=re.IGNORECASE)
    s = re.sub(r'\s+\([mwdDHh]+\)\s*$', '', s)
    s = re.sub(r'\s+e\.?\s*V\.?\s*$', '', s, flags=re.IGNORECASE)
    return s.strip()


def _autodetect_cohome(league_order: list, leagues: dict,
                       club_records: list,
                       team_verein_map: dict | None = None) -> Dict[str, Dict[str, str]]:
    """Erkennt Vereine mit Teams in mehreren Ligen automatisch.

    Methode 1: team_verein_map aus Session State (beim Hinzufügen aus DB befüllt)
    Methode 2: Vereinsdatenbank (verein-Feld)
    Methode 3: Normalisierter Teamname (Suffixe I/II/2/Damen/... entfernen)

    Gibt zurück: {vereinsname: {lid: teamname}} für Vereine in ≥2 Ligen.
    """
    # Basis: DB-Mapping teamname → verein
    team_to_verein: Dict[str, str] = {}
    for r in club_records:
        v, t = r.get('verein', ''), r.get('teamname', '')
        if v and t:
            team_to_verein[t] = v
    # Session-State-Mapping hat Vorrang (enthält auch manuell angepasste Teamnamen)
    if team_verein_map:
        team_to_verein.update(team_verein_map)

    groups: Dict[str, Dict[str, str]] = {}
    for lid in league_order:
        for team_name, _ in leagues.get(lid, {}).get('teams', []):
            norm = _normalize_club_name(team_name)
            key = (team_to_verein.get(team_name)
                   or team_to_verein.get(norm)
                   or norm)
            if not key:
                continue
            if key not in groups:
                groups[key] = {}
            if lid not in groups[key]:
                groups[key][lid] = team_name

    return {v: lm for v, lm in groups.items() if len(lm) >= 2}


def _step6():
    st.header('7. Heimspiel-Koordination für Vereine in mehreren Ligen (optional)')

    if len(S.league_order) < 2:
        st.info('Dieser Schritt ist nur relevant, wenn du mehrere Ligen konfiguriert hast.')
        go_back, go_fwd = _nav()
        if go_back:
            S.step = 5; st.rerun()
        if go_fwd:
            S.step = 7; st.rerun()
        return

    st.info('Hat ein Verein Teams in mehreren Ligen (z. B. Damen und Herren), '
            'kann der Optimierer deren Heimspiele möglichst in dieselben Kalenderwochen legen – '
            'damit Halle, Personal und Infrastruktur gemeinsam genutzt werden können.')

    club_records = load_club_db()
    clubs: Dict[str, Dict[str, str]] = dict(S.clubs)

    # ── Automatische Erkennung ─────────────────────────────────────────────────
    all_detected = _autodetect_cohome(S.league_order, S.leagues, club_records, S.team_verein_map)
    new_detected  = {v: lm for v, lm in all_detected.items() if v not in clubs}

    st.subheader('Automatisch erkannte Vereine')
    if new_detected:
        st.caption(f'{len(new_detected)} Verein(e) mit Teams in mehreren Ligen automatisch erkannt. '
                   'Häkchen setzen und auf "Ausgewählte übernehmen" klicken.')
        selected: Dict[str, Dict[str, str]] = {}
        for verein_name, liga_map in new_detected.items():
            c1, c2 = st.columns([0.08, 0.92])
            with c1:
                checked = st.checkbox(verein_name, value=True, key=f'auto_chk_{verein_name}',
                                      label_visibility='collapsed')
            with c2:
                lines = [f'**{verein_name}**']
                for lid, team in liga_map.items():
                    liga_nm = S.leagues.get(lid, {}).get('name', lid)
                    lines.append(f'&nbsp;&nbsp;· {liga_nm}: {team}')
                st.markdown('  \n'.join(lines), unsafe_allow_html=True)
            if checked:
                selected[verein_name] = liga_map

        if st.button('Ausgewählte übernehmen', type='primary', key='accept_auto_cohome',
                     disabled=not selected):
            clubs.update(selected)
            S.clubs = clubs
            st.success(f'{len(selected)} Verein(e) übernommen.')
            st.rerun()
    else:
        st.caption('Keine weiteren Vereine mit Teams in mehreren Ligen erkannt.')

    # ── Konfigurierte Vereine ──────────────────────────────────────────────────
    if clubs:
        st.subheader(f'Konfigurierte Vereine ({len(clubs)})')
        for club_name in list(clubs.keys()):
            with st.expander(f'📍 {club_name}  ({len(clubs[club_name])} Teams)'):
                for lid in list(clubs[club_name].keys()):
                    team = clubs[club_name][lid]
                    liga_nm = S.leagues.get(lid, {}).get('name', lid)
                    ca, cb = st.columns([5, 1])
                    with ca:
                        st.write(f'{liga_nm}: **{team}**')
                    with cb:
                        if st.button('✕', key=f'del_team_{club_name}_{lid}',
                                     help='Team aus diesem Verein entfernen'):
                            del clubs[club_name][lid]
                            if len(clubs[club_name]) < 2:
                                del clubs[club_name]
                            S.clubs = clubs
                            st.rerun()
                if st.button('Verein entfernen', key=f'delc_{club_name}'):
                    del clubs[club_name]
                    S.clubs = clubs
                    st.rerun()

    # ── Manuell hinzufügen ─────────────────────────────────────────────────────
    with st.expander('Verein manuell hinzufügen'):
        new_club = st.text_input('Vereinsname', '', key='new_club')
        if new_club:
            liga_map2: Dict[str, str] = {}
            for lid in S.league_order:
                opts = ['(kein Team)'] + [t for t, _ in S.leagues[lid]['teams']]
                sel  = st.selectbox(f'Team in {S.leagues[lid]["name"]}', opts,
                                    key=f'ct_{lid}')
                if sel != '(kein Team)':
                    liga_map2[lid] = sel
            if st.button('Speichern', key='save_club', type='primary'):
                if len(liga_map2) >= 2:
                    if new_club in clubs:
                        st.warning(f'"{new_club}" existiert bereits – Einträge werden überschrieben.')
                    clubs[new_club] = liga_map2
                    S.clubs = clubs
                    st.success(f'"{new_club}" gespeichert.')
                    st.rerun()
                else:
                    st.error('Mindestens 2 Teams in verschiedenen Ligen erforderlich.')

    go_back, go_fwd = _nav()
    if go_back:
        S.step = 5; st.rerun()
    if go_fwd:
        S.step = 7; st.rerun()


# ═══════════════════════════════════════════════════════════════════════════════
# SCHRITT 7 – Solver-Konfiguration
# ═══════════════════════════════════════════════════════════════════════════════
def _step7():
    st.header('8. Solver-Konfiguration')
    st.info(
        'Hier legst du fest, wie viel Rechenzeit der Optimierer bekommt. '
        'Mehr Zeit führt in der Regel zu besseren Spielplänen – '
        'du kannst also je nach Dringlichkeit zwischen Schnelligkeit und Qualität abwägen.'
    )

    sol  = dict(S.solver)
    n_lig = len(S.league_order)
    _ligen_wort = 'Liga' if n_lig == 1 else 'Ligen'

    # ── Wiederholungen ────────────────────────────────────────────────────────
    st.subheader('Wiederholungen pro Liga')
    cs1, cs2 = st.columns([3, 2])
    with cs1:
        sol['seeds'] = int(st.slider(
            'Anzahl Versuche', 1, 6, sol.get('seeds', 2), key='sol_seeds'))
        st.caption(
            f'Der Optimierer probiert für jede Liga den Schritt 1 **{sol["seeds"]}× '
            f'mit leicht unterschiedlichen Ausgangspunkten** aus und wählt am Ende das '
            f'beste Ergebnis. '
            f'**1 Versuch** = schnell, aber das Ergebnis kann zufällig schlechter ausfallen. '
            f'**2–3 Versuche** = gute Balance aus Zeit und Qualität (empfohlen). '
            f'**4–6 Versuche** = bestmögliche Qualität, braucht aber deutlich länger.'
        )
    with cs2:
        st.info(
            f'⏱ Schritt 1 läuft **{sol["seeds"]}× pro Liga** hintereinander.\n\n'
            f'Zeitbedarf Schritt 1: ~{sol["seeds"] * sol.get("p1", 900) // 60} min'
        )

    st.divider()

    # ── Schritt 1 ─────────────────────────────────────────────────────────────
    st.subheader('Schritt 1 – Spielplan für jede Liga erstellen')
    cp1, cp1b = st.columns([3, 2])
    with cp1:
        sol['p1'] = int(st.number_input(
            'Rechenzeit pro Liga und Versuch (Sekunden)',
            60, 7200, sol.get('p1', 900), 60, key='sol_p1'))
        st.caption(
            'Im ersten Schritt erstellt der Optimierer für jede Liga einen vollständigen '
            'Spielplan: Heimspiele und Auswärtsspiele werden so verteilt, dass möglichst '
            'oft gewechselt wird und die Fahrtwege insgesamt kurz bleiben. '
            'Alle Ligen werden **gleichzeitig** berechnet – es spielt also keine Rolle, '
            'ob du 2 oder 5 Ligen hast, die Wartezeit bleibt ähnlich. '
            '**15 Minuten (900 s)** reichen für die meisten Ligen. '
            'Wenn eine Liga sehr viele Teams, viele Doppelspieltage oder enge '
            'Terminvorgaben hat, helfen **20–30 Minuten**.'
        )
    with cp1b:
        st.info(
            f'⏱ Zeitbedarf Schritt 1:\n\n'
            f'{sol["seeds"]} Versuche × {sol["p1"]//60} min = '
            f'**~{sol["seeds"] * sol["p1"] // 60} min** pro Liga\n\n'
            f'_(alle {_ligen_wort} laufen gleichzeitig)_'
        )

    st.divider()

    # ── Schritt 2 ─────────────────────────────────────────────────────────────
    st.subheader('Schritt 2 – Ligen aufeinander abstimmen')
    cp2, cp2b = st.columns([3, 2])
    with cp2:
        _p2_opts = ['Standard (90 min)', 'Intensiv (3 h)', 'Nachtlauf (8 h)']
        _p2_secs = {'Standard (90 min)': 5400, 'Intensiv (3 h)': 10800, 'Nachtlauf (8 h)': 28800}
        _p2_cur  = sol.get('p2', 5400)
        _p2_idx  = 2 if _p2_cur >= 28800 else (1 if _p2_cur >= 10800 else 0)
        p2_mode  = st.radio('Rechenzeit', _p2_opts, index=_p2_idx, key='sol_p2_mode')
        sol['p2'] = _p2_secs[p2_mode]
        sol['nm'] = (p2_mode == 'Nachtlauf (8 h)')
        st.caption(
            'Im zweiten Schritt schaut der Optimierer auf alle Ligen gemeinsam und '
            'verteilt die Spieltage auf die Kalenderwochen. '
            'Hauptziel: Vereine, die in mehreren Ligen Teams haben, sollen ihre Heimspiele '
            'möglichst in denselben Wochen haben – damit Hallen und Infrastruktur '
            'gebündelt genutzt werden können. '
            '**Standard** ist ausreichend, wenn du nur wenige Ligen und wenige solcher '
            'Vereine hast. '
            'Mit **Intensiv oder Nachtlauf** bekommt der Optimierer mehr Zeit und kann '
            'die Abstimmung deutlich verbessern – sinnvoll ab 3 Ligen oder '
            'vielen Vereinen mit mehreren FBL-Teams.'
        )
    with cp2b:
        st.info(
            f'⏱ Zeitbedarf Schritt 2:\n\n'
            f'**~{sol["p2"]//60} min**\n\n'
            f'_(alle {_ligen_wort} werden hier gemeinsam berechnet)_'
        )

    st.divider()

    # ── Schritt 3 ─────────────────────────────────────────────────────────────
    st.subheader('Schritt 3 – Feinabstimmung der Heimspiele (optional)')
    cp3, cp3b = st.columns([3, 2])
    with cp3:
        sol['sa'] = int(st.number_input(
            'Rechenzeit pro Liga (Sekunden, 0 = Schritt 3 überspringen)',
            0, 600, sol.get('sa', 120), 30, key='sol_sa'))
        st.caption(
            'Im dritten Schritt wird noch einmal gezielt geprüft, ob man bei einzelnen '
            'Begegnungen Heim- und Auswärtsspiel tauschen sollte, um insgesamt weniger '
            'Kilometer zu fahren. **Der Terminplan selbst bleibt dabei unverändert** – '
            'es geht nur darum, wer an welchem Spieltag zu Hause spielt. '
            'Das bringt erfahrungsgemäß **3–8 % weniger Fahrtwege**. '
            '**2 Minuten (120 s) pro Liga** reichen fast immer aus. '
            'Auf 0 setzen, um diesen Schritt zu überspringen.'
        )
    with cp3b:
        sa_total = sol['sa'] * n_lig
        sa_min, sa_sec = sa_total // 60, sa_total % 60
        sa_str = f'{sa_min} min {sa_sec} s' if sa_sec else f'{sa_min} min'
        st.info(
            f'⏱ Zeitbedarf Schritt 3:\n\n'
            f'{n_lig} {_ligen_wort} × {sol["sa"]} s = '
            f'**~{sa_str}**\n\n'
            f'_({_ligen_wort} werden nacheinander bearbeitet)_'
        )

    st.divider()

    # ── Gesamtlaufzeit ────────────────────────────────────────────────────────
    st.subheader('Geschätzte Gesamtlaufzeit')
    p1_total  = sol['seeds'] * sol['p1']
    p2_total  = sol['p2']
    p3_total  = sol['sa'] * n_lig
    est_total = p1_total + p2_total + p3_total
    h, m = est_total // 3600, (est_total % 3600) // 60

    cg1, cg2, cg3, cg4 = st.columns(4)
    cg1.metric('Gesamt (ca.)', f'{h}h {m:02d}min')
    cg2.metric('Schritt 1', f'~{p1_total//60} min',
               help=f'Pro Liga, alle {_ligen_wort} laufen gleichzeitig')
    cg3.metric('Schritt 2', f'~{p2_total//60} min',
               help=f'Alle {_ligen_wort} gemeinsam')
    cg4.metric('Schritt 3', f'~{p3_total//60} min',
               help=f'Alle {_ligen_wort} nacheinander')
    st.caption(
        'Der Wert ist ein Richtwert. Je nach Computer und Komplexität der Spielpläne '
        'kann die tatsächliche Dauer etwas abweichen.'
    )

    S.solver = sol

    go_back, go_fwd = _nav(fwd_label='Zur Optimierung →')
    if go_back:
        S.step = 6; st.rerun()
    if go_fwd:
        S.step = 8; st.rerun()


# ═══════════════════════════════════════════════════════════════════════════════
# Sitzungs-Serialisierung (Ergebnisse speichern / laden)
# ═══════════════════════════════════════════════════════════════════════════════

def _session_to_json() -> bytes:
    """Serialisiert Konfiguration + Ergebnisse als JSON-Bytes (UTF-8)."""
    dist_mat_data = {}
    for lid, mat in (S.dist_matrices or {}).items():
        if mat is not None and isinstance(mat, np.ndarray):
            dist_mat_data[lid] = mat.tolist()

    dst_data = {
        lid: [list(b) for b in blocks]
        for lid, blocks in (S.dst_per_liga or {}).items()
    }

    kw_compat_data = {
        str(kw): kw_data for kw, kw_data in (S.kw_compat or {}).items()
    }

    routing_data = {
        lid: [bool(v[0]), int(v[1])]
        for lid, v in (S.routing or {}).items()
    }

    results_data: dict = {}
    if S.results:
        for lid, res in S.results.items():
            if res is None:
                continue
            results_data[lid] = {
                'schedule': {
                    str(d): [[ht, at] for ht, at in games]
                    for d, games in res.schedule.items()
                },
                'game_times': {
                    str(d): times for d, times in (res.game_times or {}).items()
                },
                'groups': {
                    str(d): [list(g) for g in grps]
                    for d, grps in (res.groups or {}).items()
                },
                'hosts': {
                    str(d): host for d, host in (res.hosts or {}).items()
                },
            }

    data = {
        'version': '1.0',
        'config': {
            'league_order': S.league_order,
            'leagues':      S.leagues,
            'dst_per_liga': dst_data,
            'blocked':      S.blocked,
            'forced_home':  S.forced_home,
            'pinned':       S.pinned,
            'routing':      routing_data,
            'clubs':        S.clubs,
            'kw_compat':    kw_compat_data,
            'cal_table':    {lid: rows for lid, rows in (S.cal_table or {}).items()},
            'w_cohome':     float(S.w_cohome),
            'dist_matrices': dist_mat_data,
            'weights':      S.weights,
            'solver':       S.solver,
        },
        'results': results_data,
    }
    return json.dumps(data, ensure_ascii=False, indent=2).encode('utf-8')


def _session_from_json(raw: bytes) -> str:
    """Lädt Konfiguration + Ergebnisse aus JSON. Gibt '' bei Erfolg zurück, sonst Fehlertext."""
    from ortools.sat.python import cp_model as _cp

    try:
        data = json.loads(raw.decode('utf-8'))
    except Exception as exc:
        return f'JSON-Lesefehler: {exc}'

    cfg_data = data.get('config', {})
    results_data = data.get('results', {})

    # ── Konfiguration wiederherstellen ────────────────────────────────────────
    if 'league_order' in cfg_data:
        S.league_order = cfg_data['league_order']
    if 'leagues' in cfg_data:
        S.leagues = cfg_data['leagues']
        for _ld in S.leagues.values():
            if 'teams' in _ld:
                _ld['teams'] = [tuple(e) for e in _ld['teams']]
    if 'dst_per_liga' in cfg_data:
        S.dst_per_liga = {
            lid: [tuple(b) for b in blocks]
            for lid, blocks in cfg_data['dst_per_liga'].items()
        }
    if 'blocked' in cfg_data:
        S.blocked = cfg_data['blocked']
    if 'forced_home' in cfg_data:
        S.forced_home = cfg_data['forced_home']
    if 'pinned' in cfg_data:
        S.pinned = cfg_data['pinned']
    if 'routing' in cfg_data:
        S.routing = {
            lid: (bool(v[0]), int(v[1]))
            for lid, v in cfg_data['routing'].items()
        }
    if 'clubs' in cfg_data:
        S.clubs = cfg_data['clubs']
    if 'kw_compat' in cfg_data:
        S.kw_compat = {int(kw): kw_data for kw, kw_data in cfg_data['kw_compat'].items()}
    if 'cal_table' in cfg_data:
        S.cal_table = cfg_data['cal_table']
    else:
        # Ältere Sitzung ohne cal_table: aus kw_compat migrieren
        _migrate_cal_table_from_kw_compat()
    if 'w_cohome' in cfg_data:
        S.w_cohome = float(cfg_data['w_cohome'])
    if 'weights' in cfg_data:
        S.weights = cfg_data['weights']
    if 'solver' in cfg_data:
        S.solver = cfg_data['solver']
    if 'dist_matrices' in cfg_data:
        S.dist_matrices = {
            lid: np.array(mat) for lid, mat in cfg_data['dist_matrices'].items()
        }

    if not results_data:
        return ''  # Nur Konfiguration geladen

    # ── Spielpläne wiederherstellen ───────────────────────────────────────────
    try:
        cfgs = _build_league_configs()
    except Exception as exc:
        return f'Konfiguration unvollständig – Spielpläne konnten nicht geladen werden: {exc}'

    from spielplan_multi.excel_output import (
        build_league_excel, build_cohome_summary, build_hall_schedule,
    )

    S.results    = {}
    S.excel_bytes = {}

    for lid, res_data in results_data.items():
        if lid not in cfgs:
            continue
        cfg = cfgs[lid]

        schedule = {
            int(d): [tuple(g) for g in games]
            for d, games in res_data.get('schedule', {}).items()
        }
        game_times = {
            int(d): times for d, times in res_data.get('game_times', {}).items()
        }
        groups = {
            int(d): [list(g) for g in grps]
            for d, grps in res_data.get('groups', {}).items()
        }
        hosts = {
            int(d): host for d, host in res_data.get('hosts', {}).items()
        }

        result = LeagueResult(
            league_id=lid,
            status=_cp.FEASIBLE,
            objective=0.0,
            schedule=schedule,
            sw_counts=[],
            sw_rates=[],
            travels=[],
            mins=0,
            secs=0,
            home_vals={},
            h_vals={},
            x_vals={},
            cfg=cfg,
            groups=groups,
            hosts=hosts,
            game_times=game_times,
        )
        try:
            travels, sw_counts, sw_rates = _recompute_result_stats_fn(result, cfg)
            result.travels   = travels
            result.sw_counts = sw_counts
            result.sw_rates  = sw_rates
        except Exception:
            n_teams = cfg.n_teams
            result.travels   = [0] * n_teams
            result.sw_counts = [0] * n_teams
            result.sw_rates  = [0.0] * n_teams
        S.results[lid] = result

        try:
            wb  = build_league_excel(result)
            buf = io.BytesIO()
            wb.save(buf)
            S.excel_bytes[lid] = buf.getvalue()
        except Exception:
            pass

    if S.clubs and S.results:
        try:
            wb_ch  = build_cohome_summary(S.results, S.clubs, S.kw_compat)
            buf_ch = io.BytesIO()
            wb_ch.save(buf_ch)
            S.cohome_bytes = buf_ch.getvalue()
        except Exception:
            pass

    if S.results:
        try:
            wb_hall  = build_hall_schedule(S.results)
            buf_hall = io.BytesIO()
            wb_hall.save(buf_hall)
            S.hall_bytes = buf_hall.getvalue()
        except Exception:
            pass

    S.opt_done        = True
    S._wizard_started = True
    return ''


# ═══════════════════════════════════════════════════════════════════════════════
# SCHRITT 8 – Optimierung & Ergebnisse
# ═══════════════════════════════════════════════════════════════════════════════

class _QueueWriter:
    """Leitet sys.stdout nur für den Solver-Thread in eine Queue um; andere Threads schreiben in das Original."""
    def __init__(self, q: queue.Queue, original, thread_id: int):
        import threading as _threading
        self._q = q
        self._original = original
        self._thread_id = thread_id
        self._buf = ''
        self._lock = _threading.Lock()

    def write(self, text: str):
        import threading as _threading
        if _threading.get_ident() == self._thread_id:
            with self._lock:
                self._buf += text
                while '\n' in self._buf:
                    line, self._buf = self._buf.split('\n', 1)
                    self._q.put(line)
        else:
            self._original.write(text)

    def flush(self):
        import threading as _threading
        if _threading.get_ident() != self._thread_id:
            self._original.flush()


def _build_league_configs() -> Dict[str, LeagueConfig]:
    """Baut LeagueConfig-Objekte aus dem GUI-State."""
    # cal_table → spieltage_per_liga (inkl. Datum wenn vorhanden)
    spieltage: Dict[str, Dict] = {}
    for lid in S.league_order:
        rows = S.cal_table.get(lid, [])
        if rows:
            for row in rows:
                kw = row.get('kw')
                if kw is None:
                    continue
                date = row.get('date', '').strip()
                spieltage.setdefault(lid, {})[row['spieltag']] = {
                    'kw': int(kw),
                    'week_start': date,
                    'week_end': date,
                }
        else:
            # Fallback für Ligen ohne cal_table-Eintrag (Altdaten)
            for kw, kw_data in S.kw_compat.items():
                if lid in kw_data:
                    for st_nr in kw_data[lid]:
                        spieltage.setdefault(lid, {}).setdefault(
                            st_nr, {'kw': kw, 'week_start': '', 'week_end': ''})

    cfgs = {}
    for lid in S.league_order:
        ld       = S.leagues[lid]
        n_rounds, gpd = _get_n_rounds_gpd(ld)
        teams    = [t for t, _ in ld['teams']]
        locs     = [loc for _, loc in ld['teams']]
        dst      = S.dst_per_liga.get(lid, [])
        n_md     = _calc_n_matchdays(ld)
        days     = list(range(1, n_md + 1))
        apply_r, pct = S.routing.get(lid, (False, 25))
        _w_defaults = {k: (0.0 if k == 'dst_eff' else 5.0) for k in WEIGHT_SCALES}
        raw      = S.weights.get(lid, _w_defaults)
        scaled   = {k: v * WEIGHT_SCALES[k] for k, v in raw.items() if k in WEIGHT_SCALES}
        # Spielfrei-Modus: active_per_day < n_teams
        n = len(teams)
        n_active = 0
        if ld.get('fmt') == 'Turniertag':
            apd = int(ld.get('active_per_day', n))
            if 0 < apd < n:
                n_active = apd

        # Stufe 2: k_group < n_teams → Gruppenformation; k_group==0 oder ==n → Stufe 1
        if n_active > 0:
            k_group = n_active  # 1 Gruppe mit n_active Teams, Rest hat Spielfrei
        else:
            k_group = int(ld.get('k_group', 0)) if ld.get('fmt') == 'Turniertag' else 0
            if k_group >= n:
                k_group = 0

        cfgs[lid] = LeagueConfig(
            league_id=lid,
            name=ld['name'],
            teams=teams,
            locations=locs,
            dist=S.dist_matrices[lid],
            dst_blocks=dst,
            weekends=build_weekends(days, dst),
            apply_routing=apply_r,
            f_num=100 + pct,
            f_den=100,
            w_scaled=scaled,
            raw_weights=raw,
            pinned=S.pinned.get(lid, []),
            blocked=S.blocked.get(lid, {}),
            calendar=spieltage.get(lid, {}),
            forced_home=S.forced_home.get(lid, {}),
            hier_weight=float(ld.get('hw', 1.0)),
            games_per_team_per_day=gpd,
            n_rounds=n_rounds,
            n_teams_per_group=k_group,
            n_active_per_day=n_active,
            tt_settings=ld.get('tt_settings', {}),
        )
    return cfgs


def _solver_thread(cfgs, clubs, kw_compat, w_cohome, solver_cfg,
                   result_holder: dict, log_q: queue.Queue):
    import threading as _threading
    old_out = sys.stdout
    sys.stdout = _QueueWriter(log_q, old_out, _threading.get_ident())
    try:
        result_holder['results'] = solve_all(
            cfgs=cfgs,
            clubs=clubs,
            kw_compat=kw_compat,
            w_cohome=w_cohome,
            phase1_time=solver_cfg['p1'],
            phase2_time=solver_cfg['p2'],
            night_mode=solver_cfg['nm'],
            n_seeds=solver_cfg['seeds'],
            sa_time=solver_cfg['sa'],
        )
        # Ergebnis sofort auf Disk sichern – überlebt einen Streamlit-Session-Verlust
        _pkl = _HERE / '.cache' / 'last_result.pkl'
        try:
            import pickle as _pickle
            _pkl.parent.mkdir(exist_ok=True)
            _pkl.write_bytes(_pickle.dumps({
                'results':   result_holder['results'],
                'clubs':     clubs,
                'kw_compat': kw_compat,
            }))
        except Exception:
            pass
    except Exception as exc:
        import traceback
        log_q.put(f'[FEHLER] {exc}')
        log_q.put(traceback.format_exc())
    finally:
        sys.stdout = old_out
        log_q.put('__DONE__')


def _validate_constraints() -> List[dict]:
    """Prüft die Konfiguration auf häufige Probleme vor dem Solver-Start."""
    return _validate_cfg(
        league_order  = S.league_order,
        leagues       = S.leagues,
        dst_per_liga  = S.dst_per_liga,
        blocked       = S.blocked,
        pinned        = S.pinned,
        dist_matrices = S.dist_matrices,
        kw_compat     = S.kw_compat,
        clubs         = S.clubs,
        calc_n_matchdays = _calc_n_matchdays,
        get_n_rounds_gpd = _get_n_rounds_gpd,
        routing       = S.routing,
        forced_home   = S.forced_home,
    )


def _step8():
    st.header('9. Optimierung & Ergebnisse')

    # ── Wiederherstellung nach Session-Verlust ────────────────────────────────
    _pkl = _HERE / '.cache' / 'last_result.pkl'
    if _pkl.exists() and not S.opt_done and S.results is None and not S.opt_running:
        import pickle as _pickle
        import datetime as _dt
        _mtime   = _dt.datetime.fromtimestamp(_pkl.stat().st_mtime)
        _age     = _dt.datetime.now() - _mtime
        _h, _m   = int(_age.total_seconds() // 3600), int((_age.total_seconds() % 3600) // 60)
        _age_str = f'{_h}h {_m}min' if _h else f'{_m}min'
        try:
            _preview = _pickle.loads(_pkl.read_bytes())
            _ligen   = [r.cfg.name for r in _preview.get('results', {}).values()
                        if r is not None and r.cfg]
        except Exception:
            _ligen = []
        _ligen_str = ', '.join(_ligen) if _ligen else '(unbekannt)'
        st.warning(
            f'**Nicht gespeicherte Ergebnisse gefunden** – die letzte Optimierung wurde '
            f'vor {_age_str} abgeschlossen, aber die Streamlit-Session war zu diesem '
            f'Zeitpunkt nicht mehr aktiv. Betroffene Ligen: {_ligen_str}'
        )
        _r1, _r2 = st.columns(2)
        with _r1:
            if st.button('Ergebnisse wiederherstellen', key='recover_pkl', type='primary'):
                try:
                    _data    = _pickle.loads(_pkl.read_bytes())
                    S.results   = _data.get('results', {})
                    S.clubs     = _data.get('clubs',   S.clubs)
                    S.kw_compat = _data.get('kw_compat', S.kw_compat)
                    S.excel_bytes  = {}
                    S.cohome_bytes = None
                    S.hall_bytes   = None
                    from spielplan_multi.excel_output import (
                        build_league_excel, build_cohome_summary, build_hall_schedule)
                    for _lid, _res in S.results.items():
                        if _res is None:
                            continue
                        try:
                            _wb  = build_league_excel(_res)
                            _buf = io.BytesIO()
                            _wb.save(_buf)
                            S.excel_bytes[_lid] = _buf.getvalue()
                        except Exception:
                            pass
                    if S.clubs and S.results:
                        try:
                            _wb_ch  = build_cohome_summary(S.results, S.clubs, S.kw_compat)
                            _buf_ch = io.BytesIO()
                            _wb_ch.save(_buf_ch)
                            S.cohome_bytes = _buf_ch.getvalue()
                        except Exception:
                            pass
                    try:
                        _wb_h  = build_hall_schedule(S.results)
                        _buf_h = io.BytesIO()
                        _wb_h.save(_buf_h)
                        S.hall_bytes = _buf_h.getvalue()
                    except Exception:
                        pass
                    S.opt_done = True
                    _pkl.unlink(missing_ok=True)
                    st.rerun()
                except Exception as _e:
                    st.error(f'Wiederherstellung fehlgeschlagen: {_e}')
        with _r2:
            if st.button('Verwerfen', key='discard_pkl'):
                _pkl.unlink(missing_ok=True)
                st.rerun()
        st.divider()

    # Ergebnisse anzeigen wenn fertig
    if S.opt_done and S.results is not None:
        _show_results()
        st.divider()
        _rcol_a, _rcol_b, _rcol_c = st.columns(3)
        with _rcol_a:
            if st.button('🔄  Neu berechnen', key='reopt', width='stretch',
                         help='Konfiguration behalten und Optimierung erneut starten – '
                              'z. B. nach Änderung einzelner Einstellungen.'):
                (_HERE / '.cache' / 'last_result.pkl').unlink(missing_ok=True)
                S.opt_done      = False
                S.opt_running   = False
                S.results       = None
                S.opt_queue     = None
                S.opt_process   = None
                S.opt_warnings  = []
                S.opt_best      = {}
                S.hall_bytes    = None
                S.cohome_bytes  = None
                S.excel_bytes   = {}
                st.rerun()
        with _rcol_b:
            if st.button('↺  Neuen Spielplan erstellen', key='restart', width='stretch'):
                (_HERE / '.cache' / 'last_result.pkl').unlink(missing_ok=True)
                import copy as _copy
                for k, v in _DEFAULTS.items():
                    st.session_state[k] = _copy.deepcopy(v)
                st.rerun()
        with _rcol_c:
            st.download_button(
                '💾  Sitzung speichern',
                data=_session_to_json(),
                file_name='spielplan_sitzung.json',
                mime='application/json',
                width='stretch',
                key='dl_session',
                help='Speichert Konfiguration und Spielpläne als JSON-Datei. '
                     'Kann später über "Frühere Ergebnisse laden" wieder geöffnet werden.',
            )
        return

    # Konfigurationsübersicht (immer sichtbar wenn nicht fertig)
    st.subheader('Zusammenfassung der Konfiguration')

    # ─── Ligen (max. 3 pro Zeile) ─────────────────────────────────────────────
    _lids = S.league_order
    _ncols = min(len(_lids), 3)
    for _row_start in range(0, len(_lids), _ncols):
        _row_lids = _lids[_row_start:_row_start + _ncols]
        _row_cols = st.columns(_ncols)
        for _ci, _lid in enumerate(_row_lids):
            _ld  = S.leagues[_lid]
            _nmd = _calc_n_matchdays(_ld)
            _nt  = len(_ld.get('teams', []))
            _dst = S.dst_per_liga.get(_lid, [])
            _pin = S.pinned.get(_lid, [])
            _blk = S.blocked.get(_lid, {})
            _n_blk = sum(len(v) for v in _blk.values())
            _rt_on, _rt_pct = S.routing.get(_lid, (False, 25))
            with _row_cols[_ci]:
                # **bold** verhindert, dass Markdown "1. Name" als Liste interpretiert
                st.markdown(f'**{_ld["name"]}** &nbsp;`{_lid}`',
                            unsafe_allow_html=True)
                st.write(f'Format: {_ld.get("fmt", "–")}')
                st.write(f'Teams: {_nt} · Spieltage: {_nmd}')
                if _ld.get('fmt') == 'Turniertag':
                    _apd = int(_ld.get('active_per_day', _nt) or _nt)
                    _gpd = int(_ld.get('gpd', 2) or 2)
                    _tt  = _ld.get('tt_settings') or {}
                    if _apd < _nt:
                        st.write(f'Teams/Tag: {_apd} von {_nt} · {_nt - _apd} Spielfrei')
                    else:
                        _kg = int(_ld.get('k_group', 0) or 0)
                        if 0 < _kg < _nt:
                            import math as _math2
                            st.write(f'Gruppen: {_math2.ceil(_nt / _kg)} × {_kg} Teams')
                    st.caption(f'Spiele/Team/Tag: {_gpd}')
                    if _tt:
                        _tt_parts = []
                        _tt_slots = _tt.get('host_slots', [])
                        if not _tt_slots and _tt.get('host_position'):
                            _tt_slots = ['2', 'N-1']
                        if _tt_slots:
                            _tt_parts.append(f'Ausrichterslots: {list(_tt_slots)}')
                        _tt_parts.append(f'Pause {_tt.get("min_gap", 0)}–{_tt.get("max_gap", 3)}')
                        if _tt.get('host_mode') == 'per_day' and _tt.get('host_per_day'):
                            _tt_parts.append(f'{len(_tt["host_per_day"])} ST fix')
                        elif _tt.get('host_counts') and any(v > 0 for v in _tt['host_counts'].values()):
                            _tt_parts.append(f'{sum(_tt["host_counts"].values())} Ausrichter-ST')
                        st.caption('Spielreihenf.: ' + ' · '.join(_tt_parts))
                if _dst:
                    _dst_str = ', '.join(f'ST {a}+{b}' for a, b in _dst)
                    st.write(f'DST-Blöcke ({len(_dst)}): {_dst_str}')
                else:
                    st.write('Keine DST-Blöcke')
                if _pin:
                    st.write(f'Pflichtspiele: {len(_pin)}')
                if _n_blk:
                    st.write(f'Sperrtage: {_n_blk} '
                             f'({len(_blk)} {"Team" if len(_blk)==1 else "Teams"})')
                if _rt_on:
                    st.write(f'DST-Routing: max. +{_rt_pct}% Umweg')
                # Gewichte
                _w = S.weights.get(_lid) or S.weights.get('__common__', {})
                if _w:
                    _W_DISP_DEF = {'dst_eff': 0.0}
                    _wlines = [f'{lbl}: {_w.get(k, _W_DISP_DEF.get(k, 5.0)):.0f}'
                               for k, (lbl, _) in WEIGHT_LABELS.items()]
                    st.caption('Gewichte – ' + ' · '.join(_wlines))
        if _row_start + _ncols < len(_lids):
            st.write('')

    st.divider()

    # ─── Globale Einstellungen ─────────────────────────────────────────────────
    _ga, _gb = st.columns(2)

    with _ga:
        st.markdown('**Co-Home – Vereine mit mehreren FBL-Teams**')
        _n_clubs = len(S.clubs)
        if _n_clubs:
            for _cn, _cm in S.clubs.items():
                _team_str = ', '.join(
                    f'{S.leagues.get(l, {}).get("name", l)}: {t}'
                    for l, t in _cm.items())
                st.write(f'· {_cn} ({_team_str})')
            st.caption(f'Co-Home-Bonus: {S.w_cohome:.1f} / 10')
        else:
            st.caption('Keine Vereine mit mehreren FBL-Teams konfiguriert.')

    with _gb:
        sol = S.solver
        _ligen_wort = 'Liga' if len(_lids) == 1 else 'Ligen'
        st.markdown('**Solver**')
        st.write(f'Versuche pro Liga: {sol["seeds"]}')
        st.write(f'Schritt 1: {sol["p1"]//60} min pro Versuch '
                 f'(alle {_ligen_wort} gleichzeitig)')
        _p2_label = {5400: 'Standard', 10800: 'Intensiv', 28800: 'Nachtlauf'}.get(
            sol["p2"], f'{sol["p2"]//60} min')
        st.write(f'Schritt 2: {_p2_label} ({sol["p2"]//60} min, '
                 f'alle {_ligen_wort} gemeinsam)')
        if sol['sa']:
            st.write(f'Schritt 3: {sol["sa"]} s pro Liga '
                     f'({len(_lids)} {_ligen_wort} × {sol["sa"]} s = '
                     f'~{sol["sa"] * len(_lids) // 60} min)')
        else:
            st.write('Schritt 3: deaktiviert')
        _p1_tot = sol["seeds"] * sol["p1"]
        _p3_tot = sol["sa"] * len(_lids)
        _est    = _p1_tot + sol["p2"] + _p3_tot
        _h, _m  = _est // 3600, (_est % 3600) // 60
        st.caption(f'Geschätzte Gesamtlaufzeit: ~{_h}h {_m:02d}min')

    st.divider()

    # ── Unterer Bereich: Fortschritt ODER Startbereit – st.empty() sorgt für atomaren Tausch ──
    _lower = st.empty()

    if S.opt_running:
        # Queue zuerst leeren (kein Rendern) – Daten für Container bereitstellen
        _q = S.opt_queue
        _done_flag = False
        if _q is None:
            S.opt_running = False
            st.rerun()
            return
        try:
            while True:
                _line = _q.get_nowait()
                if _line == '__DONE__':
                    _done_flag = True
                    break
                if isinstance(_line, tuple) and len(_line) == 2 and _line[0] == '__RESULTS__':
                    S.opt_result_holder = {'results': _line[1]}
                    continue
                S.opt_log.append(_line)
                if _line.startswith('[BEST]'):
                    _parts = _line.split()
                    if len(_parts) >= 4:
                        _lid_b = _parts[1]
                        _obj_b = next((p.split('=')[1] for p in _parts
                                       if p.startswith('obj=')), None)
                        _t_b   = next((p.split('=')[1] for p in _parts
                                       if p.startswith('t=')), '')
                        _cnt_b = next((p.strip('(#)') for p in _parts
                                       if p.startswith('(#')), '')
                        if _obj_b:
                            S.opt_best[_lid_b] = {
                                'obj': float(_obj_b), 'elapsed': _t_b, 'count': _cnt_b
                            }
        except (queue.Empty, EOFError, OSError):
            pass

        # Zeitberechnung & Phase
        _elapsed = time.time() - (S.opt_start_time or time.time())
        _p1_tot  = S.solver['seeds'] * S.solver['p1']
        _p2_tot  = S.solver['p2']
        _p3_tot  = S.solver['sa'] * len(S.league_order)
        _total   = max(_p1_tot + _p2_tot + _p3_tot, 1)
        _pct     = min(_elapsed / _total, 0.99)
        if any('PHASE 3' in l for l in S.opt_log):
            _phase_name = 'Phase 3 – SA-Nachbearbeitung'
        elif any(l.startswith('[BEST] P2') for l in S.opt_log):
            _phase_name = 'Phase 2 – Kalender-Koordination'
        else:
            _phase_name = 'Phase 1 – Heimrecht-Optimierung'
        _h_e = int(_elapsed // 3600)
        _m_e = int((_elapsed % 3600) // 60)
        _s_e = int(_elapsed % 60)
        _elapsed_str = (f'{_h_e}h {_m_e:02d}min {_s_e:02d}s' if _h_e
                        else f'{_m_e}min {_s_e:02d}s')

        # Container ersetzt den alten Constraint-Check + Button atomisch
        with _lower.container():
            st.markdown(f'### ⏳ {_phase_name}')
            st.progress(_pct)
            st.markdown(
                f'<p style="font-size:1.25rem; margin:-0.3rem 0 0.6rem 0;">'
                f'<b>{int(_pct * 100)}&thinsp;%</b>&ensp;·&ensp;{_elapsed_str} vergangen'
                f'</p>',
                unsafe_allow_html=True,
            )
            if S.opt_best:
                _m_cols = st.columns(max(1, min(4, len(S.opt_best))))
                for _ci, (_lid_m, _bst) in enumerate(S.opt_best.items()):
                    _name_m = (S.leagues.get(_lid_m, {}).get('name', _lid_m)
                               if _lid_m != 'P2' else 'Phase 2 gesamt')
                    with _m_cols[_ci % len(_m_cols)]:
                        st.metric(
                            label=f'Beste Lösung: {_name_m}',
                            value=f'{_bst["obj"]:,.0f}',
                            help=f't={_bst["elapsed"]}  |  #{_bst["count"]} Lösung(en)',
                        )
            st.code('\n'.join(S.opt_log[-80:]), language=None)
            st.divider()
            if st.button('⏹  OPTIMIERUNG ABBRECHEN', type='primary', width='stretch',
                         help='Bricht die laufende Berechnung ab. '
                              'Einstellungen bleiben erhalten – die Optimierung kann '
                              'danach angepasst und neu gestartet werden.'):
                _proc = S.opt_process
                if _proc is not None:
                    _proc.terminate()
                    _proc.join(timeout=3)
                    if _proc.is_alive():
                        _proc.kill()
                        _proc.join(timeout=1)
                S.opt_running       = False
                S.opt_done          = False
                S.opt_queue         = None
                S.opt_process       = None
                S.opt_result_holder = None
                S.opt_log           = []
                S.opt_best          = {}
                S.opt_warnings      = []
                S.opt_start_time    = None
                st.rerun()

        if _done_flag:
            S.opt_running = False
            S.opt_done    = True
            S.results     = S.opt_result_holder.get('results', {})
            S.opt_warnings = []
            for _wl in S.opt_log:
                if '[!!]' in _wl:
                    S.opt_warnings.append(_wl.replace('[!!]', '').strip())
                elif '[FEHLER]' in _wl:
                    S.opt_warnings.append(f'Fehler: {_wl.replace("[FEHLER]", "").strip()}')
            try:
                from spielplan_multi.excel_output import (
                    build_league_excel, build_cohome_summary, build_hall_schedule,
                )
                for lid, res in (S.results or {}).items():
                    if res is not None:
                        wb  = build_league_excel(res)
                        buf = io.BytesIO()
                        wb.save(buf)
                        S.excel_bytes[lid] = buf.getvalue()
                if S.clubs and S.results:
                    wb_ch  = build_cohome_summary(S.results, S.clubs, S.kw_compat)
                    buf_ch = io.BytesIO()
                    wb_ch.save(buf_ch)
                    S.cohome_bytes = buf_ch.getvalue()
                if S.results:
                    wb_hall  = build_hall_schedule(S.results)
                    buf_hall = io.BytesIO()
                    wb_hall.save(buf_hall)
                    S.hall_bytes = buf_hall.getvalue()
            except Exception as _exc_excel:
                import traceback as _tb_excel
                S.opt_warnings.append(f'Excel-Erzeugung fehlgeschlagen: {_exc_excel}')
                S.opt_log.append(f'[!!] Excel-Fehler: {_tb_excel.format_exc()}')
            st.rerun()
        else:
            time.sleep(2)
            st.rerun()
        return

    # ── Nicht laufend: Constraint-Prüfung + Startknopf im gleichen Placeholder ──
    with _lower.container():
        issues = _validate_constraints()
        _has_errors = any(i['level'] == 'error' for i in issues)
        if issues:
            with st.expander(
                f'{"❌" if _has_errors else "⚠️"}  Konfigurations-Prüfung '
                f'({sum(1 for i in issues if i["level"]=="error")} Fehler, '
                f'{sum(1 for i in issues if i["level"]=="warning")} Hinweise)',
                expanded=True,
            ):
                for issue in issues:
                    prefix = f'`{issue["lid"]}`  ' if issue.get('lid') else ''
                    if issue['level'] == 'error':
                        st.error(prefix + issue['msg'])
                    else:
                        st.warning(prefix + issue['msg'])
                if _has_errors:
                    st.info('Bitte die Fehler (rot) beheben, bevor die Optimierung gestartet wird.')
        else:
            st.success('✅ Konfiguration geprüft – keine Probleme gefunden.')

        st.markdown('### Bereit zur Optimierung')
        if st.button('🖩  OPTIMIERUNG STARTEN', type='primary',
                     width='stretch', disabled=_has_errors):
            try:
                cfgs = _build_league_configs()
            except Exception as exc:
                st.error(f'Konfigurationsfehler: {exc}')
                return

            from spielplan_multi._worker import run_solver as _run_solver
            log_q = multiprocessing.Queue()
            proc  = multiprocessing.Process(
                target=_run_solver,
                args=(cfgs, S.clubs, S.kw_compat,
                      S.w_cohome, S.solver, log_q, str(_HERE)),
                daemon=True,
            )
            proc.start()

            S.opt_queue         = log_q
            S.opt_process       = proc
            S.opt_result_holder = {}
            S.opt_running       = True
            S.opt_start_time    = time.time()
            S.opt_done          = False
            S.opt_log           = []
            S.opt_warnings      = []
            S.results           = None
            S.excel_bytes       = {}
            S.cohome_bytes      = None
            st.rerun()


def _recompute_result_stats(result, cfg) -> tuple:
    return _recompute_result_stats_fn(result, cfg)


def _swap_home_away(result, cfg, day: int, match_idx: int) -> None:
    _swap_home_away_fn(result, cfg, day, match_idx)


def _build_ics_bytes(result, season_year: int) -> bytes:
    return _build_ics_bytes_fn(result, season_year)


def _move_game(result, cfg, old_day: int, match_idx: int, new_day: int) -> str:
    return _move_game_fn(result, cfg, old_day, match_idx, new_day)


def _cancel_game(result, cfg, day: int, match_idx: int):
    return _cancel_game_fn(result, cfg, day, match_idx)


def _reschedule_game(result, cfg, day: int, home_team: str, away_team: str) -> str:
    return _reschedule_game_fn(result, cfg, day, home_team, away_team)


def _find_free_days(result, cfg, team_a: str, team_b: str) -> list:
    return _find_free_days_fn(result, cfg, team_a, team_b)


def _find_schedule_warnings(result, cfg) -> list:
    return _find_schedule_warnings_fn(result, cfg)


def _regen_league_excel(lid: str, res) -> None:
    from spielplan_multi.excel_output import build_league_excel as _ble
    _wb  = _ble(res)
    _buf = io.BytesIO()
    _wb.save(_buf)
    S.excel_bytes[lid] = _buf.getvalue()


def _diagnose_infeasible_league(lid: str) -> None:
    """Zeigt Diagnose-Hinweise wenn eine Liga keine Lösung hat."""
    ld     = S.leagues.get(lid, {})
    name   = ld.get('name', lid)
    teams  = [t for t, _ in ld.get('teams', [])]
    n      = len(teams)
    n_rounds, gpd = _get_n_rounds_gpd(ld)
    n_md   = _calc_n_matchdays(ld)
    total_games = n * (n - 1) // 2 * n_rounds

    # Status aus Log ableiten – zeilenweise prüfen um False Positives zu vermeiden
    _log_lines = S.opt_log or []
    _lid_pat   = r'(?<!\w)' + re.escape(lid) + r'(?!\w)'
    _lid_in    = lambda t: bool(re.search(_lid_pat, t))
    is_infeasible = any('INFEASIBLE'    in ln and _lid_in(ln) for ln in _log_lines)
    is_unknown    = any(
        ('UNKNOWN'      in ln and _lid_in(ln)) or
        ('Keine Loesung' in ln and _lid_in(ln))
        for ln in _log_lines
    )

    hints: list = []

    if is_infeasible:
        hints.append(('error',
            'Der Solver hat bewiesen, dass mit diesen Einstellungen kein gültiger '
            'Spielplan existiert (INFEASIBLE). Es liegt ein echter Widerspruch in den '
            'Constraints vor.'))
    elif is_unknown:
        hints.append(('warn',
            'Der Solver hat das Zeitlimit erreicht, ohne eine Lösung zu finden (UNKNOWN). '
            'Das bedeutet nicht zwingend, dass keine Lösung existiert.'))
        hints.append(('suggestion', 'Zeitlimit für Phase 1 erhöhen (Schritt 8, empfohlen: 30–60 min pro Liga).'))
        hints.append(('suggestion', 'Seeds erhöhen (3–5) um mehr Suchstrategien zu versuchen.'))

    # Pflichtspiele-Dichte
    pins = S.pinned.get(lid, [])
    if total_games > 0 and len(pins) > total_games * 0.3:
        pct = len(pins) * 100 // total_games
        hints.append(('warn',
            f'{len(pins)} von {total_games} Spielen ({pct}%) sind Pflichtspiele – '
            f'das schränkt den Solver stark ein.'))
        hints.append(('suggestion', 'Pflichtspiele auf die wirklich notwendigen reduzieren.'))

    # Sperrtage-Dichte
    blk = S.blocked.get(lid, {})
    for team, bdays in blk.items():
        cnt = sum(1 for d in bdays if 1 <= d <= n_md)
        if cnt >= n_md // 2:
            hints.append(('warn',
                f'Team «{team}» hat {cnt} von {n_md} Spieltagen gesperrt '
                f'({cnt * 100 // n_md}%). '
                'Heimspiele sind für dieses Team kaum planbar.'))
            hints.append(('suggestion',
                f'Sperrtage für «{team}» reduzieren oder Zeitlimit erhöhen.'))

    # Pflichtheim-Dichte
    frc = S.forced_home.get(lid, {})
    for team, fdays in frc.items():
        cnt = sum(1 for d in fdays if 1 <= d <= n_md)
        if cnt > n_md // 2:
            hints.append(('warn',
                f'Team «{team}» hat {cnt} Pflichtheim-Tage bei nur {n_md} Spieltagen – '
                'dies kann mit anderen Constraints kollidieren.'))
            hints.append(('suggestion',
                f'Pflichtheim-Tage für «{team}» reduzieren.'))

    # DST + Routing
    rt_on, rt_pct = S.routing.get(lid, (False, 25))
    dst = S.dst_per_liga.get(lid, [])
    if rt_on and dst and rt_pct < 100:
        hints.append(('warn',
            f'DST-Routing aktiv mit {rt_pct}% Toleranz. '
            'Bei niedrigen Werten kann die Routing-Einschränkung in Verbindung mit '
            'Sperrtagen unlösbar werden.'))
        hints.append(('suggestion',
            f'Routing-Toleranz erhöhen (≥ 100%, aktuell {rt_pct}%) oder deaktivieren.'))

    # DST + Sperrtage kombiniert
    for d1, d2 in dst:
        for team in teams:
            bdays_set = set(blk.get(team, []))
            if d1 in bdays_set and d2 in bdays_set:
                hints.append(('error',
                    f'Team «{team}» hat BEIDE DST-Tage (ST{d1}/ST{d2}) gesperrt. '
                    'Der DST-Constraint erzwingt gleiches Heimrecht an beiden Tagen, '
                    'Sperrtage erzwingen Auswärts → unlösbar.'))
                hints.append(('suggestion',
                    f'Mindestens einen der Sperrtage ST{d1}/ST{d2} für «{team}» entfernen.'))

    if not hints:
        hints.append(('warn',
            'Keine offensichtliche Ursache ermittelt. Der Solver hat keine Lösung gefunden.'))
        hints.append(('suggestion',
            'Zeitlimit und Seeds erhöhen. Falls das nicht hilft, Constraints schrittweise lockern.'))

    with st.expander(f'🔍  Diagnose: **{name}** – Keine Lösung', expanded=True):
        for kind, msg in hints:
            if kind == 'error':
                st.error(msg)
            elif kind == 'warn':
                st.warning(msg)
            elif kind == 'suggestion':
                st.info(f'💡 {msg}')


def _show_results():
    st.success('✅ Optimierung abgeschlossen!')

    for _w in (S.opt_warnings or []):
        st.warning(_w)

    if not S.results:
        st.error('Keine Ergebnisse vorhanden.')
        return

    # Diagnose für Ligen ohne Lösung
    for _flid, _fres in S.results.items():
        if _fres is None:
            _diagnose_infeasible_league(_flid)

    # Kennzahlen
    cols = st.columns(max(1, len(S.results)))
    for i, (lid, res) in enumerate(S.results.items()):
        if res is None:
            continue
        with cols[i]:
            name = res.cfg.name if res.cfg else lid
            _sw = res.sw_counts or []
            _sw_label = f'Wechsel: {min(_sw)}–{max(_sw)}' if _sw else 'Wechsel: –'
            st.metric(name,
                      f'{sum(res.travels):,} km gesamt',
                      _sw_label)

    # Warnungen zur Plan-Qualität
    _all_warnings: list = []
    for _wlid, _wres in S.results.items():
        if _wres is not None and _wres.cfg:
            for _w in _find_schedule_warnings(_wres, _wres.cfg):
                _w['liga'] = _wres.cfg.name
                _all_warnings.append(_w)
    if _all_warnings:
        _n_warn = sum(1 for _w in _all_warnings if _w['severity'] == 'warn')
        _lbl = f'⚠ Hinweise zur Plan-Qualität ({len(_all_warnings)}' + (
            f', davon {_n_warn} Warnungen' if _n_warn else '') + ')'
        with st.expander(_lbl, expanded=_n_warn > 0):
            st.caption(
                '**Warnung (orange):** Auffälligkeit, die den Spielbetrieb oder die Fairness '
                'beeinträchtigen kann – z. B. 4 oder mehr Auswärtsspiele hintereinander oder '
                'ein Team mit deutlich höherem Reiseaufwand als der Durchschnitt. '
                'Prüfen ob dies akzeptabel ist, oder Konfiguration anpassen und neu optimieren. '
                '**Hinweis (blau):** Weniger kritisch, aber zur Kenntnis nehmen.'
            )
            for _w in _all_warnings:
                _msg = f'**{_w["liga"]}** · {_w["team"]}: {_w["text"]}'
                if _w['severity'] == 'warn':
                    st.warning(_msg)
                else:
                    st.info(_msg)

    # Fairness-Überblick
    st.subheader('Fairness-Überblick')
    for lid, res in S.results.items():
        if res is None or not res.cfg:
            continue
        cfg_f  = res.cfg
        n_f    = cfg_f.n_teams
        avg_km = sum(res.travels) / n_f if n_f else 1
        rows_f = []
        for ti, t in enumerate(cfg_f.teams):
            km   = res.travels[ti]
            pct  = (km - avg_km) / avg_km * 100 if avg_km else 0
            sw   = res.sw_counts[ti]
            rate = res.sw_rates[ti]
            # home%
            total_games = sum(1 for d in cfg_f.days
                              for ht, at in res.schedule.get(d, [])
                              if ht == t or at == t)
            home_games  = sum(1 for d in cfg_f.days
                              for ht, _at in res.schedule.get(d, []) if ht == t)
            home_pct = round(home_games / total_games * 100) if total_games else 0
            rows_f.append({
                'Team':       t,
                'km':         km,
                'Abw.':       f'{pct:+.0f}%',
                'Heim':       home_games,
                'Heim %':     f'{home_pct}%',
                'Wechsel':    sw,
                'Quote':      f'{rate:.0f}%',
            })
        name_f = cfg_f.name
        with st.expander(f'Fairness: {name_f}', expanded=True):
            st.dataframe(pd.DataFrame(rows_f), hide_index=True, width='stretch')
            st.caption(
                f'Ø {avg_km:.0f} km/Team · '
                f'Ø Wechselquote {sum(res.sw_rates)/n_f:.0f}% · '
                f'Reisefairness (max. Abw.): '
                f'{(max(abs(res.travels[i]-avg_km)/avg_km*100 for i in range(n_f)) if avg_km else 0):.0f}%'
            )

    # Spielpläne
    st.subheader('Spielpläne')
    for lid, res in S.results.items():
        if res is None:
            st.warning(
                f'**{lid}: Kein Spielplan gefunden.** '
                f'Mögliche Ursachen: zu viele Pflichtspiele oder Sperrtage auf denselben Spieltagen, '
                f'DST-Blöcke decken alle Spieltage ab, oder das Zeitlimit war zu kurz. '
                f'Empfehlung: Rechenzeit in Schritt 8 erhöhen, Pflichtspiele reduzieren, '
                f'oder Sperrtage prüfen und neu optimieren.'
            )
            continue
        n_rounds    = res.cfg.n_rounds    if res.cfg else 2
        n_per_round = max(1, res.cfg.n_matchdays // n_rounds) if res.cfg else 1
        dst_days    = res.cfg.dst_days    if res.cfg else set()
        phase_lbl   = {1: 'Hin', 2: 'Rue'} if n_rounds == 2 else {}
        is_tt       = bool(res.cfg and res.cfg.games_per_team_per_day > 1)
        typ_col     = 'Ausrichter' if is_tt else 'Typ'
        rows = []
        for d in (res.cfg.days if res.cfg else sorted(res.schedule.keys())):
            rnd   = min(n_rounds, (d - 1) // n_per_round + 1)
            phase = phase_lbl.get(rnd, f'R{rnd}')
            typ   = res.hosts.get(d, '') if is_tt else ('DST' if d in dst_days else 'EST')
            for ht, at in res.schedule.get(d, []):
                rows.append({'ST': d, 'Phase': phase, typ_col: typ,
                             'Heimteam': ht, 'Gastteam': at})
        name = res.cfg.name if res.cfg else lid
        with st.expander(f'📅 {name}  ({len(rows)} Spiele)', expanded=False):
            st.dataframe(pd.DataFrame(rows), width='stretch', hide_index=True)

    # Downloads
    st.subheader('Excel-Dateien')
    _n_dl = (len(S.excel_bytes)
             + (1 if S.cohome_bytes else 0)
             + (1 if S.hall_bytes else 0))
    dl_cols = st.columns(max(1, min(4, _n_dl)))
    idx = 0
    for lid, res in S.results.items():
        if lid not in S.excel_bytes or res is None:
            continue
        name = res.cfg.name if res.cfg else lid
        with dl_cols[idx % len(dl_cols)]:
            st.download_button(
                f'⬇ {name}',
                data=S.excel_bytes[lid],
                file_name=f'Spielplan_{lid}.xlsx',
                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                width='stretch',
                key=f'dl_{lid}',
            )
        idx += 1
    if S.cohome_bytes:
        with dl_cols[idx % len(dl_cols)]:
            st.download_button(
                '⬇ Co-Home-Übersicht',
                data=S.cohome_bytes,
                file_name='CoHome_Uebersicht.xlsx',
                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                width='stretch',
                key='dl_cohome',
            )
        idx += 1
    if S.hall_bytes:
        with dl_cols[idx % len(dl_cols)]:
            st.download_button(
                '⬇ Hallenbelegungsplan',
                data=S.hall_bytes,
                file_name='Hallenbelegungsplan.xlsx',
                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                width='stretch',
                key='dl_hall',
            )

    # ── Kalender-Export (.ics) ────────────────────────────────────────────────
    _valid_res = [(lid, res) for lid, res in S.results.items() if res is not None]
    if _valid_res:
        st.subheader('Kalender-Export (.ics)')
        _has_cal = any(res.cfg and res.cfg.calendar for _, res in _valid_res)
        if not _has_cal:
            st.caption(
                'Kein Rahmenterminplan geladen – die .ics-Datei enthält Spieltag-Nummern ohne '
                'konkrete Daten. Für datierte Kalender-Einträge bitte in Schritt 3 einen '
                'Rahmenterminplan hochladen und laden.'
            )
        _ics_year = st.number_input(
            'Saison-Startjahr', 2020, 2035, 2026, 1, key='ics_year',
            help=(
                'Jahr des Saisonstarts (z. B. 2026 für Saison 2026/27). '
                'Kalenderwochen > 26 werden diesem Jahr, Wochen ≤ 26 dem Folgejahr zugeordnet.'
            ),
        )
        _ics_dl_cols = st.columns(max(1, min(4, len(_valid_res))))
        for _ci, (_lid, _res) in enumerate(_valid_res):
            with _ics_dl_cols[_ci % len(_ics_dl_cols)]:
                st.download_button(
                    f'📅 {_res.cfg.name if _res.cfg else _lid}',
                    data=_build_ics_bytes(_res, int(_ics_year)),
                    file_name=f'Spielplan_{_lid}.ics',
                    mime='text/calendar',
                    width='stretch',
                    key=f'dl_ics_{_lid}',
                )

    # ── Druckbarer Spielplan (HTML) ───────────────────────────────────────────
    if _valid_res:
        st.subheader('Druckbarer Spielplan')
        st.caption(
            'HTML-Datei herunterladen, im Browser öffnen und mit Strg+P / Cmd+P drucken '
            'oder als PDF speichern. Enthält Gesamtspielplan und Einzelansicht je Team.'
        )
        _print_year = int(st.session_state.get('ics_year', 2026))
        _print_cols = st.columns(max(1, min(4, len(_valid_res))))
        for _ci, (_lid, _res) in enumerate(_valid_res):
            if not _res.cfg:
                continue
            with _print_cols[_ci % len(_print_cols)]:
                _html_bytes = _build_print_html_fn(_res, _print_year).encode('utf-8')
                st.download_button(
                    f'🖨 {_res.cfg.name}',
                    data=_html_bytes,
                    file_name=f'Spielplan_{_lid}_druckbar.html',
                    mime='text/html',
                    width='stretch',
                    key=f'dl_print_{_lid}',
                )

    # ── Spielzeiten zuweisen ─────────────────────────────────────────────────
    with st.expander('Spielzeiten zuweisen', expanded=False):
        st.caption(
            'Uhrzeiten für die Spiele eines Spieltags festlegen. '
            'Die Zeiten werden in Excel und im druckbaren Spielplan angezeigt. '
            'Format: "14:00, 16:00, 18:00" (eine Uhrzeit je Spiel, durch Komma getrennt).'
        )
        _time_lids = [l for l, r in S.results.items() if r is not None]
        for _tl in _time_lids:
            _res_t  = S.results[_tl]
            _name_t = _res_t.cfg.name if _res_t.cfg else _tl
            _gpd    = _res_t.cfg.games_per_team_per_day if _res_t.cfg else 1
            _n_t    = _res_t.cfg.n_teams if _res_t.cfg else 2
            _n_active = _res_t.cfg.n_active_per_day if _res_t.cfg and _res_t.cfg.n_active_per_day > 0 else _n_t
            _gspd   = _n_active * _gpd // 2     # Spiele pro Spieltag
            _default_slots = ', '.join(
                f'{(14 + i * 2):02d}:00' for i in range(_gspd)
            )
            _tpl = st.text_input(
                f'Uhrzeiten für **{_name_t}** ({_gspd} Spiel{"e" if _gspd != 1 else ""}/Tag)',
                value=S.time_templates.get(_tl, _default_slots),
                placeholder='z. B. 14:00, 16:00, 18:00',
                key=f'time_tpl_{_tl}',
            )
            S.time_templates[_tl] = _tpl

        if st.button('Spielzeiten automatisch zuweisen', key='assign_times',
                     type='secondary'):
            _changed = False
            for _tl in _time_lids:
                _res_t = S.results[_tl]
                _slots = [s.strip() for s in S.time_templates.get(_tl, '').split(',')
                          if s.strip()]
                if _slots:
                    _assign_game_times_fn(_res_t, _slots)
                    # Excel-Bytes neu erzeugen
                    from spielplan_multi.excel_output import build_league_excel
                    _wb  = build_league_excel(_res_t)
                    _buf = io.BytesIO()
                    _wb.save(_buf)
                    S.excel_bytes[_tl] = _buf.getvalue()
                    _changed = True
            if _changed:
                from spielplan_multi.excel_output import build_hall_schedule as _bhs
                _wb_hall  = _bhs(S.results)
                _buf_hall = io.BytesIO()
                _wb_hall.save(_buf_hall)
                S.hall_bytes = _buf_hall.getvalue()
                st.success('Spielzeiten zugewiesen. Excel-Dateien aktualisiert.')
                st.rerun()

        # Vorschau der aktuellen Zuteilung
        _any_times = any(
            bool(S.results[l].game_times) for l in _time_lids if S.results[l] is not None
        )
        if _any_times:
            for _tl in _time_lids:
                _res_t = S.results.get(_tl)
                if _res_t and _res_t.game_times and _res_t.cfg:
                    _prev_days = sorted(_res_t.game_times.keys())[:3]
                    _prev_rows = []
                    for _pd in _prev_days:
                        for _gi, (_ht, _at) in enumerate(_res_t.schedule.get(_pd, [])):
                            _uhr = _res_t.game_times[_pd][_gi] if _gi < len(_res_t.game_times[_pd]) else ''
                            _prev_rows.append({'ST': _pd, 'Uhrzeit': _uhr,
                                               'Heimteam': _ht, 'Gastteam': _at})
                    if _prev_rows:
                        st.caption(f'Vorschau {_res_t.cfg.name} (erste 3 Spieltage):')
                        st.dataframe(pd.DataFrame(_prev_rows), hide_index=True)

    # ── Spielplan manuell anpassen ────────────────────────────────────────────
    with st.expander('Spielplan manuell anpassen', expanded=False):
        st.caption(
            'Einzelne Heimrecht-Entscheidungen nach der Optimierung korrigieren. '
            'Bei Doppelspieltagen werden beide Spieltage gemeinsam getauscht. '
            'Reisekilometer und Wechselquoten werden automatisch neu berechnet.'
        )
        _adj_lids = [l for l, r in S.results.items() if r is not None]
        if not _adj_lids:
            st.info('Keine Ergebnisse zum Anpassen verfügbar.')
        else:
            _adj_lid = (
                st.selectbox(
                    'Liga', _adj_lids,
                    format_func=lambda l: (S.results[l].cfg.name
                                           if S.results.get(l) and S.results[l].cfg else l),
                    key='adj_lid',
                )
                if len(_adj_lids) > 1 else _adj_lids[0]
            )
            _res_adj = S.results.get(_adj_lid)
            if _res_adj and _res_adj.cfg:
                _cfg_adj  = _res_adj.cfg
                _all_days = sorted(_res_adj.schedule.keys())
                _sel_day  = st.selectbox(
                    'Spieltag', _all_days,
                    format_func=lambda d: (
                        f'Spieltag {d} (Doppelspieltag)' if d in _cfg_adj.dst_days
                        else f'Spieltag {d}'
                    ),
                    key='adj_day',
                )
                # DST-Hinweis
                for _d1, _d2 in _cfg_adj.dst_blocks:
                    if _sel_day in (_d1, _d2):
                        _partner = _d2 if _sel_day == _d1 else _d1
                        st.info(
                            f'Doppelspieltag: Ein Tausch hier gilt automatisch auch '
                            f'für Spieltag {_partner}.'
                        )
                        break
                _day_games = _res_adj.schedule.get(_sel_day, [])
                if not _day_games:
                    st.caption('Keine Spiele an diesem Spieltag.')
                else:
                    for _mi, (_ht, _at) in enumerate(_day_games):
                        _c1, _c2, _c3 = st.columns([4, 4, 1])
                        _c1.write(f'**{_ht}** (Heim)')
                        _c2.write(f'{_at} (Gast)')
                        if _c3.button(
                            '⇄', key=f'swap_{_adj_lid}_{_sel_day}_{_mi}',
                            help=f'Tauschen: {_at} spielt zu Hause gegen {_ht}',
                        ):
                            _swap_home_away(_res_adj, _cfg_adj, _sel_day, _mi)
                            from spielplan_multi.excel_output import build_league_excel as _ble
                            _wb2 = _ble(_res_adj)
                            _buf2 = io.BytesIO()
                            _wb2.save(_buf2)
                            S.excel_bytes[_adj_lid] = _buf2.getvalue()
                            st.rerun()


    # ── Spiel verschieben / Absagen & Nachholspiele ───────────────────────────
    with st.expander('Spiel verschieben / Absagen & Nachholspiele', expanded=False):
        st.caption(
            'Einzelne Spiele nachträglich anpassen: auf einen anderen Spieltag verschieben (📅) '
            'oder als ausgefallen markieren und optional einen Nachholtermin einplanen (❌). '
            'Der Optimierer prüft, ob der Zieltermin für beide Teams frei ist. '
            'Reisekilometer und Wechselquoten werden nach jeder Änderung automatisch neu berechnet. '
            'Die aktualisierte Excel-Datei wird sofort neu erzeugt.'
        )
        _mv_lids = [l for l, r in S.results.items() if r is not None]
        if not _mv_lids:
            st.info('Keine Ergebnisse verfügbar.')
        else:
            _mv_lid = (
                st.selectbox(
                    'Liga', _mv_lids,
                    format_func=lambda l: (S.results[l].cfg.name
                                           if S.results.get(l) and S.results[l].cfg else l),
                    key='mv_lid',
                )
                if len(_mv_lids) > 1 else _mv_lids[0]
            )
            # Pending-States für andere Ligen löschen
            _mp = S.move_pending
            _cp = S.cancel_pending
            if _mp and _mp.get('lid') != _mv_lid:
                S.move_pending = None; _mp = None
            if _cp and _cp.get('lid') != _mv_lid:
                S.cancel_pending = None; _cp = None

            _res_mv = S.results.get(_mv_lid)
            if _res_mv and _res_mv.cfg:
                _cfg_mv  = _res_mv.cfg
                _mv_days = sorted(_res_mv.schedule.keys())
                _sel_day_mv = st.selectbox(
                    'Spieltag', _mv_days,
                    format_func=lambda d: f'Spieltag {d}',
                    key='mv_day',
                )
                _mv_games = _res_mv.schedule.get(_sel_day_mv, [])

                # ── Pending: Nachholtermin wählen ─────────────────────────────
                if _cp and _cp.get('lid') == _mv_lid:
                    st.info(
                        f'**{_cp["ht"]} vs. {_cp["at"]}** ist ausgefallen. '
                        f'Nachholtermin wählen oder ohne Ersatz bestätigen.'
                    )
                    _free_cp = _find_free_days(_res_mv, _cfg_mv, _cp['ht'], _cp['at'])
                    if _free_cp:
                        _nhl_day = st.selectbox(
                            'Nachholtermin',
                            _free_cp,
                            format_func=lambda d: f'Spieltag {d}',
                            key='mv_nachhol_day',
                        )
                        _nc1, _nc2 = st.columns(2)
                        if _nc1.button('Nachholspiel einplanen', type='primary',
                                       key='btn_nachhol_ok'):
                            _err_nhl = _reschedule_game(
                                _res_mv, _cfg_mv, _nhl_day, _cp['ht'], _cp['at']
                            )
                            if _err_nhl:
                                st.error(_err_nhl)
                            else:
                                S.cancel_pending = None
                                _regen_league_excel(_mv_lid, _res_mv)
                                st.success(
                                    f'Nachholspiel {_cp["ht"]} vs. {_cp["at"]} '
                                    f'an Spieltag {_nhl_day} eingetragen.'
                                )
                                st.rerun()
                        if _nc2.button('Ohne Nachholtermin bestätigen', key='btn_nachhol_skip'):
                            S.cancel_pending = None
                            st.rerun()
                    else:
                        st.warning('Kein freier Spieltag für beide Teams gefunden.')
                        if st.button('Ohne Nachholtermin bestätigen', key='btn_nachhol_nf'):
                            S.cancel_pending = None
                            st.rerun()
                    st.markdown('---')

                # ── Pending: Zieltag für Verschieben wählen ───────────────────
                if _mp and _mp.get('lid') == _mv_lid:
                    st.markdown(
                        f'**{_mp["ht"]} vs. {_mp["at"]}** '
                        f'(Spieltag {_mp["day"]}) verschieben auf:'
                    )
                    _free_mv = [
                        d for d in _find_free_days(_res_mv, _cfg_mv, _mp['ht'], _mp['at'])
                        if d != _mp['day']
                    ]
                    if _free_mv:
                        _target_day = st.selectbox(
                            'Ziel-Spieltag',
                            _free_mv,
                            format_func=lambda d: f'Spieltag {d}',
                            key='mv_target_sel',
                        )
                        _mc1, _mc2 = st.columns(2)
                        if _mc1.button('Verschieben bestätigen', type='primary',
                                       key='btn_mv_ok'):
                            _err_mv = _move_game(
                                _res_mv, _cfg_mv, _mp['day'], _mp['idx'], _target_day
                            )
                            if _err_mv:
                                st.error(_err_mv)
                            else:
                                S.move_pending = None
                                _regen_league_excel(_mv_lid, _res_mv)
                                st.success(
                                    f'Spiel auf Spieltag {_target_day} verschoben.'
                                )
                                st.rerun()
                        if _mc2.button('Abbrechen', key='btn_mv_abort'):
                            S.move_pending = None
                            st.rerun()
                    else:
                        st.warning('Kein freier Spieltag für beide Teams gefunden.')
                        if st.button('Abbrechen', key='btn_mv_abort_nf'):
                            S.move_pending = None
                            st.rerun()
                    st.markdown('---')

                # ── Spiel-Liste mit Aktions-Buttons ───────────────────────────
                if not _mv_games:
                    st.caption('Keine Spiele an diesem Spieltag.')
                else:
                    for _mvi, (_mht, _mat) in enumerate(_mv_games):
                        _gc1, _gc2, _gc3, _gc4 = st.columns([4, 4, 1, 1])
                        _gc1.write(f'**{_mht}** (Heim)')
                        _gc2.write(f'{_mat} (Gast)')
                        if _gc3.button(
                            '📅', key=f'mv_btn_{_mv_lid}_{_sel_day_mv}_{_mvi}',
                            help='Auf anderen Spieltag verschieben',
                        ):
                            S.move_pending = {
                                'lid': _mv_lid, 'day': _sel_day_mv,
                                'idx': _mvi, 'ht': _mht, 'at': _mat,
                            }
                            S.cancel_pending = None
                            st.rerun()
                        if _gc4.button(
                            '❌', key=f'cancel_btn_{_mv_lid}_{_sel_day_mv}_{_mvi}',
                            help='Als ausgefallen markieren',
                        ):
                            _ht_c, _at_c = _cancel_game(
                                _res_mv, _cfg_mv, _sel_day_mv, _mvi
                            )
                            if _ht_c:
                                S.cancel_pending = {
                                    'lid': _mv_lid, 'ht': _ht_c, 'at': _at_c,
                                }
                                S.move_pending = None
                                _regen_league_excel(_mv_lid, _res_mv)
                                st.rerun()
                            else:
                                st.error('Spiel konnte nicht entfernt werden (Spieltag oder Index ungültig).')

    # ── Spielplan vergleichen ─────────────────────────────────────────────────
    with st.expander('Spielplan vergleichen', expanded=False):
        st.caption(
            'Ergebnis-Excel eines früheren Optimierungslaufs hochladen, um Kilometer '
            'und Wechselquoten direkt nebeneinander zu vergleichen. '
            'Die Vergleichsdatei muss ein Sheet namens **„Kilometerstatistik"** enthalten – '
            'dieses Sheet wird bei jedem Excel-Export automatisch erzeugt. '
            'Nützlich, um verschiedene Konfigurationen gegeneinander abzuwägen.'
        )
        _cmp_valid = [(l, r) for l, r in S.results.items()
                      if r is not None and r.cfg]
        if not _cmp_valid:
            st.info('Keine Ergebnisse vorhanden.')
        else:
            for _cl, _cr in _cmp_valid:
                _cmp_up = st.file_uploader(
                    f'Vergleichsdatei für **{_cr.cfg.name}**',
                    type=['xlsx', 'xls'],
                    key=f'cmp_upload_{_cl}',
                )
                if not _cmp_up:
                    continue
                try:
                    _cmp_xl = pd.ExcelFile(io.BytesIO(_cmp_up.getvalue()))
                    if 'Kilometerstatistik' not in _cmp_xl.sheet_names:
                        st.error(f'Sheet "Kilometerstatistik" nicht in der Datei gefunden.')
                        continue
                    _cmp_df = _cmp_xl.parse('Kilometerstatistik').fillna('')
                    # Lookup: team → (km, sw_rate)
                    _cmp_lookup: dict = {}
                    for _, _row in _cmp_df.iterrows():
                        _tm = str(_row.iloc[0]).strip()
                        if not _tm or _tm.lower() in ('team', 'gesamt', 'durchschnitt', 'nan'):
                            continue
                        try:
                            _cmp_lookup[_tm] = (
                                int(float(str(_row.iloc[1]))),
                                float(str(_row.iloc[3])),
                            )
                        except (ValueError, IndexError):
                            pass

                    _rows_cmp = []
                    for _ti, _t in enumerate(_cr.cfg.teams):
                        _cur_km = _cr.travels[_ti]
                        _cur_sw = round(_cr.sw_rates[_ti], 1)
                        _prev   = _cmp_lookup.get(_t)
                        if _prev:
                            _prev_km, _prev_sw = _prev
                            _dkm = _cur_km - _prev_km
                            _dsw = round(_cur_sw - _prev_sw, 1)
                            _rows_cmp.append({
                                'Team':          _t,
                                'Aktuell km':    _cur_km,
                                'Vorher km':     _prev_km,
                                'Δ km':          f'{_dkm:+d}',
                                'Aktuell SW%':   _cur_sw,
                                'Vorher SW%':    _prev_sw,
                                'Δ SW%':         f'{_dsw:+.1f}',
                            })
                        else:
                            _rows_cmp.append({
                                'Team':          _t,
                                'Aktuell km':    _cur_km,
                                'Vorher km':     '—',
                                'Δ km':          '—',
                                'Aktuell SW%':   _cur_sw,
                                'Vorher SW%':    '—',
                                'Δ SW%':         '—',
                            })
                    if _rows_cmp:
                        st.markdown(f'**{_cr.cfg.name}**')
                        _cmp_frame = pd.DataFrame(_rows_cmp)
                        st.dataframe(_cmp_frame, hide_index=True, width='stretch')
                        _cur_tot = sum(_cr.travels)
                        _prev_total_km = sum(
                            _cmp_lookup[t][0] for t in _cr.cfg.teams if t in _cmp_lookup
                        )
                        if _prev_total_km:
                            _d_total = _cur_tot - _prev_total_km
                            st.caption(
                                f'Gesamt: aktuell **{_cur_tot:,} km** · '
                                f'vorher {_prev_total_km:,} km · '
                                f'Differenz **{_d_total:+,} km**'
                            )
                except zipfile.BadZipFile:
                    st.error('Die Vergleichsdatei ist beschädigt oder kein gültiges Excel-Format (.xlsx).')
                except ValueError as _ve:
                    st.error(f'Sheet nicht gefunden oder ungültiges Format: {_ve}')
                except Exception as _ce:
                    st.error(f'Fehler beim Lesen der Vergleichsdatei: {_ce}')


# ═══════════════════════════════════════════════════════════════════════════════
# Übersichtsseite
# ═══════════════════════════════════════════════════════════════════════════════
def _step_intro():
    st.image(_LOGO_PATH, width=420)
    st.title('Spielplan-Optimierer')
    st.markdown('### Automatische Spielplanerstellung für Floorball-Ligen')
    st.markdown('---')

    st.markdown(
        'Der Spielplan-Optimierer von **FLOORBALL VERBAND DEUTSCHLAND (FD)** erstellt '
        'automatisch Spielpläne für eine oder mehrere Ligen gleichzeitig. '
        'Ziel ist ein fairer, reisearmer Spielplan, der alle sportlichen und '
        'organisatorischen Vorgaben berücksichtigt.'
    )

    st.markdown('---')

    # ── Was kann der Optimierer? ──────────────────────────────────────────────
    col_a, col_b = st.columns(2)

    with col_a:
        st.subheader('Was der Optimierer kann')
        st.markdown('''
**Spielplan-Formate**
- Hin-/Rückrunde (jede Paarung 2×)
- Einfachrunde (jede Paarung 1×)
- Dreifachrunde (jede Paarung 3×)
- Turniertag: alle Teams an einem Ort, mehrere Spiele pro Team pro Tag –
  mit optionaler Gruppenaufteilung (z. B. 2 Gruppen à 4 Teams)

**Optimierungsziele**
- Reisekilometer aller Teams minimieren
- Heimrecht möglichst oft wechseln (abwechselnd Heim/Auswärts)
- Faire Verteilung von Heim- und Reisekilometern über alle Teams

**Terminplanung**
- Mehrere Ligen gleichzeitig optimieren
- Doppelspieltage (DST) mit konsistentem Heimrecht
- Vereine mit Teams in mehreren Ligen koordinieren (gleiche Heimspielwochen)
- Pflichtspiele auf bestimmte Spieltage festlegen
- Heimspiel-Sperrtage pro Team (z. B. Hallensperrung, kein Heimspiel an Spieltag X)
- Heimspiel-Pflichttage pro Team (z. B. Halle nur an bestimmten Spieltagen verfügbar)
- Rahmenterminplan aus Excel einlesen (Kalenderwochen-Zuordnung)

**Distanzberechnung**
- Automatisch per Google Maps API
- Manuell über Excel-/CSV-Datei
- Direkte Eingabe im Tool

**Nach der Optimierung**
- Einzelne Spiele manuell verschieben
- Spiele als ausgefallen markieren und Nachholtermin einplanen
- Spielzeiten (Anstoßzeiten) je Spieltag zuweisen
- Spielplan-Qualität prüfen (automatische Warnungen)
- Zwei Läufe direkt nebeneinander vergleichen
- Sitzung speichern und jederzeit wieder laden – inkl. aller Bearbeitungsoptionen
- Excel, iCal-Kalender und druckbare HTML-Ansicht herunterladen
        ''')

    with col_b:
        st.subheader('Wie die Optimierung funktioniert')
        st.markdown('''
Der Optimierer arbeitet in **3 Schritten**:

**Schritt 1 – Spielplan je Liga erstellen**
Für jede Liga wird unabhängig ein vollständiger Spielplan berechnet.
Alle Ligen laufen dabei gleichzeitig – egal ob 2 oder 5 Ligen,
die Wartezeit bleibt ähnlich. Mehrere Versuche mit verschiedenen
Ausgangspunkten erhöhen die Qualität des Ergebnisses.

**Schritt 2 – Ligen aufeinander abstimmen**
Alle Ligen werden gemeinsam betrachtet. Spieltage werden auf
Kalenderwochen verteilt. Hauptziel: Vereine mit Teams in mehreren
Ligen sollen ihre Heimspiele möglichst in denselben Wochen haben,
um Halle und Personal zu bündeln.

**Schritt 3 – Feinabstimmung der Heimspiele**
Ein ergänzender Durchlauf prüft, ob einzelne Hin-/Rückspiele
getauscht werden sollten, um Fahrtwege zu verkürzen.
Der Terminplan selbst bleibt dabei unverändert.
Erfahrungsgemäß werden dabei 3–8 % der Gesamtkilometer eingespart.

---

**Ergebnis**
Je Liga wird eine Excel-Datei erstellt mit:
- Vollständigem Spielplan (alle Spieltage, Heim/Auswärts)
- Heimrecht-Heatmap (Übersicht pro Team und Spieltag)
- Kilometerstatistik und Fahrtkostenausgleich
- Distanzmatrix aller Standorte

Außerdem je Liga verfügbar:
- **iCal-Export** (.ics) zum Import in Google Calendar, Outlook usw.
- **Druckbare HTML-Ansicht** (PDF-Export über Browser)
        ''')

    st.markdown('---')

    # ── Grenzen ───────────────────────────────────────────────────────────────
    st.subheader('Grenzen & Hinweise')
    lim_a, lim_b = st.columns(2)
    with lim_a:
        st.markdown('''
**Was der Optimierer nicht übernimmt**
- Mindestens **4 Teams** pro Liga erforderlich
- Konkrete Spieltermine (Uhrzeit, genaues Datum) werden **nicht** vergeben –
  der Spielplan weist Spieltage Kalenderwochen zu; genaue Ansetzungen
  macht der zuständige Staffelleiter
- Schiedsrichter-Ansetzungen sind nicht Teil der Optimierung

**Speichern & Wiederherstellen**
- Konfiguration: als Excel-Datei jederzeit speichern und wieder laden
- Ergebnisse & Spielpläne: als Sitzungs-Datei (.json) speichern –
  über die Seitenleiste jederzeit wieder laden, um Spielzeiten zuzuweisen,
  Spiele zu verschieben oder Absagen zu verwalten
- Distanzmatrix muss einmalig bereitgestellt werden
  (Google Maps API-Schlüssel oder manuelle Eingabe)
        ''')
    with lim_b:
        st.markdown('''
**Laufzeit & Lösungsqualität**
- Die Optimierung dauert je nach Ligagröße und Einstellungen
  **mehrere Stunden** – ein Nachtlauf wird für mehrere Ligen empfohlen
- Das Ergebnis ist ein sehr guter, aber nicht zwingend der mathematisch
  optimale Plan – das wäre bei dieser Problemgröße nicht in vertretbarer
  Zeit berechenbar
- Vor dem Start prüft der Optimierer die Konfiguration automatisch auf
  Widersprüche und gibt verständliche Fehlermeldungen

**„Keine Lösung gefunden"**
- Tritt auf, wenn Vorgaben widersprüchlich sind (z. B. zu viele Sperr-
  oder Pflichttage, zu viele Pflichtspiele) oder das Zeitlimit erreicht wird
- Der Optimierer zeigt dann eine **automatische Diagnose** mit den
  wahrscheinlichen Ursachen und konkreten Verbesserungsvorschlägen

**Turniertag**
- Beim Turniertag-Format mit Gruppen (Stufe 2) werden die Gruppen je
  Spieltag automatisch neu ausgelost – wer gegen wen in welcher Gruppe
  spielt, ergibt sich aus der Optimierung
- Spielreihenfolge innerhalb eines Turniertags wird konfiguriert,
  aber nicht weiter optimiert
        ''')

    st.markdown('---')

    # ── Start-Button ──────────────────────────────────────────────────────────
    _, btn_col, _ = st.columns([2, 3, 2])
    with btn_col:
        if st.button('Konfiguration starten →', type='primary',
                     width='stretch', key='intro_start'):
            S._wizard_started = True
            st.rerun()


# ═══════════════════════════════════════════════════════════════════════════════
# Haupt-Rendering
# ═══════════════════════════════════════════════════════════════════════════════
_sidebar()

if not S._wizard_started:
    _step_intro()
else:
    [_step0, _step1, _step2, _step3, _step4, _step5, _step6, _step7, _step8][max(0, min(S.step, 8))]()
